"""
AIO LBO Excel Generator

Generates Excel workbooks for LBO models from validated SEC EDGAR data.
Currently implements: Assumptions tab, Sources & Uses tab.
"""

import os
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.comments import Comment


# =============================================================================
# STYLE CONSTANTS - IB/PE Modeling Convention
# =============================================================================

# Blue font: hardcoded inputs (API-fetched data + user assumptions)
BLUE_FONT = Font(name='Calibri', size=11, color='0000CC', bold=False)

# Black font: formulas (calculated from other cells)
BLACK_FONT = Font(name='Calibri', size=11, color='000000', bold=False)

# Bold fonts for headers
BLUE_FONT_BOLD = Font(name='Calibri', size=11, color='0000CC', bold=True)
BLACK_FONT_BOLD = Font(name='Calibri', size=11, color='000000', bold=True)

# Section header styling
SECTION_HEADER_FONT = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
SECTION_HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

# Sub-section header styling
SUBSECTION_FONT = Font(name='Calibri', size=11, bold=True, color='000000')
SUBSECTION_FILL = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')

# Defaulted value flag (soft requirement defaulted by validator) - Yellow
DEFAULTED_FILL = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')

# User-provided value flag (manually entered by user to fill missing data) - Light Green
USER_PROVIDED_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

# Error fill (for balance check)
ERROR_FILL = PatternFill(start_color='FF6666', end_color='FF6666', fill_type='solid')

# Implausible calculation flag (e.g., COVID-distorted growth rate)
# Orange fill - distinct from yellow "defaulted" and matches Interest Expense placeholder
IMPLAUSIBLE_CALC_FILL = PatternFill(start_color='FFB366', end_color='FFB366', fill_type='solid')

# Borders
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Alignments
RIGHT_ALIGN = Alignment(horizontal='right')
LEFT_ALIGN = Alignment(horizontal='left')
CENTER_ALIGN = Alignment(horizontal='center')

# Number formats
CURRENCY_FORMAT = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'
CURRENCY_MILLIONS_FORMAT = '_($* #,##0.0,,"M"_);_($* (#,##0.0,,"M");_($* "-"??_);_(@_)'
PERCENT_FORMAT = '0.0%'
MULTIPLE_FORMAT = '0.0"x"'
NUMBER_FORMAT = '#,##0'
SHARES_FORMAT = '#,##0.0,,"M"'  # Shares in millions


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def safe_get(data: dict, *keys, default=None):
    """Safely navigate nested dict structure."""
    current = data
    for key in keys:
        if isinstance(current, dict) and key in current:
            current = current[key]
        else:
            return default
    return current if current is not None else default


def get_most_recent_fy(summary: dict) -> Optional[int]:
    """Get the most recent fiscal year from summary."""
    fiscal_years = summary.get('fiscal_years', {})
    if not fiscal_years:
        return None
    return max(fiscal_years.keys())


def get_fy_data(summary: dict, fy: int) -> dict:
    """Get data for a specific fiscal year."""
    return summary.get('fiscal_years', {}).get(fy, {})


def calculate_revenue_growth_rate(summary: dict) -> tuple[float, bool, str]:
    """
    Calculate revenue growth rate from 5-year history using MEDIAN of YoY changes.

    Uses median instead of mean to reduce sensitivity to extreme outlier years
    (e.g., COVID-era revenue collapses/recoveries).

    Applies a sanity band: if calculated rate is outside -15% to +25%,
    falls back to a conservative 3% default.

    Returns:
        (growth_rate, is_fallback, comment)
        - growth_rate: the rate to use (either calculated median or 3% fallback)
        - is_fallback: True if the sanity band was triggered and 3% fallback was used
        - comment: description of how the value was derived
    """
    import statistics

    fiscal_years = summary.get('fiscal_years', {})
    if len(fiscal_years) < 2:
        return 0.03, True, "Insufficient historical data — defaulted to 3%"

    # Sort years in ascending order
    sorted_years = sorted(fiscal_years.keys())
    growth_rates = []

    for i in range(1, len(sorted_years)):
        prev_year = sorted_years[i - 1]
        curr_year = sorted_years[i]
        prev_rev = fiscal_years.get(prev_year, {}).get('revenue')
        curr_rev = fiscal_years.get(curr_year, {}).get('revenue')

        if prev_rev and curr_rev and prev_rev > 0:
            growth = (curr_rev - prev_rev) / prev_rev
            growth_rates.append(growth)

    if not growth_rates:
        return 0.03, True, "No valid revenue history — defaulted to 3%"

    # Use MEDIAN instead of mean to reduce outlier sensitivity
    median_growth = statistics.median(growth_rates)

    # Sanity band: -15% to +25%
    GROWTH_MIN = -0.15
    GROWTH_MAX = 0.25
    FALLBACK_RATE = 0.03

    if median_growth < GROWTH_MIN or median_growth > GROWTH_MAX:
        return (
            FALLBACK_RATE,
            True,
            f"Historical growth rate ({median_growth*100:.1f}%) was implausible "
            f"(likely COVID-era distortion) — defaulted to 3%, please review and adjust manually."
        )

    return (
        median_growth,
        False,
        f"Default: median of historical YoY growth rates ({median_growth*100:.1f}%)"
    )


def is_user_provided_field(summary: dict, field_name: str) -> bool:
    """Check if a field was user-provided (manually entered by user to fill missing data)."""
    user_provided_fields = summary.get('user_provided_fields', [])
    return field_name in user_provided_fields


def is_defaulted_field(summary: dict, validation_result: dict, field_name: str) -> bool:
    """Check if a field was defaulted by the validator.
    Returns False if the field is user-provided (user data takes precedence).
    """
    # User-provided fields are NOT defaulted
    if is_user_provided_field(summary, field_name):
        return False

    defaults = validation_result.get('defaults_applied', [])

    # Check if field is mentioned in defaults
    for default_msg in defaults:
        if field_name.lower() in default_msg.lower():
            return True

    # Also check if the field is None in the most recent FY
    most_recent_fy = get_most_recent_fy(summary)
    if most_recent_fy:
        fy_data = get_fy_data(summary, most_recent_fy)
        if fy_data.get(field_name) is None:
            return True

    return False


def create_defined_name(wb: Workbook, name: str, sheet_name: str, cell: str):
    """Create a workbook-level defined name for a cell."""
    from openpyxl.workbook.defined_name import DefinedName

    # Clean the name to be Excel-compatible
    clean_name = name.replace(' ', '_').replace('%', 'Pct').replace('/', '_')
    clean_name = ''.join(c for c in clean_name if c.isalnum() or c == '_')

    # Ensure name doesn't start with a number
    if clean_name[0].isdigit():
        clean_name = '_' + clean_name

    # Build the cell reference - parse cell like "C15" to "$C$15"
    col_letter = ''.join(c for c in cell if c.isalpha())
    row_num = ''.join(c for c in cell if c.isdigit())
    cell_ref = f"${col_letter}${row_num}"

    # Create and add the defined name
    defn = DefinedName(clean_name, attr_text=f"'{sheet_name}'!{cell_ref}")
    wb.defined_names[clean_name] = defn

    return clean_name


# =============================================================================
# ASSUMPTIONS TAB BUILDER
# =============================================================================

def build_assumptions_tab(wb: Workbook, summary: dict, validation_result: dict) -> dict:
    """
    Build the Assumptions tab.

    Returns a dict mapping friendly names to Excel cell references for use by other tabs.
    """
    ws = wb.create_sheet("Assumptions", 0)

    # Column widths
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 38  # Wide enough for longest labels
    ws.column_dimensions['C'].width = 22  # Wide enough for $X,XXX,XXX,XXX,XXX format
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['F'].width = 22
    ws.column_dimensions['G'].width = 22
    ws.column_dimensions['H'].width = 22

    # Track named ranges
    named_ranges = {}
    current_row = 1

    # ==========================================================================
    # SECTION A: COMPANY DATA
    # ==========================================================================

    # Section header
    ws.cell(row=current_row, column=2, value="SECTION A: COMPANY DATA (Fetched from SEC EDGAR)")
    ws.cell(row=current_row, column=2).font = SECTION_HEADER_FONT
    ws.cell(row=current_row, column=2).fill = SECTION_HEADER_FILL
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=7)
    current_row += 2

    # Company identification
    ws.cell(row=current_row, column=2, value="Company Name").font = BLUE_FONT
    ws.cell(row=current_row, column=3, value=summary.get('company_name', 'Unknown')).font = BLUE_FONT
    current_row += 1

    ws.cell(row=current_row, column=2, value="Ticker").font = BLUE_FONT
    ws.cell(row=current_row, column=3, value=summary.get('ticker', '')).font = BLUE_FONT
    current_row += 1

    ws.cell(row=current_row, column=2, value="CIK").font = BLUE_FONT
    ws.cell(row=current_row, column=3, value=summary.get('cik', '')).font = BLUE_FONT
    current_row += 1

    ws.cell(row=current_row, column=2, value="SIC Code / Sector").font = BLUE_FONT
    sic_code = summary.get('sic_code', '')
    sic_desc = summary.get('sic_description', '')
    ws.cell(row=current_row, column=3, value=f"{sic_code} - {sic_desc}").font = BLUE_FONT
    current_row += 2

    # Current market data
    ws.cell(row=current_row, column=2, value="Market Data").font = SUBSECTION_FONT
    ws.cell(row=current_row, column=2).fill = SUBSECTION_FILL
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
    current_row += 1

    # Current share price
    ws.cell(row=current_row, column=2, value="Current Share Price").font = BLUE_FONT
    price_cell = ws.cell(row=current_row, column=3)
    current_price = summary.get('current_price')
    price_user_provided = is_user_provided_field(summary, 'current_price')
    if current_price is not None:
        price_cell.value = current_price
        price_cell.number_format = '$#,##0.00'
        if price_user_provided:
            price_cell.fill = USER_PROVIDED_FILL
            price_cell.comment = Comment("User-provided value", "User")
    else:
        price_cell.value = 0
        price_cell.comment = Comment("Price not available - update manually", "System")
    price_cell.font = BLUE_FONT
    price_ref = f"C{current_row}"
    named_ranges['CurrentPrice'] = create_defined_name(wb, 'CurrentPrice', 'Assumptions', price_ref)
    current_row += 1

    # Shares outstanding
    ws.cell(row=current_row, column=2, value="Shares Outstanding").font = BLUE_FONT
    shares_cell = ws.cell(row=current_row, column=3)
    shares = summary.get('shares_outstanding')
    if shares is not None:
        shares_cell.value = shares
        shares_cell.number_format = '#,##0'
    else:
        shares_cell.value = 0
        shares_cell.comment = Comment("Shares not available - update manually", "System")
    shares_cell.font = BLUE_FONT
    shares_ref = f"C{current_row}"
    named_ranges['SharesOutstanding'] = create_defined_name(wb, 'SharesOutstanding', 'Assumptions', shares_ref)
    current_row += 1

    # Market cap (FORMULA - black font)
    ws.cell(row=current_row, column=2, value="Market Cap").font = BLACK_FONT
    mktcap_cell = ws.cell(row=current_row, column=3)
    mktcap_cell.value = f"={price_ref}*{shares_ref}"
    mktcap_cell.font = BLACK_FONT
    mktcap_cell.number_format = CURRENCY_FORMAT
    named_ranges['MarketCap'] = create_defined_name(wb, 'MarketCap', 'Assumptions', f"C{current_row}")
    current_row += 2

    # ==========================================================================
    # 5-YEAR HISTORICAL FINANCIALS
    # ==========================================================================

    ws.cell(row=current_row, column=2, value="Historical Financials (5-Year)").font = SUBSECTION_FONT
    ws.cell(row=current_row, column=2).fill = SUBSECTION_FILL
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=7)
    current_row += 1

    # Get fiscal years sorted (most recent first)
    fiscal_years = summary.get('fiscal_years', {})
    sorted_fys = sorted(fiscal_years.keys(), reverse=True)[:5]

    # Header row with fiscal years
    ws.cell(row=current_row, column=2, value="Metric").font = BLUE_FONT_BOLD
    for i, fy in enumerate(sorted_fys):
        col = 3 + i
        ws.cell(row=current_row, column=col, value=f"FY{fy}").font = BLUE_FONT_BOLD
        ws.cell(row=current_row, column=col).alignment = CENTER_ALIGN
    current_row += 1

    # Revenue row
    revenue_row = current_row
    ws.cell(row=current_row, column=2, value="Revenue").font = BLUE_FONT
    for i, fy in enumerate(sorted_fys):
        col = 3 + i
        cell = ws.cell(row=current_row, column=col)
        fy_data = fiscal_years.get(fy, {})
        revenue = fy_data.get('revenue')
        if revenue is not None:
            cell.value = revenue
        else:
            cell.value = 0
            cell.fill = DEFAULTED_FILL
            cell.comment = Comment("Defaulted — not reported", "System")
        cell.font = BLUE_FONT
        cell.number_format = CURRENCY_FORMAT
    current_row += 1

    # Operating Income row
    ws.cell(row=current_row, column=2, value="Operating Income").font = BLUE_FONT
    for i, fy in enumerate(sorted_fys):
        col = 3 + i
        cell = ws.cell(row=current_row, column=col)
        fy_data = fiscal_years.get(fy, {})
        op_income = fy_data.get('operating_income')
        if op_income is not None:
            cell.value = op_income
        else:
            cell.value = 0
            cell.fill = DEFAULTED_FILL
            cell.comment = Comment("Defaulted — not reported", "System")
        cell.font = BLUE_FONT
        cell.number_format = CURRENCY_FORMAT
    op_income_row = current_row
    current_row += 1

    # D&A row
    ws.cell(row=current_row, column=2, value="D&A").font = BLUE_FONT
    for i, fy in enumerate(sorted_fys):
        col = 3 + i
        cell = ws.cell(row=current_row, column=col)
        fy_data = fiscal_years.get(fy, {})
        da = fy_data.get('da')
        if da is not None:
            cell.value = da
        else:
            cell.value = 0
            cell.fill = DEFAULTED_FILL
            cell.comment = Comment("Defaulted — not reported", "System")
        cell.font = BLUE_FONT
        cell.number_format = CURRENCY_FORMAT
    da_row = current_row
    current_row += 1

    # EBITDA row (FORMULA - black font)
    ebitda_row = current_row
    ws.cell(row=current_row, column=2, value="EBITDA (Calculated)").font = BLACK_FONT
    for i, fy in enumerate(sorted_fys):
        col = 3 + i
        col_letter = get_column_letter(col)
        cell = ws.cell(row=current_row, column=col)
        cell.value = f"={col_letter}{op_income_row}+{col_letter}{da_row}"
        cell.font = BLACK_FONT
        cell.number_format = CURRENCY_FORMAT
    current_row += 2

    # Balance sheet items - most recent only
    ws.cell(row=current_row, column=2, value="Most Recent Balance Sheet Items").font = SUBSECTION_FONT
    ws.cell(row=current_row, column=2).fill = SUBSECTION_FILL
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
    current_row += 1

    most_recent_fy = sorted_fys[0] if sorted_fys else None
    most_recent_data = fiscal_years.get(most_recent_fy, {}) if most_recent_fy else {}

    # Total Debt
    ws.cell(row=current_row, column=2, value="Total Debt").font = BLUE_FONT
    debt_cell = ws.cell(row=current_row, column=3)
    total_debt = most_recent_data.get('total_debt')
    debt_user_provided = is_user_provided_field(summary, 'total_debt')
    debt_defaulted = is_defaulted_field(summary, validation_result, 'total_debt')
    if total_debt is not None:
        debt_cell.value = total_debt
    else:
        debt_cell.value = 0
        if not debt_user_provided:
            debt_defaulted = True
    if debt_user_provided:
        debt_cell.fill = USER_PROVIDED_FILL
        debt_cell.comment = Comment("User-provided value", "User")
    elif debt_defaulted:
        debt_cell.fill = DEFAULTED_FILL
        debt_cell.comment = Comment("Defaulted — not reported (may be debt-free)", "System")
    debt_cell.font = BLUE_FONT
    debt_cell.number_format = CURRENCY_FORMAT
    named_ranges['TotalDebt'] = create_defined_name(wb, 'TotalDebt', 'Assumptions', f"C{current_row}")
    current_row += 1

    # Cash
    ws.cell(row=current_row, column=2, value="Cash & Equivalents").font = BLUE_FONT
    cash_cell = ws.cell(row=current_row, column=3)
    cash = most_recent_data.get('cash')
    cash_user_provided = is_user_provided_field(summary, 'cash')
    cash_defaulted = is_defaulted_field(summary, validation_result, 'cash')
    if cash is not None:
        cash_cell.value = cash
    else:
        cash_cell.value = 0
        if not cash_user_provided:
            cash_defaulted = True
    if cash_user_provided:
        cash_cell.fill = USER_PROVIDED_FILL
        cash_cell.comment = Comment("User-provided value", "User")
    elif cash_defaulted:
        cash_cell.fill = DEFAULTED_FILL
        cash_cell.comment = Comment("Defaulted — not reported", "System")
    cash_cell.font = BLUE_FONT
    cash_cell.number_format = CURRENCY_FORMAT
    named_ranges['Cash'] = create_defined_name(wb, 'Cash', 'Assumptions', f"C{current_row}")
    current_row += 1

    # CapEx
    ws.cell(row=current_row, column=2, value="Capital Expenditures").font = BLUE_FONT
    capex_cell = ws.cell(row=current_row, column=3)
    capex = most_recent_data.get('capex')
    capex_user_provided = is_user_provided_field(summary, 'capex')
    capex_defaulted = is_defaulted_field(summary, validation_result, 'capex')
    if capex is not None:
        capex_cell.value = capex
    else:
        capex_cell.value = 0
        if not capex_user_provided:
            capex_defaulted = True
    if capex_user_provided:
        capex_cell.fill = USER_PROVIDED_FILL
        capex_cell.comment = Comment("User-provided value", "User")
    elif capex_defaulted:
        capex_cell.fill = DEFAULTED_FILL
        capex_cell.comment = Comment("Defaulted — not reported", "System")
    capex_cell.font = BLUE_FONT
    capex_cell.number_format = CURRENCY_FORMAT
    named_ranges['CapEx'] = create_defined_name(wb, 'CapEx', 'Assumptions', f"C{current_row}")
    capex_row = current_row
    current_row += 3

    # ==========================================================================
    # SECTION B: DEAL ASSUMPTIONS
    # ==========================================================================

    ws.cell(row=current_row, column=2, value="SECTION B: DEAL ASSUMPTIONS (User Inputs)")
    ws.cell(row=current_row, column=2).font = SECTION_HEADER_FONT
    ws.cell(row=current_row, column=2).fill = SECTION_HEADER_FILL
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)
    current_row += 2

    # Calculate default values
    growth_rate, growth_is_fallback, growth_comment = calculate_revenue_growth_rate(summary)

    # Most recent revenue and EBITDA for ratio calculations
    most_recent_revenue = most_recent_data.get('revenue', 0) or 1  # Avoid div by zero
    most_recent_ebitda = most_recent_data.get('ebitda_calculated')
    if most_recent_ebitda is None:
        op_inc = most_recent_data.get('operating_income', 0) or 0
        da = most_recent_data.get('da', 0) or 0
        most_recent_ebitda = op_inc + da

    most_recent_capex = most_recent_data.get('capex', 0) or 0
    most_recent_da = most_recent_data.get('da', 0) or 0

    ebitda_margin = most_recent_ebitda / most_recent_revenue if most_recent_revenue > 0 else 0.15
    capex_pct = most_recent_capex / most_recent_revenue if most_recent_revenue > 0 else 0.02
    da_pct = most_recent_da / most_recent_revenue if most_recent_revenue > 0 else 0.03

    # If capex was defaulted, use 2%
    if capex_defaulted:
        capex_pct = 0.02

    # --- Deal Assumptions Inputs ---

    # Revenue Growth Rate
    ws.cell(row=current_row, column=2, value="Revenue Growth Rate").font = BLUE_FONT
    growth_cell = ws.cell(row=current_row, column=3)
    growth_cell.value = growth_rate
    growth_cell.font = BLUE_FONT
    growth_cell.number_format = PERCENT_FORMAT
    growth_cell.comment = Comment(growth_comment, "System")
    if growth_is_fallback:
        # Apply orange fill for implausible/fallback calculation
        growth_cell.fill = IMPLAUSIBLE_CALC_FILL
    named_ranges['RevenueGrowthRate'] = create_defined_name(wb, 'RevenueGrowthRate', 'Assumptions', f"C{current_row}")
    current_row += 1

    # EBITDA Margin
    ws.cell(row=current_row, column=2, value="EBITDA Margin %").font = BLUE_FONT
    margin_cell = ws.cell(row=current_row, column=3)
    margin_cell.value = ebitda_margin
    margin_cell.font = BLUE_FONT
    margin_cell.number_format = PERCENT_FORMAT
    margin_cell.comment = Comment(f"Default: most recent FY EBITDA / Revenue", "System")
    named_ranges['EBITDAMargin'] = create_defined_name(wb, 'EBITDAMargin', 'Assumptions', f"C{current_row}")
    current_row += 1

    # Entry EV/EBITDA Multiple
    ws.cell(row=current_row, column=2, value="Entry EV/EBITDA Multiple").font = BLUE_FONT
    entry_mult_cell = ws.cell(row=current_row, column=3)
    entry_mult_cell.value = 8.0
    entry_mult_cell.font = BLUE_FONT
    entry_mult_cell.number_format = MULTIPLE_FORMAT
    entry_mult_cell.comment = Comment("Default: 8.0x placeholder — adjust based on sector comps", "System")
    entry_mult_ref = f"C{current_row}"
    named_ranges['EntryMultiple'] = create_defined_name(wb, 'EntryMultiple', 'Assumptions', entry_mult_ref)
    current_row += 1

    # Offer Premium
    ws.cell(row=current_row, column=2, value="Offer Premium %").font = BLUE_FONT
    premium_cell = ws.cell(row=current_row, column=3)
    premium_cell.value = 0.25
    premium_cell.font = BLUE_FONT
    premium_cell.number_format = PERCENT_FORMAT
    named_ranges['OfferPremium'] = create_defined_name(wb, 'OfferPremium', 'Assumptions', f"C{current_row}")
    current_row += 1

    # Leverage Multiple
    ws.cell(row=current_row, column=2, value="Leverage Multiple").font = BLUE_FONT
    leverage_cell = ws.cell(row=current_row, column=3)
    leverage_cell.value = 5.5
    leverage_cell.font = BLUE_FONT
    leverage_cell.number_format = MULTIPLE_FORMAT
    named_ranges['LeverageMultiple'] = create_defined_name(wb, 'LeverageMultiple', 'Assumptions', f"C{current_row}")
    current_row += 1

    # Interest Rate
    ws.cell(row=current_row, column=2, value="Interest Rate").font = BLUE_FONT
    interest_cell = ws.cell(row=current_row, column=3)
    interest_cell.value = 0.08
    interest_cell.font = BLUE_FONT
    interest_cell.number_format = PERCENT_FORMAT
    named_ranges['InterestRate'] = create_defined_name(wb, 'InterestRate', 'Assumptions', f"C{current_row}")
    current_row += 1

    # Tax Rate
    ws.cell(row=current_row, column=2, value="Tax Rate").font = BLUE_FONT
    tax_cell = ws.cell(row=current_row, column=3)
    tax_cell.value = 0.25
    tax_cell.font = BLUE_FONT
    tax_cell.number_format = PERCENT_FORMAT
    named_ranges['TaxRate'] = create_defined_name(wb, 'TaxRate', 'Assumptions', f"C{current_row}")
    current_row += 1

    # CapEx %
    ws.cell(row=current_row, column=2, value="CapEx % of Revenue").font = BLUE_FONT
    capex_pct_cell = ws.cell(row=current_row, column=3)
    capex_pct_cell.value = capex_pct
    capex_pct_cell.font = BLUE_FONT
    capex_pct_cell.number_format = PERCENT_FORMAT
    if capex_defaulted:
        capex_pct_cell.comment = Comment("Default: 2% (CapEx not reported)", "System")
    else:
        capex_pct_cell.comment = Comment("Default: most recent FY CapEx / Revenue", "System")
    named_ranges['CapExPct'] = create_defined_name(wb, 'CapExPct', 'Assumptions', f"C{current_row}")
    current_row += 1

    # D&A %
    ws.cell(row=current_row, column=2, value="D&A % of Revenue").font = BLUE_FONT
    da_pct_cell = ws.cell(row=current_row, column=3)
    da_pct_cell.value = da_pct
    da_pct_cell.font = BLUE_FONT
    da_pct_cell.number_format = PERCENT_FORMAT
    da_pct_cell.comment = Comment("Default: most recent FY D&A / Revenue", "System")
    named_ranges['DAPct'] = create_defined_name(wb, 'DAPct', 'Assumptions', f"C{current_row}")
    current_row += 1

    # Change in NWC %
    ws.cell(row=current_row, column=2, value="Change in NWC % of Revenue").font = BLUE_FONT
    nwc_cell = ws.cell(row=current_row, column=3)
    nwc_cell.value = 0.0
    nwc_cell.font = BLUE_FONT
    nwc_cell.number_format = PERCENT_FORMAT
    named_ranges['NWCPct'] = create_defined_name(wb, 'NWCPct', 'Assumptions', f"C{current_row}")
    current_row += 1

    # Exit Year
    ws.cell(row=current_row, column=2, value="Exit Year").font = BLUE_FONT
    exit_year_cell = ws.cell(row=current_row, column=3)
    exit_year_cell.value = 5
    exit_year_cell.font = BLUE_FONT
    exit_year_cell.number_format = '0'
    named_ranges['ExitYear'] = create_defined_name(wb, 'ExitYear', 'Assumptions', f"C{current_row}")
    current_row += 1

    # Exit EV/EBITDA Multiple (FORMULA by default - references entry multiple)
    ws.cell(row=current_row, column=2, value="Exit EV/EBITDA Multiple").font = BLACK_FONT
    exit_mult_cell = ws.cell(row=current_row, column=3)
    exit_mult_cell.value = f"={entry_mult_ref}"  # Formula referencing entry multiple
    exit_mult_cell.font = BLACK_FONT
    exit_mult_cell.number_format = MULTIPLE_FORMAT
    exit_mult_cell.comment = Comment(
        "Default: equals Entry Multiple (no expansion/contraction). "
        "Override with hardcoded value to model multiple expansion/contraction.",
        "System"
    )
    named_ranges['ExitMultiple'] = create_defined_name(wb, 'ExitMultiple', 'Assumptions', f"C{current_row}")
    current_row += 1

    # Transaction Fee %
    ws.cell(row=current_row, column=2, value="Transaction Fee %").font = BLUE_FONT
    txn_fee_cell = ws.cell(row=current_row, column=3)
    txn_fee_cell.value = 0.02
    txn_fee_cell.font = BLUE_FONT
    txn_fee_cell.number_format = PERCENT_FORMAT
    named_ranges['TransactionFeePct'] = create_defined_name(wb, 'TransactionFeePct', 'Assumptions', f"C{current_row}")
    current_row += 1

    # Mandatory Amortization %
    ws.cell(row=current_row, column=2, value="Mandatory Amortization %").font = BLUE_FONT
    amort_cell = ws.cell(row=current_row, column=3)
    amort_cell.value = 0.01
    amort_cell.font = BLUE_FONT
    amort_cell.number_format = PERCENT_FORMAT
    named_ranges['AmortizationPct'] = create_defined_name(wb, 'AmortizationPct', 'Assumptions', f"C{current_row}")
    current_row += 2

    # Store Entry Revenue and EBITDA references (most recent FY values for projections)
    # Column C is most recent fiscal year
    if sorted_fys:
        most_recent_col = get_column_letter(3)  # Column C is most recent
        entry_revenue_ref = f"{most_recent_col}{revenue_row}"
        entry_ebitda_ref = f"{most_recent_col}{ebitda_row}"
        named_ranges['EntryRevenue'] = create_defined_name(wb, 'EntryRevenue', 'Assumptions', entry_revenue_ref)
        named_ranges['EntryEBITDA'] = create_defined_name(wb, 'EntryEBITDA', 'Assumptions', entry_ebitda_ref)

    return named_ranges


# =============================================================================
# SOURCES & USES TAB BUILDER
# =============================================================================

def build_sources_uses_tab(wb: Workbook, named_ranges: dict):
    """
    Build the Sources & Uses tab.

    All values are formulas (black font) referencing the Assumptions tab.
    """
    ws = wb.create_sheet("Sources & Uses", 1)

    # Column widths
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 38  # Wide enough for longest labels
    ws.column_dimensions['C'].width = 22  # Wide enough for $X,XXX,XXX,XXX,XXX format
    ws.column_dimensions['D'].width = 22

    current_row = 1

    # ==========================================================================
    # USES OF FUNDS
    # ==========================================================================

    ws.cell(row=current_row, column=2, value="USES OF FUNDS")
    ws.cell(row=current_row, column=2).font = SECTION_HEADER_FONT
    ws.cell(row=current_row, column=2).fill = SECTION_HEADER_FILL
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
    current_row += 2

    # Purchase Enterprise Value = Entry EBITDA × Entry Multiple
    # Use direct cell references instead of named ranges for cross-sheet reliability
    ws.cell(row=current_row, column=2, value="Purchase Enterprise Value").font = BLACK_FONT
    pev_cell = ws.cell(row=current_row, column=3)
    pev_cell.value = "='Assumptions'!$C$18*'Assumptions'!$C$30"  # EntryEBITDA * EntryMultiple
    pev_cell.font = BLACK_FONT
    pev_cell.number_format = CURRENCY_FORMAT
    pev_row = current_row
    current_row += 1

    # Transaction Fees = Purchase EV × Fee %
    ws.cell(row=current_row, column=2, value="Transaction Fees").font = BLACK_FONT
    fees_cell = ws.cell(row=current_row, column=3)
    fees_cell.value = f"=C{pev_row}*'Assumptions'!$C$40"  # TransactionFeePct
    fees_cell.font = BLACK_FONT
    fees_cell.number_format = CURRENCY_FORMAT
    fees_row = current_row
    current_row += 1

    # Total Uses = sum
    current_row += 1
    ws.cell(row=current_row, column=2, value="Total Uses").font = BLACK_FONT_BOLD
    total_uses_cell = ws.cell(row=current_row, column=3)
    total_uses_cell.value = f"=C{pev_row}+C{fees_row}"
    total_uses_cell.font = BLACK_FONT_BOLD
    total_uses_cell.number_format = CURRENCY_FORMAT
    total_uses_cell.border = Border(top=Side(style='thin'), bottom=Side(style='double'))
    total_uses_row = current_row

    # Create named range for Total Uses
    from openpyxl.workbook.defined_name import DefinedName
    defn = DefinedName('TotalUses', attr_text=f"'Sources & Uses'!C{total_uses_row}")
    wb.defined_names['TotalUses'] = defn

    current_row += 3

    # ==========================================================================
    # SOURCES OF FUNDS
    # ==========================================================================

    ws.cell(row=current_row, column=2, value="SOURCES OF FUNDS")
    ws.cell(row=current_row, column=2).font = SECTION_HEADER_FONT
    ws.cell(row=current_row, column=2).fill = SECTION_HEADER_FILL
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
    current_row += 2

    # New Debt Raised = Entry EBITDA × Leverage Multiple
    # Use direct cell references instead of named ranges for cross-sheet reliability
    ws.cell(row=current_row, column=2, value="New Debt Raised").font = BLACK_FONT
    debt_cell = ws.cell(row=current_row, column=3)
    debt_cell.value = "='Assumptions'!$C$18*'Assumptions'!$C$32"  # EntryEBITDA * LeverageMultiple
    debt_cell.font = BLACK_FONT
    debt_cell.number_format = CURRENCY_FORMAT
    debt_row = current_row

    # Create named range for NewDebtRaised
    defn = DefinedName('NewDebtRaised', attr_text=f"'Sources & Uses'!C{debt_row}")
    wb.defined_names['NewDebtRaised'] = defn

    current_row += 1

    # Sponsor Equity Check = Total Uses - New Debt Raised
    ws.cell(row=current_row, column=2, value="Sponsor Equity Check").font = BLACK_FONT
    equity_cell = ws.cell(row=current_row, column=3)
    equity_cell.value = f"=C{total_uses_row}-C{debt_row}"
    equity_cell.font = BLACK_FONT
    equity_cell.number_format = CURRENCY_FORMAT
    equity_row = current_row

    # Create named range for SponsorEquity
    defn = DefinedName('SponsorEquity', attr_text=f"'Sources & Uses'!C{equity_row}")
    wb.defined_names['SponsorEquity'] = defn

    current_row += 1

    # Total Sources = New Debt + Sponsor Equity
    current_row += 1
    ws.cell(row=current_row, column=2, value="Total Sources").font = BLACK_FONT_BOLD
    total_sources_cell = ws.cell(row=current_row, column=3)
    total_sources_cell.value = f"=C{debt_row}+C{equity_row}"
    total_sources_cell.font = BLACK_FONT_BOLD
    total_sources_cell.number_format = CURRENCY_FORMAT
    total_sources_cell.border = Border(top=Side(style='thin'), bottom=Side(style='double'))
    total_sources_row = current_row

    # Create named range for TotalSources
    defn = DefinedName('TotalSources', attr_text=f"'Sources & Uses'!C{total_sources_row}")
    wb.defined_names['TotalSources'] = defn

    current_row += 3

    # ==========================================================================
    # BALANCE CHECK
    # ==========================================================================

    ws.cell(row=current_row, column=2, value="Balance Check").font = SUBSECTION_FONT
    ws.cell(row=current_row, column=2).fill = SUBSECTION_FILL
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
    current_row += 1

    ws.cell(row=current_row, column=2, value="Sources - Uses (should = 0)").font = BLACK_FONT
    balance_cell = ws.cell(row=current_row, column=3)
    balance_cell.value = f"=C{total_sources_row}-C{total_uses_row}"
    balance_cell.font = BLACK_FONT
    balance_cell.number_format = CURRENCY_FORMAT
    balance_row = current_row

    # Add conditional formatting: red fill if not zero
    ws.conditional_formatting.add(
        f"C{balance_row}",
        FormulaRule(
            formula=[f"C{balance_row}<>0"],
            fill=ERROR_FILL
        )
    )

    # Create named range for BalanceCheck
    defn = DefinedName('BalanceCheck', attr_text=f"'Sources & Uses'!C{balance_row}")
    wb.defined_names['BalanceCheck'] = defn

    return {
        'pev_row': pev_row,
        'total_uses_row': total_uses_row,
        'debt_row': debt_row,
        'equity_row': equity_row,
        'total_sources_row': total_sources_row,
        'balance_row': balance_row
    }


# =============================================================================
# OPERATING MODEL TAB BUILDER
# =============================================================================

# Placeholder styling for Interest Expense (will be wired to Debt Schedule later)
PLACEHOLDER_FILL = PatternFill(start_color='FFB366', end_color='FFB366', fill_type='solid')


def build_operating_model_tab(wb: Workbook, summary: dict, named_ranges: dict) -> dict:
    """
    Build the Operating Model tab.

    Projects Revenue through Free Cash Flow for each year from Year 1 to Exit Year.
    All values are formulas (black font) referencing the Assumptions tab.

    Returns a dict with row references for use by future tabs (Debt Schedule, Returns).
    """
    ws = wb.create_sheet("Operating Model", 2)

    # Get exit year from assumptions (default 5)
    exit_year = 5  # Default - will be read from named range in formulas

    # Get most recent fiscal year for calendar year labels
    fiscal_years = summary.get('fiscal_years', {})
    if fiscal_years:
        most_recent_fy = max(fiscal_years.keys())
    else:
        most_recent_fy = 2024  # Fallback

    # Column widths
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 38  # Wide enough for longest labels

    # Year columns start at C - set wide enough for $X,XXX,XXX,XXX,XXX format
    for i in range(exit_year):
        col_letter = get_column_letter(3 + i)
        ws.column_dimensions[col_letter].width = 22

    current_row = 1

    # ==========================================================================
    # HEADER
    # ==========================================================================

    ws.cell(row=current_row, column=2, value="OPERATING MODEL (Projected)")
    ws.cell(row=current_row, column=2).font = SECTION_HEADER_FONT
    ws.cell(row=current_row, column=2).fill = SECTION_HEADER_FILL
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=2 + exit_year)
    current_row += 2

    # Year headers
    ws.cell(row=current_row, column=2, value="").font = BLACK_FONT_BOLD
    for year in range(1, exit_year + 1):
        col = 2 + year  # Column C = Year 1, D = Year 2, etc.
        calendar_year = most_recent_fy + year
        cell = ws.cell(row=current_row, column=col)
        cell.value = f"Year {year}\n(FY{calendar_year})"
        cell.font = BLACK_FONT_BOLD
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    header_row = current_row
    current_row += 2

    # ==========================================================================
    # INCOME STATEMENT SECTION
    # ==========================================================================

    ws.cell(row=current_row, column=2, value="Income Statement").font = SUBSECTION_FONT
    ws.cell(row=current_row, column=2).fill = SUBSECTION_FILL
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=2 + exit_year)
    current_row += 1

    # --- Revenue ---
    revenue_row = current_row
    ws.cell(row=current_row, column=2, value="Revenue").font = BLACK_FONT
    for year in range(1, exit_year + 1):
        col = 2 + year
        col_letter = get_column_letter(col)
        cell = ws.cell(row=current_row, column=col)

        if year == 1:
            # Year 1: Most recent FY Revenue × (1 + Growth Rate)
            # Use direct cell references instead of named ranges for cross-sheet reliability
            cell.value = "='Assumptions'!$C$15*(1+'Assumptions'!$C$28)"  # EntryRevenue * (1+RevenueGrowthRate)
        else:
            # Year 2+: Prior year Revenue × (1 + Growth Rate)
            prev_col = get_column_letter(col - 1)
            cell.value = f"={prev_col}{revenue_row}*(1+'Assumptions'!$C$28)"  # RevenueGrowthRate

        cell.font = BLACK_FONT
        cell.number_format = CURRENCY_FORMAT
    current_row += 1

    # --- EBITDA ---
    ebitda_row = current_row
    ws.cell(row=current_row, column=2, value="EBITDA").font = BLACK_FONT
    for year in range(1, exit_year + 1):
        col = 2 + year
        col_letter = get_column_letter(col)
        cell = ws.cell(row=current_row, column=col)
        cell.value = f"={col_letter}{revenue_row}*'Assumptions'!$C$29"  # EBITDAMargin
        cell.font = BLACK_FONT
        cell.number_format = CURRENCY_FORMAT
    current_row += 1

    # --- D&A ---
    da_row = current_row
    ws.cell(row=current_row, column=2, value="Less: D&A").font = BLACK_FONT
    for year in range(1, exit_year + 1):
        col = 2 + year
        col_letter = get_column_letter(col)
        cell = ws.cell(row=current_row, column=col)
        cell.value = f"={col_letter}{revenue_row}*'Assumptions'!$C$36"  # DAPct
        cell.font = BLACK_FONT
        cell.number_format = CURRENCY_FORMAT
    current_row += 1

    # --- EBIT ---
    ebit_row = current_row
    ws.cell(row=current_row, column=2, value="EBIT").font = BLACK_FONT_BOLD
    for year in range(1, exit_year + 1):
        col = 2 + year
        col_letter = get_column_letter(col)
        cell = ws.cell(row=current_row, column=col)
        cell.value = f"={col_letter}{ebitda_row}-{col_letter}{da_row}"
        cell.font = BLACK_FONT_BOLD
        cell.number_format = CURRENCY_FORMAT
        cell.border = Border(top=Side(style='thin'))
    current_row += 1

    # --- Interest Expense (PLACEHOLDER) ---
    interest_row = current_row
    ws.cell(row=current_row, column=2, value="Less: Interest Expense").font = BLACK_FONT
    for year in range(1, exit_year + 1):
        col = 2 + year
        cell = ws.cell(row=current_row, column=col)
        cell.value = 0  # Hardcoded placeholder
        cell.font = BLACK_FONT
        cell.number_format = CURRENCY_FORMAT
        cell.fill = PLACEHOLDER_FILL
        cell.comment = Comment(
            "PLACEHOLDER - will be replaced with a reference to the Debt Schedule tab once built.",
            "System"
        )
    current_row += 1

    # --- Pre-Tax Income ---
    pretax_row = current_row
    ws.cell(row=current_row, column=2, value="Pre-Tax Income").font = BLACK_FONT
    for year in range(1, exit_year + 1):
        col = 2 + year
        col_letter = get_column_letter(col)
        cell = ws.cell(row=current_row, column=col)
        cell.value = f"={col_letter}{ebit_row}-{col_letter}{interest_row}"
        cell.font = BLACK_FONT
        cell.number_format = CURRENCY_FORMAT
    current_row += 1

    # --- Taxes (floored at 0) ---
    taxes_row = current_row
    ws.cell(row=current_row, column=2, value="Less: Taxes").font = BLACK_FONT
    for year in range(1, exit_year + 1):
        col = 2 + year
        col_letter = get_column_letter(col)
        cell = ws.cell(row=current_row, column=col)
        # MAX(0, Pre-Tax Income × TaxRate) - no negative taxes
        cell.value = f"=MAX(0,{col_letter}{pretax_row}*'Assumptions'!$C$34)"  # TaxRate
        cell.font = BLACK_FONT
        cell.number_format = CURRENCY_FORMAT
    current_row += 1

    # --- Net Income ---
    net_income_row = current_row
    ws.cell(row=current_row, column=2, value="Net Income").font = BLACK_FONT_BOLD
    for year in range(1, exit_year + 1):
        col = 2 + year
        col_letter = get_column_letter(col)
        cell = ws.cell(row=current_row, column=col)
        cell.value = f"={col_letter}{pretax_row}-{col_letter}{taxes_row}"
        cell.font = BLACK_FONT_BOLD
        cell.number_format = CURRENCY_FORMAT
        cell.border = Border(top=Side(style='thin'), bottom=Side(style='double'))
    current_row += 2

    # ==========================================================================
    # FREE CASH FLOW SECTION
    # ==========================================================================

    ws.cell(row=current_row, column=2, value="Free Cash Flow Build").font = SUBSECTION_FONT
    ws.cell(row=current_row, column=2).fill = SUBSECTION_FILL
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=2 + exit_year)
    current_row += 1

    # --- Net Income (reference) ---
    fcf_net_income_row = current_row
    ws.cell(row=current_row, column=2, value="Net Income").font = BLACK_FONT
    for year in range(1, exit_year + 1):
        col = 2 + year
        col_letter = get_column_letter(col)
        cell = ws.cell(row=current_row, column=col)
        cell.value = f"={col_letter}{net_income_row}"
        cell.font = BLACK_FONT
        cell.number_format = CURRENCY_FORMAT
    current_row += 1

    # --- Plus: D&A (add back) ---
    fcf_da_row = current_row
    ws.cell(row=current_row, column=2, value="Plus: D&A (non-cash add-back)").font = BLACK_FONT
    for year in range(1, exit_year + 1):
        col = 2 + year
        col_letter = get_column_letter(col)
        cell = ws.cell(row=current_row, column=col)
        cell.value = f"={col_letter}{da_row}"
        cell.font = BLACK_FONT
        cell.number_format = CURRENCY_FORMAT
    current_row += 1

    # --- Less: CapEx ---
    capex_row = current_row
    ws.cell(row=current_row, column=2, value="Less: CapEx").font = BLACK_FONT
    for year in range(1, exit_year + 1):
        col = 2 + year
        col_letter = get_column_letter(col)
        cell = ws.cell(row=current_row, column=col)
        cell.value = f"={col_letter}{revenue_row}*'Assumptions'!$C$35"  # CapExPct
        cell.font = BLACK_FONT
        cell.number_format = CURRENCY_FORMAT
    current_row += 1

    # --- Less: Change in NWC ---
    nwc_row = current_row
    ws.cell(row=current_row, column=2, value="Less: Change in NWC").font = BLACK_FONT
    for year in range(1, exit_year + 1):
        col = 2 + year
        col_letter = get_column_letter(col)
        cell = ws.cell(row=current_row, column=col)
        # NWC change = Revenue × NWCPct (consumes cash when positive)
        cell.value = f"={col_letter}{revenue_row}*'Assumptions'!$C$37"  # NWCPct
        cell.font = BLACK_FONT
        cell.number_format = CURRENCY_FORMAT
    current_row += 1

    # --- Free Cash Flow ---
    fcf_row = current_row
    ws.cell(row=current_row, column=2, value="Free Cash Flow for Debt Paydown").font = BLACK_FONT_BOLD
    for year in range(1, exit_year + 1):
        col = 2 + year
        col_letter = get_column_letter(col)
        cell = ws.cell(row=current_row, column=col)
        # FCF = Net Income + D&A - CapEx - Change in NWC
        cell.value = f"={col_letter}{fcf_net_income_row}+{col_letter}{fcf_da_row}-{col_letter}{capex_row}-{col_letter}{nwc_row}"
        cell.font = BLACK_FONT_BOLD
        cell.number_format = CURRENCY_FORMAT
        cell.border = Border(top=Side(style='thin'), bottom=Side(style='double'))
    current_row += 2

    # ==========================================================================
    # KEY METRICS (for quick reference)
    # ==========================================================================

    ws.cell(row=current_row, column=2, value="Key Metrics").font = SUBSECTION_FONT
    ws.cell(row=current_row, column=2).fill = SUBSECTION_FILL
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=2 + exit_year)
    current_row += 1

    # --- EBITDA Margin % ---
    ws.cell(row=current_row, column=2, value="EBITDA Margin %").font = BLACK_FONT
    for year in range(1, exit_year + 1):
        col = 2 + year
        col_letter = get_column_letter(col)
        cell = ws.cell(row=current_row, column=col)
        cell.value = f"={col_letter}{ebitda_row}/{col_letter}{revenue_row}"
        cell.font = BLACK_FONT
        cell.number_format = PERCENT_FORMAT
    current_row += 1

    # --- FCF Conversion % (FCF / EBITDA) ---
    ws.cell(row=current_row, column=2, value="FCF Conversion % (FCF/EBITDA)").font = BLACK_FONT
    for year in range(1, exit_year + 1):
        col = 2 + year
        col_letter = get_column_letter(col)
        cell = ws.cell(row=current_row, column=col)
        cell.value = f"=IF({col_letter}{ebitda_row}=0,0,{col_letter}{fcf_row}/{col_letter}{ebitda_row})"
        cell.font = BLACK_FONT
        cell.number_format = PERCENT_FORMAT

    # Create named ranges for key rows that will be used by Debt Schedule and Returns tabs
    from openpyxl.workbook.defined_name import DefinedName

    # Create named ranges for the Operating Model outputs
    # These use the exit year column (last year) for exit calculations
    exit_col = get_column_letter(2 + exit_year)

    # Exit Year EBITDA (for exit valuation)
    defn = DefinedName('ExitEBITDA', attr_text=f"'Operating Model'!{exit_col}{ebitda_row}")
    wb.defined_names['ExitEBITDA'] = defn

    # Create row references for each year's FCF (for debt paydown)
    # Store these as a dict to return
    row_refs = {
        'revenue_row': revenue_row,
        'ebitda_row': ebitda_row,
        'da_row': da_row,
        'ebit_row': ebit_row,
        'interest_row': interest_row,
        'pretax_row': pretax_row,
        'taxes_row': taxes_row,
        'net_income_row': net_income_row,
        'capex_row': capex_row,
        'nwc_row': nwc_row,
        'fcf_row': fcf_row,
        'exit_year': exit_year,
        'year_start_col': 3,  # Column C
    }

    return row_refs


# =============================================================================
# DEBT SCHEDULE TAB BUILDER
# =============================================================================

def build_debt_schedule_tab(wb: Workbook, summary: dict, op_model_refs: dict) -> dict:
    """
    Build the Debt Schedule tab.

    Projects debt balance, interest expense, mandatory amortization, and cash sweep
    for each year from Year 1 to Exit Year.

    Interest Expense is calculated on BEGINNING balance only (not average) to avoid
    circular reference with Operating Model's FCF calculation. This is a disclosed
    simplification that slightly overstates interest in heavy-paydown years.

    Returns a dict with row references for use by other tabs.
    """
    ws = wb.create_sheet("Debt Schedule", 3)

    exit_year = op_model_refs['exit_year']
    fcf_row = op_model_refs['fcf_row']
    ebitda_row = op_model_refs['ebitda_row']
    year_start_col = op_model_refs['year_start_col']

    # Get most recent fiscal year for calendar year labels
    fiscal_years = summary.get('fiscal_years', {})
    if fiscal_years:
        most_recent_fy = max(fiscal_years.keys())
    else:
        most_recent_fy = 2024  # Fallback

    # Column widths
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 38  # Wide enough for longest labels

    # Year columns start at C - set wide enough for $X,XXX,XXX,XXX,XXX format
    for i in range(exit_year):
        col_letter = get_column_letter(3 + i)
        ws.column_dimensions[col_letter].width = 22

    current_row = 1

    # ==========================================================================
    # HEADER
    # ==========================================================================

    ws.cell(row=current_row, column=2, value="DEBT SCHEDULE")
    ws.cell(row=current_row, column=2).font = SECTION_HEADER_FONT
    ws.cell(row=current_row, column=2).fill = SECTION_HEADER_FILL
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=2 + exit_year)
    current_row += 2

    # Year headers
    ws.cell(row=current_row, column=2, value="").font = BLACK_FONT_BOLD
    for year in range(1, exit_year + 1):
        col = 2 + year  # Column C = Year 1, D = Year 2, etc.
        calendar_year = most_recent_fy + year
        cell = ws.cell(row=current_row, column=col)
        cell.value = f"Year {year}\n(FY{calendar_year})"
        cell.font = BLACK_FONT_BOLD
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    current_row += 2

    # ==========================================================================
    # DEBT PAYDOWN SECTION
    # ==========================================================================

    ws.cell(row=current_row, column=2, value="Debt Paydown Schedule").font = SUBSECTION_FONT
    ws.cell(row=current_row, column=2).fill = SUBSECTION_FILL
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=2 + exit_year)
    current_row += 1

    # --- Beginning Debt Balance ---
    beginning_balance_row = current_row
    ws.cell(row=current_row, column=2, value="Beginning Debt Balance").font = BLACK_FONT
    for year in range(1, exit_year + 1):
        col = 2 + year
        col_letter = get_column_letter(col)
        cell = ws.cell(row=current_row, column=col)

        if year == 1:
            # Year 1: Beginning balance = NewDebtRaised from Sources & Uses
            # Use direct cell reference instead of named range for cross-sheet reliability
            cell.value = "='Sources & Uses'!$C$11"  # NewDebtRaised
        else:
            # Year 2+: Prior year's Ending Debt Balance (same-sheet reference, OK as-is)
            prev_col = get_column_letter(col - 1)
            # ending_balance_row will be defined below, so we need to calculate it
            # Ending balance row is: beginning + 5 rows down (interest, mandatory, cash avail, optional, ending)
            ending_row = beginning_balance_row + 5
            cell.value = f"={prev_col}{ending_row}"

        cell.font = BLACK_FONT
        cell.number_format = CURRENCY_FORMAT
    current_row += 1

    # --- Interest Expense ---
    interest_expense_row = current_row
    ws.cell(row=current_row, column=2, value="Interest Expense").font = BLACK_FONT
    for year in range(1, exit_year + 1):
        col = 2 + year
        col_letter = get_column_letter(col)
        cell = ws.cell(row=current_row, column=col)
        # Interest = Beginning Balance × InterestRate (on beginning balance only to avoid circularity)
        # Use direct cell reference instead of named range for cross-sheet reliability
        cell.value = f"={col_letter}{beginning_balance_row}*'Assumptions'!$C$33"  # InterestRate
        cell.font = BLACK_FONT
        cell.number_format = CURRENCY_FORMAT
    current_row += 1

    # --- Mandatory Amortization ---
    mandatory_amort_row = current_row
    ws.cell(row=current_row, column=2, value="Mandatory Amortization").font = BLACK_FONT
    for year in range(1, exit_year + 1):
        col = 2 + year
        cell = ws.cell(row=current_row, column=col)
        # Mandatory = Original principal (NewDebtRaised) × MandatoryAmortPct
        # This stays constant every year (based on original loan amount, not declining balance)
        # Use direct cell references instead of named ranges for cross-sheet reliability
        cell.value = "='Sources & Uses'!$C$11*'Assumptions'!$C$41"  # NewDebtRaised * AmortizationPct
        cell.font = BLACK_FONT
        cell.number_format = CURRENCY_FORMAT
    current_row += 1

    # --- Cash Available for Sweep ---
    cash_available_row = current_row
    ws.cell(row=current_row, column=2, value="Cash Available for Sweep").font = BLACK_FONT
    for year in range(1, exit_year + 1):
        col = 2 + year
        col_letter = get_column_letter(col)
        cell = ws.cell(row=current_row, column=col)
        # Pull FCF for Debt Paydown from Operating Model (same year column)
        cell.value = f"='Operating Model'!{col_letter}{fcf_row}"
        cell.font = BLACK_FONT
        cell.number_format = CURRENCY_FORMAT
    current_row += 1

    # --- Optional Paydown (Cash Sweep) ---
    optional_paydown_row = current_row
    ws.cell(row=current_row, column=2, value="Optional Paydown (Cash Sweep)").font = BLACK_FONT
    for year in range(1, exit_year + 1):
        col = 2 + year
        col_letter = get_column_letter(col)
        cell = ws.cell(row=current_row, column=col)
        # Optional = MIN of:
        #   (a) Cash Available - Mandatory Amortization (can't sweep more cash than you have after mandatory)
        #   (b) Beginning Balance - Mandatory Amortization (can't pay down more debt than exists)
        # Wrapped in MAX(0, ...) to prevent negative paydown in bad years
        cash_minus_mand = f"{col_letter}{cash_available_row}-{col_letter}{mandatory_amort_row}"
        begin_minus_mand = f"{col_letter}{beginning_balance_row}-{col_letter}{mandatory_amort_row}"
        cell.value = f"=MAX(0,MIN({cash_minus_mand},{begin_minus_mand}))"
        cell.font = BLACK_FONT
        cell.number_format = CURRENCY_FORMAT
    current_row += 1

    # --- Ending Debt Balance ---
    ending_balance_row = current_row
    ws.cell(row=current_row, column=2, value="Ending Debt Balance").font = BLACK_FONT_BOLD
    for year in range(1, exit_year + 1):
        col = 2 + year
        col_letter = get_column_letter(col)
        cell = ws.cell(row=current_row, column=col)
        # Ending = Beginning - Mandatory - Optional
        cell.value = f"={col_letter}{beginning_balance_row}-{col_letter}{mandatory_amort_row}-{col_letter}{optional_paydown_row}"
        cell.font = BLACK_FONT_BOLD
        cell.number_format = CURRENCY_FORMAT
        cell.border = Border(top=Side(style='thin'), bottom=Side(style='double'))
    current_row += 2

    # ==========================================================================
    # LEVERAGE METRICS SECTION
    # ==========================================================================

    ws.cell(row=current_row, column=2, value="Leverage Metrics").font = SUBSECTION_FONT
    ws.cell(row=current_row, column=2).fill = SUBSECTION_FILL
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=2 + exit_year)
    current_row += 1

    # --- Leverage Ratio (Ending Debt / EBITDA) ---
    leverage_ratio_row = current_row
    ws.cell(row=current_row, column=2, value="Leverage Ratio (Debt / EBITDA)").font = BLACK_FONT
    for year in range(1, exit_year + 1):
        col = 2 + year
        col_letter = get_column_letter(col)
        cell = ws.cell(row=current_row, column=col)
        # Ending Debt Balance / EBITDA from Operating Model
        cell.value = f"=IF('Operating Model'!{col_letter}{ebitda_row}=0,0,{col_letter}{ending_balance_row}/'Operating Model'!{col_letter}{ebitda_row})"
        cell.font = BLACK_FONT
        cell.number_format = MULTIPLE_FORMAT

    # Create named ranges for key debt schedule values
    from openpyxl.workbook.defined_name import DefinedName

    # Exit year ending debt balance (for returns calculation)
    exit_col = get_column_letter(2 + exit_year)
    defn = DefinedName('ExitDebtBalance', attr_text=f"'Debt Schedule'!{exit_col}{ending_balance_row}")
    wb.defined_names['ExitDebtBalance'] = defn

    # Return row references
    return {
        'beginning_balance_row': beginning_balance_row,
        'interest_expense_row': interest_expense_row,
        'mandatory_amort_row': mandatory_amort_row,
        'cash_available_row': cash_available_row,
        'optional_paydown_row': optional_paydown_row,
        'ending_balance_row': ending_balance_row,
        'leverage_ratio_row': leverage_ratio_row,
        'exit_year': exit_year,
        'year_start_col': year_start_col,
    }


# =============================================================================
# RETURNS TAB BUILDER
# =============================================================================

def build_returns_tab(wb: Workbook, summary: dict, sources_uses_refs: dict,
                      op_model_refs: dict, debt_schedule_refs: dict) -> dict:
    """
    Build the Returns tab.

    Calculates IRR (via XIRR) and MOIC based on:
    - Entry investment (Sponsor Equity from Sources & Uses)
    - Exit proceeds (Exit Equity Value = Exit EV - Exit Debt)

    All formulas use DIRECT cell references (no named ranges) for cross-sheet
    references to avoid the cross-sheet named range resolution bug.

    Returns a dict with row references.
    """
    ws = wb.create_sheet("Returns", 4)

    exit_year = op_model_refs['exit_year']
    ebitda_row = op_model_refs['ebitda_row']
    ending_balance_row = debt_schedule_refs['ending_balance_row']
    equity_row = sources_uses_refs['equity_row']

    # Exit year column (e.g., G for 5-year exit)
    exit_col = get_column_letter(2 + exit_year)

    # Get most recent fiscal year for date calculations
    fiscal_years = summary.get('fiscal_years', {})
    if fiscal_years:
        most_recent_fy = max(fiscal_years.keys())
    else:
        most_recent_fy = 2024  # Fallback

    # Column widths - consistent with other tabs
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 38  # Wide enough for longest labels

    # Data columns - need columns for Year 0 through Exit Year for XIRR
    for i in range(exit_year + 2):  # +2 to cover Year 0 through Exit Year plus some buffer
        col_letter = get_column_letter(3 + i)
        ws.column_dimensions[col_letter].width = 22

    current_row = 1

    # ==========================================================================
    # RETURNS SUMMARY (headline numbers at top)
    # ==========================================================================

    ws.cell(row=current_row, column=2, value="RETURNS SUMMARY")
    ws.cell(row=current_row, column=2).font = SECTION_HEADER_FONT
    ws.cell(row=current_row, column=2).fill = SECTION_HEADER_FILL
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)
    current_row += 2

    # IRR (large, prominent)
    irr_display_row = current_row
    ws.cell(row=current_row, column=2, value="IRR").font = Font(bold=True, size=14)
    irr_display_cell = ws.cell(row=current_row, column=3)
    # This will reference the calculated IRR cell below (we'll set row number after we know it)
    # For now, leave as placeholder - will update after IRR row is created
    current_row += 1

    # MOIC (large, prominent)
    moic_display_row = current_row
    ws.cell(row=current_row, column=2, value="MOIC").font = Font(bold=True, size=14)
    moic_display_cell = ws.cell(row=current_row, column=3)
    # This will reference the calculated MOIC cell below
    current_row += 2

    # ==========================================================================
    # EXIT VALUATION SECTION
    # ==========================================================================

    ws.cell(row=current_row, column=2, value="Exit Valuation").font = SUBSECTION_FONT
    ws.cell(row=current_row, column=2).fill = SUBSECTION_FILL
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)
    current_row += 1

    # Exit Year EBITDA - direct reference to Operating Model
    exit_ebitda_row = current_row
    ws.cell(row=current_row, column=2, value="Exit Year EBITDA").font = BLACK_FONT
    cell = ws.cell(row=current_row, column=3)
    # DIRECT cell reference - no named range
    cell.value = f"='Operating Model'!{exit_col}{ebitda_row}"
    cell.font = BLACK_FONT
    cell.number_format = CURRENCY_FORMAT
    current_row += 1

    # Exit Multiple - direct reference to Assumptions
    exit_multiple_row = current_row
    ws.cell(row=current_row, column=2, value="Exit EV/EBITDA Multiple").font = BLACK_FONT
    cell = ws.cell(row=current_row, column=3)
    # DIRECT cell reference - Assumptions C39 is Exit Multiple
    cell.value = "='Assumptions'!$C$39"
    cell.font = BLACK_FONT
    cell.number_format = MULTIPLE_FORMAT
    current_row += 1

    # Exit Enterprise Value = Exit EBITDA × Exit Multiple
    exit_ev_row = current_row
    ws.cell(row=current_row, column=2, value="Exit Enterprise Value").font = BLACK_FONT
    cell = ws.cell(row=current_row, column=3)
    cell.value = f"=C{exit_ebitda_row}*C{exit_multiple_row}"
    cell.font = BLACK_FONT
    cell.number_format = CURRENCY_FORMAT
    current_row += 1

    # Exit Debt Balance - direct reference to Debt Schedule
    exit_debt_row = current_row
    ws.cell(row=current_row, column=2, value="Exit Debt Balance").font = BLACK_FONT
    cell = ws.cell(row=current_row, column=3)
    # DIRECT cell reference - no named range
    cell.value = f"='Debt Schedule'!{exit_col}{ending_balance_row}"
    cell.font = BLACK_FONT
    cell.number_format = CURRENCY_FORMAT
    current_row += 1

    # Exit Equity Value = Exit EV - Exit Debt
    exit_equity_row = current_row
    ws.cell(row=current_row, column=2, value="Exit Equity Value").font = BLACK_FONT_BOLD
    cell = ws.cell(row=current_row, column=3)
    cell.value = f"=C{exit_ev_row}-C{exit_debt_row}"
    cell.font = BLACK_FONT_BOLD
    cell.number_format = CURRENCY_FORMAT
    cell.border = Border(top=Side(style='thin'), bottom=Side(style='double'))
    current_row += 2

    # ==========================================================================
    # ENTRY INVESTMENT SECTION
    # ==========================================================================

    ws.cell(row=current_row, column=2, value="Entry Investment").font = SUBSECTION_FONT
    ws.cell(row=current_row, column=2).fill = SUBSECTION_FILL
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)
    current_row += 1

    # Sponsor Equity (Year 0 investment) - direct reference to Sources & Uses
    sponsor_equity_row = current_row
    ws.cell(row=current_row, column=2, value="Sponsor Equity (Year 0)").font = BLACK_FONT
    cell = ws.cell(row=current_row, column=3)
    # DIRECT cell reference - Sources & Uses C{equity_row}
    cell.value = f"='Sources & Uses'!$C${equity_row}"
    cell.font = BLACK_FONT
    cell.number_format = CURRENCY_FORMAT
    current_row += 2

    # ==========================================================================
    # IRR CALCULATION (using XIRR)
    # ==========================================================================

    ws.cell(row=current_row, column=2, value="IRR Calculation").font = SUBSECTION_FONT
    ws.cell(row=current_row, column=2).fill = SUBSECTION_FILL
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3 + exit_year)
    current_row += 1

    # Date row for XIRR - Year 0 through Exit Year
    dates_row = current_row
    ws.cell(row=current_row, column=2, value="Dates").font = BLACK_FONT
    for year in range(0, exit_year + 1):
        col = 3 + year  # Column C = Year 0, D = Year 1, etc.
        cell = ws.cell(row=current_row, column=col)
        # Use DATE function: entry date is end of most recent FY, then add years
        # DATE(year, month, day) - assume fiscal year end is Dec 31 for simplicity
        entry_year = most_recent_fy
        cell.value = f"=DATE({entry_year + year},12,31)"
        cell.font = BLACK_FONT
        cell.number_format = "YYYY-MM-DD"
    current_row += 1

    # Cash flow row for XIRR
    cashflows_row = current_row
    ws.cell(row=current_row, column=2, value="Cash Flows").font = BLACK_FONT
    for year in range(0, exit_year + 1):
        col = 3 + year
        col_letter = get_column_letter(col)
        cell = ws.cell(row=current_row, column=col)

        if year == 0:
            # Year 0: Negative investment (outflow)
            cell.value = f"=-C{sponsor_equity_row}"
        elif year == exit_year:
            # Exit Year: Positive proceeds (inflow)
            cell.value = f"=C{exit_equity_row}"
        else:
            # Interim years: No cash flow (hold period)
            cell.value = 0

        cell.font = BLACK_FONT
        cell.number_format = CURRENCY_FORMAT
    current_row += 1

    # IRR using XIRR
    irr_row = current_row
    ws.cell(row=current_row, column=2, value="IRR (XIRR)").font = BLACK_FONT_BOLD
    cell = ws.cell(row=current_row, column=3)
    # XIRR(values, dates, [guess])
    # Values range: C{cashflows_row} to {exit_col_for_cf}{cashflows_row}
    exit_cf_col = get_column_letter(3 + exit_year)
    cell.value = f"=XIRR(C{cashflows_row}:{exit_cf_col}{cashflows_row},C{dates_row}:{exit_cf_col}{dates_row})"
    cell.font = BLACK_FONT_BOLD
    cell.number_format = PERCENT_FORMAT
    cell.border = Border(top=Side(style='thin'), bottom=Side(style='double'))
    current_row += 2

    # ==========================================================================
    # MOIC CALCULATION
    # ==========================================================================

    ws.cell(row=current_row, column=2, value="MOIC Calculation").font = SUBSECTION_FONT
    ws.cell(row=current_row, column=2).fill = SUBSECTION_FILL
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)
    current_row += 1

    # MOIC = Exit Equity / Sponsor Equity
    moic_row = current_row
    ws.cell(row=current_row, column=2, value="MOIC").font = BLACK_FONT_BOLD
    cell = ws.cell(row=current_row, column=3)
    cell.value = f"=IF(C{sponsor_equity_row}=0,0,C{exit_equity_row}/C{sponsor_equity_row})"
    cell.font = BLACK_FONT_BOLD
    cell.number_format = MULTIPLE_FORMAT
    cell.border = Border(top=Side(style='thin'), bottom=Side(style='double'))

    # ==========================================================================
    # UPDATE SUMMARY DISPLAY CELLS
    # ==========================================================================

    # Now update the summary cells at the top to reference the calculated values
    irr_display_cell.value = f"=C{irr_row}"
    irr_display_cell.font = Font(bold=True, size=14)
    irr_display_cell.number_format = PERCENT_FORMAT

    moic_display_cell.value = f"=C{moic_row}"
    moic_display_cell.font = Font(bold=True, size=14)
    moic_display_cell.number_format = MULTIPLE_FORMAT

    # Return row references
    return {
        'exit_ebitda_row': exit_ebitda_row,
        'exit_ev_row': exit_ev_row,
        'exit_debt_row': exit_debt_row,
        'exit_equity_row': exit_equity_row,
        'sponsor_equity_row': sponsor_equity_row,
        'irr_row': irr_row,
        'moic_row': moic_row,
        'dates_row': dates_row,
        'cashflows_row': cashflows_row,
    }


# =============================================================================
# SENSITIVITY TAB BUILDER
# =============================================================================

# Base case highlight fill (light green)
BASE_CASE_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')


def build_sensitivity_tab(wb: Workbook, summary: dict, sources_uses_refs: dict,
                          op_model_refs: dict, debt_schedule_refs: dict,
                          returns_refs: dict) -> dict:
    """
    Build the Sensitivity tab with Entry Multiple x Exit Multiple grids.

    Creates two 5x5 grids showing MOIC and IRR for different combinations of
    Entry and Exit multiples, centered on the base case assumptions.

    Key design choices:
    - NO Excel Data Tables (unreliable cross-platform)
    - Each cell contains a fully self-contained formula
    - All cross-sheet references are DIRECT cell references (no named ranges)
    - Uses simplified IRR: (ExitEquity/SponsorEquity)^(1/ExitYear)-1

    For each (Entry Multiple, Exit Multiple) combination:
    - Sponsor Equity = (EntryEBITDA × EntryMult) × (1 + TxnFee%) - NewDebt
    - Exit Equity = (ExitYearEBITDA × ExitMult) - ExitDebt
    - MOIC = Exit Equity / Sponsor Equity
    - IRR = MOIC^(1/ExitYear) - 1

    Note: New Debt and Exit Debt are fixed (based on Leverage Multiple and the
    original deal's FCF trajectory), so they reference the actual Sources & Uses
    and Debt Schedule tabs.
    """
    ws = wb.create_sheet("Sensitivity", 5)

    exit_year = op_model_refs['exit_year']
    ebitda_row = op_model_refs['ebitda_row']
    ending_balance_row = debt_schedule_refs['ending_balance_row']
    debt_row = sources_uses_refs['debt_row']

    # Exit year column in Operating Model / Debt Schedule
    exit_col = get_column_letter(2 + exit_year)

    # Column widths
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 22
    for i in range(5):  # 5 columns for the grid
        col_letter = get_column_letter(3 + i)
        ws.column_dimensions[col_letter].width = 14

    # Gap column between MOIC and IRR grids
    ws.column_dimensions['H'].width = 5

    # IRR grid columns
    ws.column_dimensions['I'].width = 22
    for i in range(5):
        col_letter = get_column_letter(10 + i)
        ws.column_dimensions[col_letter].width = 14

    current_row = 1

    # ==========================================================================
    # SECTION HEADER
    # ==========================================================================

    ws.cell(row=current_row, column=2, value="SENSITIVITY ANALYSIS")
    ws.cell(row=current_row, column=2).font = SECTION_HEADER_FONT
    ws.cell(row=current_row, column=2).fill = SECTION_HEADER_FILL
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=7)

    ws.cell(row=current_row, column=9, value="SENSITIVITY ANALYSIS")
    ws.cell(row=current_row, column=9).font = SECTION_HEADER_FONT
    ws.cell(row=current_row, column=9).fill = SECTION_HEADER_FILL
    ws.merge_cells(start_row=current_row, start_column=9, end_row=current_row, end_column=14)
    current_row += 2

    # ==========================================================================
    # GRID LABELS
    # ==========================================================================

    # MOIC grid label
    ws.cell(row=current_row, column=2, value="MOIC").font = BLACK_FONT_BOLD
    # IRR grid label
    ws.cell(row=current_row, column=9, value="IRR").font = BLACK_FONT_BOLD
    current_row += 1

    # Sub-labels
    ws.cell(row=current_row, column=2, value="Exit Multiple →").font = BLACK_FONT
    ws.cell(row=current_row, column=9, value="Exit Multiple →").font = BLACK_FONT

    # ==========================================================================
    # BUILD SENSITIVITY GRIDS
    # ==========================================================================

    # The grid will be:
    # - Rows: Entry Multiple variations (5 values: base-1, base-0.5, base, base+0.5, base+1)
    # - Columns: Exit Multiple variations (same pattern)
    # - Center cell = base case (should match Returns tab)

    # Multiple step size: 0.5x increments
    # Grid offsets from base case: [-1.0, -0.5, 0, +0.5, +1.0]
    offsets = [-1.0, -0.5, 0.0, 0.5, 1.0]

    # Row where grid starts (after labels)
    grid_start_row = current_row + 1

    # Column header row (Exit Multiple values)
    header_row = grid_start_row
    for col_idx, offset in enumerate(offsets):
        # MOIC grid header
        moic_col = 3 + col_idx
        cell = ws.cell(row=header_row, column=moic_col)
        # Formula: base Exit Multiple + offset
        # Exit Multiple is at 'Assumptions'!$C$39
        if offset == 0:
            cell.value = "='Assumptions'!$C$39"
        elif offset > 0:
            cell.value = f"='Assumptions'!$C$39+{offset}"
        else:
            cell.value = f"='Assumptions'!$C$39{offset}"  # offset is negative, so minus sign automatic
        cell.font = BLACK_FONT_BOLD
        cell.number_format = MULTIPLE_FORMAT
        cell.alignment = CENTER_ALIGN

        # IRR grid header (same values)
        irr_col = 10 + col_idx
        cell = ws.cell(row=header_row, column=irr_col)
        if offset == 0:
            cell.value = "='Assumptions'!$C$39"
        elif offset > 0:
            cell.value = f"='Assumptions'!$C$39+{offset}"
        else:
            cell.value = f"='Assumptions'!$C$39{offset}"
        cell.font = BLACK_FONT_BOLD
        cell.number_format = MULTIPLE_FORMAT
        cell.alignment = CENTER_ALIGN

    # Row labels column (Entry Multiple values) and grid cells
    for row_idx, entry_offset in enumerate(offsets):
        data_row = grid_start_row + 1 + row_idx

        # Row label for MOIC grid (Entry Multiple value)
        label_cell = ws.cell(row=data_row, column=2)
        if entry_offset == 0:
            label_cell.value = "='Assumptions'!$C$30"
        elif entry_offset > 0:
            label_cell.value = f"='Assumptions'!$C$30+{entry_offset}"
        else:
            label_cell.value = f"='Assumptions'!$C$30{entry_offset}"
        label_cell.font = BLACK_FONT_BOLD
        label_cell.number_format = MULTIPLE_FORMAT
        label_cell.alignment = RIGHT_ALIGN

        # Row label for IRR grid (Entry Multiple value)
        irr_label_cell = ws.cell(row=data_row, column=9)
        if entry_offset == 0:
            irr_label_cell.value = "='Assumptions'!$C$30"
        elif entry_offset > 0:
            irr_label_cell.value = f"='Assumptions'!$C$30+{entry_offset}"
        else:
            irr_label_cell.value = f"='Assumptions'!$C$30{entry_offset}"
        irr_label_cell.font = BLACK_FONT_BOLD
        irr_label_cell.number_format = MULTIPLE_FORMAT
        irr_label_cell.alignment = RIGHT_ALIGN

        # Fill in grid cells for this row
        for col_idx, exit_offset in enumerate(offsets):
            moic_col = 3 + col_idx
            irr_col = 10 + col_idx

            # Build the Entry Multiple expression for this cell
            if entry_offset == 0:
                entry_mult_expr = "'Assumptions'!$C$30"
            elif entry_offset > 0:
                entry_mult_expr = f"('Assumptions'!$C$30+{entry_offset})"
            else:
                entry_mult_expr = f"('Assumptions'!$C$30{entry_offset})"

            # Build the Exit Multiple expression for this cell
            if exit_offset == 0:
                exit_mult_expr = "'Assumptions'!$C$39"
            elif exit_offset > 0:
                exit_mult_expr = f"('Assumptions'!$C$39+{exit_offset})"
            else:
                exit_mult_expr = f"('Assumptions'!$C$39{exit_offset})"

            # ============================================================
            # MOIC Formula
            # ============================================================
            # MOIC = Exit Equity / Sponsor Equity
            #
            # Sponsor Equity = (EntryEBITDA × EntryMult) × (1 + TxnFee%) - NewDebt
            # Exit Equity = (ExitYearEBITDA × ExitMult) - ExitDebt
            #
            # Cell references:
            # - EntryEBITDA: 'Assumptions'!$C$18
            # - TxnFee%: 'Assumptions'!$C$40
            # - NewDebt: 'Sources & Uses'!$C${debt_row}
            # - ExitYearEBITDA: 'Operating Model'!{exit_col}${ebitda_row}
            # - ExitDebt: 'Debt Schedule'!{exit_col}${ending_balance_row}

            sponsor_equity_formula = (
                f"('Assumptions'!$C$18*{entry_mult_expr})*(1+'Assumptions'!$C$40)"
                f"-'Sources & Uses'!$C${debt_row}"
            )

            exit_equity_formula = (
                f"('Operating Model'!{exit_col}${ebitda_row}*{exit_mult_expr})"
                f"-'Debt Schedule'!{exit_col}${ending_balance_row}"
            )

            moic_formula = f"=IF({sponsor_equity_formula}<=0,0,({exit_equity_formula})/({sponsor_equity_formula}))"

            moic_cell = ws.cell(row=data_row, column=moic_col)
            moic_cell.value = moic_formula
            moic_cell.font = BLACK_FONT
            moic_cell.number_format = MULTIPLE_FORMAT
            moic_cell.alignment = CENTER_ALIGN

            # Highlight base case (center cell where both offsets are 0)
            if entry_offset == 0 and exit_offset == 0:
                moic_cell.fill = BASE_CASE_FILL
                moic_cell.font = BLACK_FONT_BOLD

            # ============================================================
            # IRR Formula
            # ============================================================
            # IRR = MOIC^(1/ExitYear) - 1
            # Since no interim distributions, this simplified formula works
            #
            # ExitYear: 'Assumptions'!$C$38

            irr_formula = (
                f"=IF({sponsor_equity_formula}<=0,0,"
                f"(({exit_equity_formula})/({sponsor_equity_formula}))^(1/'Assumptions'!$C$38)-1)"
            )

            irr_cell = ws.cell(row=data_row, column=irr_col)
            irr_cell.value = irr_formula
            irr_cell.font = BLACK_FONT
            irr_cell.number_format = PERCENT_FORMAT
            irr_cell.alignment = CENTER_ALIGN

            # Highlight base case
            if entry_offset == 0 and exit_offset == 0:
                irr_cell.fill = BASE_CASE_FILL
                irr_cell.font = BLACK_FONT_BOLD

    # Update current_row past the grid
    current_row = grid_start_row + 1 + len(offsets) + 2

    # ==========================================================================
    # LEGEND / NOTES
    # ==========================================================================

    ws.cell(row=current_row, column=2, value="Notes:").font = BLACK_FONT_BOLD
    current_row += 1

    ws.cell(row=current_row, column=2,
            value="• Green highlighted cell = Base Case (matches Returns tab)").font = BLACK_FONT
    current_row += 1

    ws.cell(row=current_row, column=2,
            value="• Entry/Exit Multiple steps: ±0.5x, ±1.0x from base case").font = BLACK_FONT
    current_row += 1

    ws.cell(row=current_row, column=2,
            value="• IRR uses simplified formula: MOIC^(1/Years)-1 (no interim distributions)").font = BLACK_FONT
    current_row += 1

    # Return refs for potential future use
    return {
        'grid_start_row': grid_start_row,
        'moic_start_col': 3,
        'irr_start_col': 10,
    }


# =============================================================================
# WIRE UP INTEREST EXPENSE (after Debt Schedule is built)
# =============================================================================

def wire_interest_expense(wb: Workbook, op_model_refs: dict, debt_schedule_refs: dict):
    """
    Update Operating Model's Interest Expense row to reference Debt Schedule.

    This is called after both tabs are built to wire up the cross-reference.
    Removes the placeholder styling (orange fill) and replaces hardcoded 0 with
    a formula referencing the Debt Schedule's Interest Expense row.

    Dependency chain verification (no circular reference):
    1. Debt Schedule Beginning Balance (Year t) = NewDebtRaised (Year 1) or prior Ending Balance
    2. Debt Schedule Interest Expense = Beginning Balance × InterestRate
    3. Operating Model Interest Expense = Debt Schedule Interest Expense (same year)
    4. Operating Model FCF = ... - Interest Expense - ...
    5. Debt Schedule Cash Available = Operating Model FCF
    6. Debt Schedule Ending Balance = Beginning - Mandatory - Optional Paydown
    7. NEXT year's Beginning Balance = This year's Ending Balance

    The key insight: Interest for Year t depends on Beginning Balance for Year t,
    which is fully determined before Year t's FCF is calculated. The dependency
    flows one direction: Beginning -> Interest -> FCF -> Ending -> Next Beginning.
    """
    ws = wb['Operating Model']

    exit_year = op_model_refs['exit_year']
    interest_row = op_model_refs['interest_row']
    debt_interest_row = debt_schedule_refs['interest_expense_row']

    for year in range(1, exit_year + 1):
        col = 2 + year  # Column C = Year 1, D = Year 2, etc.
        col_letter = get_column_letter(col)
        cell = ws.cell(row=interest_row, column=col)

        # Replace placeholder with formula referencing Debt Schedule
        cell.value = f"='Debt Schedule'!{col_letter}{debt_interest_row}"
        cell.font = BLACK_FONT
        cell.number_format = CURRENCY_FORMAT
        # Remove placeholder fill (set to no fill)
        cell.fill = PatternFill(fill_type=None)
        # Remove placeholder comment
        cell.comment = None


# =============================================================================
# MAIN GENERATOR FUNCTION
# =============================================================================

def generate_workbook(validated_summary: dict, output_dir: str = None) -> str:
    """
    Generate an LBO model Excel workbook from a validated summary.

    Args:
        validated_summary: Dict containing both the summary data and validation result.
                          Expected structure:
                          {
                              'summary': {...},  # From fetch_ticker_data()
                              'validation': {...}  # From validate()
                          }
                          OR just the summary dict directly (validation defaults to empty)
        output_dir: Directory to save the file (defaults to current directory)

    Returns:
        Path to the generated .xlsx file
    """
    # Handle both formats: wrapped or direct summary
    if 'summary' in validated_summary and 'validation' in validated_summary:
        summary = validated_summary['summary']
        validation = validated_summary['validation']
    else:
        summary = validated_summary
        validation = {}

    # Create workbook
    wb = Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # Build Assumptions tab
    named_ranges = build_assumptions_tab(wb, summary, validation)

    # Build Sources & Uses tab
    su_refs = build_sources_uses_tab(wb, named_ranges)

    # Build Operating Model tab
    op_model_refs = build_operating_model_tab(wb, summary, named_ranges)

    # Build Debt Schedule tab
    debt_schedule_refs = build_debt_schedule_tab(wb, summary, op_model_refs)

    # Wire up Operating Model's Interest Expense to reference Debt Schedule
    # (replaces the placeholder 0 values with actual formulas)
    wire_interest_expense(wb, op_model_refs, debt_schedule_refs)

    # Build Returns tab
    returns_refs = build_returns_tab(wb, summary, su_refs, op_model_refs, debt_schedule_refs)

    # Build Sensitivity tab
    sensitivity_refs = build_sensitivity_tab(wb, summary, su_refs, op_model_refs, debt_schedule_refs, returns_refs)

    # Generate filename
    ticker = summary.get('ticker', 'UNKNOWN')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"LBO_{ticker}_{timestamp}.xlsx"

    # Determine output path
    if output_dir:
        output_path = Path(output_dir) / filename
    else:
        output_path = Path.cwd() / filename

    # Save workbook
    wb.save(str(output_path))

    print(f"Generated workbook: {output_path}")

    return str(output_path)


# =============================================================================
# FORMULA VERIFICATION (since openpyxl doesn't evaluate formulas)
# =============================================================================

def verify_workbook_formulas(filepath: str) -> dict:
    """
    Load a generated workbook and print the formula strings for verification.

    Since openpyxl writes formulas as text and can't evaluate them,
    this function reads back the formulas so you can spot-check the logic.

    Returns a dict with key formula strings.
    """
    wb = load_workbook(filepath)

    result = {
        'filepath': filepath,
        'sheets': list(wb.sheetnames),
        'assumptions_formulas': {},
        'sources_uses_formulas': {},
        'operating_model_formulas': {},
        'debt_schedule_formulas': {},
        'returns_formulas': {},
        'named_ranges': {}
    }

    # Get named ranges
    for name in wb.defined_names:
        defn = wb.defined_names[name]
        result['named_ranges'][name] = str(defn.attr_text)

    # Check Assumptions tab
    if 'Assumptions' in wb.sheetnames:
        ws = wb['Assumptions']
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    cell_ref = f"{get_column_letter(cell.column)}{cell.row}"
                    result['assumptions_formulas'][cell_ref] = cell.value

    # Check Sources & Uses tab
    if 'Sources & Uses' in wb.sheetnames:
        ws = wb['Sources & Uses']
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    cell_ref = f"{get_column_letter(cell.column)}{cell.row}"
                    result['sources_uses_formulas'][cell_ref] = cell.value

    # Check Operating Model tab
    if 'Operating Model' in wb.sheetnames:
        ws = wb['Operating Model']
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    cell_ref = f"{get_column_letter(cell.column)}{cell.row}"
                    result['operating_model_formulas'][cell_ref] = cell.value

    # Check Debt Schedule tab
    if 'Debt Schedule' in wb.sheetnames:
        ws = wb['Debt Schedule']
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    cell_ref = f"{get_column_letter(cell.column)}{cell.row}"
                    result['debt_schedule_formulas'][cell_ref] = cell.value

    # Check Returns tab
    if 'Returns' in wb.sheetnames:
        ws = wb['Returns']
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    cell_ref = f"{get_column_letter(cell.column)}{cell.row}"
                    result['returns_formulas'][cell_ref] = cell.value

    return result


def print_verification_report(verification: dict):
    """Print a human-readable verification report."""
    print(f"\n{'='*70}")
    print(f"FORMULA VERIFICATION REPORT")
    print(f"{'='*70}")
    print(f"File: {verification['filepath']}")
    print(f"Sheets: {', '.join(verification['sheets'])}")

    print(f"\n--- Named Ranges ({len(verification['named_ranges'])}) ---")
    for name, ref in sorted(verification['named_ranges'].items()):
        print(f"  {name}: {ref}")

    print(f"\n--- Assumptions Tab Formulas ---")
    for cell, formula in verification['assumptions_formulas'].items():
        print(f"  {cell}: {formula}")

    print(f"\n--- Sources & Uses Tab Formulas ---")
    for cell, formula in verification['sources_uses_formulas'].items():
        print(f"  {cell}: {formula}")

    print(f"\n--- Returns Tab Formulas ---")
    for cell, formula in verification.get('returns_formulas', {}).items():
        print(f"  {cell}: {formula}")

    # Key formulas to verify
    print(f"\n--- KEY FORMULAS TO VERIFY ---")
    su_formulas = verification['sources_uses_formulas']

    print("\nExpected logic:")
    print("  Purchase EV = EntryEBITDA × EntryMultiple")
    print("  Transaction Fees = Purchase EV × TransactionFeePct")
    print("  Total Uses = Purchase EV + Fees")
    print("  New Debt = EntryEBITDA × LeverageMultiple")
    print("  Sponsor Equity = Total Uses - New Debt")
    print("  Total Sources = New Debt + Sponsor Equity")
    print("  Balance Check = Total Sources - Total Uses (should = 0)")
    print("\nReturns logic:")
    print("  Exit EV = Exit Year EBITDA × Exit Multiple")
    print("  Equity at Exit = Exit EV - Ending Debt")
    print("  IRR = XIRR of cash flows (negative equity at entry, positive at exit)")
    print("  MOIC = Equity at Exit / Sponsor Equity")


def evaluate_model_values(summary: dict, validation: dict) -> dict:
    """
    Evaluate the key model values manually to verify formula logic.

    Since openpyxl can't evaluate Excel formulas, this function computes
    the expected values using the same logic as the formulas. This allows
    verification without opening Excel.

    Returns a dict with computed values.
    """
    fiscal_years = summary.get('fiscal_years', {})
    if not fiscal_years:
        return {"error": "No fiscal year data"}

    most_recent_fy = max(fiscal_years.keys())
    fy_data = fiscal_years[most_recent_fy]

    # Get Entry EBITDA (most recent FY)
    ebitda = fy_data.get('ebitda_calculated')
    if ebitda is None:
        op_inc = fy_data.get('operating_income', 0) or 0
        da = fy_data.get('da', 0) or 0
        ebitda = op_inc + da

    # Deal assumptions (defaults from build_assumptions_tab)
    entry_multiple = 8.0
    leverage_multiple = 5.5
    transaction_fee_pct = 0.02

    # Computed values
    purchase_ev = ebitda * entry_multiple
    transaction_fees = purchase_ev * transaction_fee_pct
    total_uses = purchase_ev + transaction_fees
    new_debt = ebitda * leverage_multiple
    sponsor_equity = total_uses - new_debt
    total_sources = new_debt + sponsor_equity
    balance_check = total_sources - total_uses

    return {
        "ticker": summary.get('ticker', 'Unknown'),
        "most_recent_fy": most_recent_fy,
        "entry_ebitda": ebitda,
        "entry_multiple": entry_multiple,
        "leverage_multiple": leverage_multiple,
        "transaction_fee_pct": transaction_fee_pct,
        "purchase_ev": purchase_ev,
        "transaction_fees": transaction_fees,
        "total_uses": total_uses,
        "new_debt": new_debt,
        "sponsor_equity": sponsor_equity,
        "total_sources": total_sources,
        "balance_check": balance_check,
        "balance_ok": abs(balance_check) < 0.01  # Should be exactly 0
    }


def print_computed_values(values: dict):
    """Print the computed model values in a readable format."""
    if "error" in values:
        print(f"  ERROR: {values['error']}")
        return

    def fmt_currency(val):
        if val >= 1e9:
            return f"${val/1e9:,.2f}B"
        elif val >= 1e6:
            return f"${val/1e6:,.2f}M"
        else:
            return f"${val:,.0f}"

    print(f"\n  Computed Values for {values['ticker']} (FY{values['most_recent_fy']}):")
    print(f"  {'-'*50}")
    print(f"  Entry EBITDA:      {fmt_currency(values['entry_ebitda'])}")
    print(f"  Entry Multiple:    {values['entry_multiple']:.1f}x")
    print(f"  Leverage Multiple: {values['leverage_multiple']:.1f}x")
    print(f"  Transaction Fee %: {values['transaction_fee_pct']*100:.1f}%")
    print(f"  {'-'*50}")
    print(f"  USES:")
    print(f"    Purchase EV:     {fmt_currency(values['purchase_ev'])}")
    print(f"    Transaction Fees:{fmt_currency(values['transaction_fees'])}")
    print(f"    Total Uses:      {fmt_currency(values['total_uses'])}")
    print(f"  {'-'*50}")
    print(f"  SOURCES:")
    print(f"    New Debt:        {fmt_currency(values['new_debt'])}")
    print(f"    Sponsor Equity:  {fmt_currency(values['sponsor_equity'])}")
    print(f"    Total Sources:   {fmt_currency(values['total_sources'])}")
    print(f"  {'-'*50}")
    print(f"  Balance Check:     {fmt_currency(values['balance_check'])} {'[OK]' if values['balance_ok'] else '[ERROR]'}")


def evaluate_operating_model(summary: dict, validation: dict, exit_year: int = 5) -> dict:
    """
    Evaluate the Operating Model projections manually to verify formula logic.

    Returns a dict with projected values for each year.
    """
    fiscal_years = summary.get('fiscal_years', {})
    if not fiscal_years:
        return {"error": "No fiscal year data"}

    most_recent_fy = max(fiscal_years.keys())
    fy_data = fiscal_years[most_recent_fy]

    # Entry values
    entry_revenue = fy_data.get('revenue', 0) or 0
    entry_ebitda = fy_data.get('ebitda_calculated')
    if entry_ebitda is None:
        op_inc = fy_data.get('operating_income', 0) or 0
        da = fy_data.get('da', 0) or 0
        entry_ebitda = op_inc + da

    # Calculate growth rate using the same logic as the Excel generator
    revenue_growth, growth_is_fallback, _ = calculate_revenue_growth_rate(summary)

    # Assumptions (defaults)
    ebitda_margin = entry_ebitda / entry_revenue if entry_revenue > 0 else 0.15
    da_pct = (fy_data.get('da', 0) or 0) / entry_revenue if entry_revenue > 0 else 0.03
    capex_pct = (fy_data.get('capex', 0) or 0) / entry_revenue if entry_revenue > 0 else 0.02
    nwc_pct = 0.0
    tax_rate = 0.25
    interest_expense = 0  # Placeholder

    projections = []
    prev_revenue = entry_revenue

    for year in range(1, exit_year + 1):
        # Revenue
        revenue = prev_revenue * (1 + revenue_growth)

        # EBITDA
        ebitda = revenue * ebitda_margin

        # D&A
        da = revenue * da_pct

        # EBIT
        ebit = ebitda - da

        # Pre-Tax Income
        pretax = ebit - interest_expense

        # Taxes (floored at 0)
        taxes = max(0, pretax * tax_rate)

        # Net Income
        net_income = pretax - taxes

        # FCF components
        capex = revenue * capex_pct
        nwc_change = revenue * nwc_pct

        # FCF
        fcf = net_income + da - capex - nwc_change

        projections.append({
            'year': year,
            'calendar_year': most_recent_fy + year,
            'revenue': revenue,
            'ebitda': ebitda,
            'da': da,
            'ebit': ebit,
            'interest': interest_expense,
            'pretax': pretax,
            'taxes': taxes,
            'net_income': net_income,
            'capex': capex,
            'nwc_change': nwc_change,
            'fcf': fcf,
        })

        prev_revenue = revenue

    return {
        'ticker': summary.get('ticker', 'Unknown'),
        'most_recent_fy': most_recent_fy,
        'entry_revenue': entry_revenue,
        'entry_ebitda': entry_ebitda,
        'revenue_growth': revenue_growth,
        'growth_is_fallback': growth_is_fallback,
        'ebitda_margin': ebitda_margin,
        'da_pct': da_pct,
        'capex_pct': capex_pct,
        'tax_rate': tax_rate,
        'projections': projections,
    }


def print_operating_model(op_model: dict):
    """Print Operating Model projections in a readable format."""
    if "error" in op_model:
        print(f"  ERROR: {op_model['error']}")
        return

    def fmt_currency(val):
        if abs(val) >= 1e9:
            return f"${val/1e9:,.2f}B"
        elif abs(val) >= 1e6:
            return f"${val/1e6:,.2f}M"
        else:
            return f"${val:,.0f}"

    print(f"\n  Operating Model for {op_model['ticker']}:")
    print(f"  Entry FY: {op_model['most_recent_fy']}")
    print(f"  Entry Revenue: {fmt_currency(op_model['entry_revenue'])}")
    growth_flag = " [FALLBACK - implausible calc]" if op_model.get('growth_is_fallback') else ""
    print(f"  Revenue Growth: {op_model['revenue_growth']*100:.1f}%{growth_flag}")
    print(f"  EBITDA Margin: {op_model['ebitda_margin']*100:.1f}%")
    print()

    # Print header
    years = op_model['projections']
    header = f"  {'Metric':<20}"
    for p in [years[0], years[-1]]:  # Year 1 and final year
        header += f"  Year {p['year']:>3}"
    print(header)
    print("  " + "-" * (20 + 12 * 2))

    # Print key rows
    metrics = [
        ('Revenue', 'revenue'),
        ('EBITDA', 'ebitda'),
        ('D&A', 'da'),
        ('EBIT', 'ebit'),
        ('Interest', 'interest'),
        ('Pre-Tax', 'pretax'),
        ('Taxes', 'taxes'),
        ('Net Income', 'net_income'),
        ('CapEx', 'capex'),
        ('NWC Change', 'nwc_change'),
        ('FCF', 'fcf'),
    ]

    for label, key in metrics:
        row = f"  {label:<20}"
        for p in [years[0], years[-1]]:
            row += f"  {fmt_currency(p[key]):>10}"
        print(row)

    # FCF conversion check
    year1 = years[0]
    final = years[-1]
    print()
    print(f"  FCF/EBITDA Year 1: {year1['fcf']/year1['ebitda']*100:.1f}%")
    print(f"  FCF/EBITDA Year {final['year']}: {final['fcf']/final['ebitda']*100:.1f}%")


# =============================================================================
# INTEGRATED DEBT SCHEDULE + OPERATING MODEL EVALUATION
# =============================================================================

def evaluate_full_model(summary: dict, validation: dict, exit_year: int = 5) -> dict:
    """
    Evaluate the full integrated model including Operating Model and Debt Schedule.

    This computes the actual values accounting for the feedback between:
    - Debt Schedule's Interest Expense (depends on beginning balance)
    - Operating Model's FCF (depends on interest expense)
    - Debt Schedule's paydown (depends on FCF)

    Since interest is calculated on BEGINNING balance (not average), there's no
    true circularity - each year's values can be computed sequentially.

    Returns a dict with both operating model projections and debt schedule values.
    """
    fiscal_years = summary.get('fiscal_years', {})
    if not fiscal_years:
        return {"error": "No fiscal year data"}

    most_recent_fy = max(fiscal_years.keys())
    fy_data = fiscal_years[most_recent_fy]

    # Entry values
    entry_revenue = fy_data.get('revenue', 0) or 0
    entry_ebitda = fy_data.get('ebitda_calculated')
    if entry_ebitda is None:
        op_inc = fy_data.get('operating_income', 0) or 0
        da = fy_data.get('da', 0) or 0
        entry_ebitda = op_inc + da

    # Calculate growth rate
    revenue_growth, growth_is_fallback, _ = calculate_revenue_growth_rate(summary)

    # Assumptions (defaults from build_assumptions_tab)
    ebitda_margin = entry_ebitda / entry_revenue if entry_revenue > 0 else 0.15
    da_pct = (fy_data.get('da', 0) or 0) / entry_revenue if entry_revenue > 0 else 0.03
    capex_pct = (fy_data.get('capex', 0) or 0) / entry_revenue if entry_revenue > 0 else 0.02
    nwc_pct = 0.0
    tax_rate = 0.25
    interest_rate = 0.08
    leverage_multiple = 5.5
    mandatory_amort_pct = 0.01

    # Calculate new debt raised
    new_debt_raised = entry_ebitda * leverage_multiple

    projections = []
    debt_schedule = []
    prev_revenue = entry_revenue
    beginning_balance = new_debt_raised

    for year in range(1, exit_year + 1):
        # =====================================================================
        # DEBT SCHEDULE - Interest (based on BEGINNING balance - no circularity)
        # =====================================================================
        interest_expense = beginning_balance * interest_rate
        mandatory_amort = new_debt_raised * mandatory_amort_pct

        # =====================================================================
        # OPERATING MODEL
        # =====================================================================
        # Revenue
        revenue = prev_revenue * (1 + revenue_growth)

        # EBITDA
        ebitda = revenue * ebitda_margin

        # D&A
        da = revenue * da_pct

        # EBIT
        ebit = ebitda - da

        # Pre-Tax Income (now with actual interest!)
        pretax = ebit - interest_expense

        # Taxes (floored at 0)
        taxes = max(0, pretax * tax_rate)

        # Net Income
        net_income = pretax - taxes

        # FCF components
        capex = revenue * capex_pct
        nwc_change = revenue * nwc_pct

        # FCF for Debt Paydown
        fcf = net_income + da - capex - nwc_change

        # =====================================================================
        # DEBT SCHEDULE - Paydown (depends on FCF which is now calculated)
        # =====================================================================
        cash_available = fcf
        # Optional paydown = MIN(cash after mandatory, debt remaining after mandatory)
        # Floor at 0 to prevent negative paydown
        optional_paydown = max(0, min(
            cash_available - mandatory_amort,
            beginning_balance - mandatory_amort
        ))
        ending_balance = beginning_balance - mandatory_amort - optional_paydown

        # Handle edge case where ending balance goes slightly negative due to rounding
        ending_balance = max(0, ending_balance)

        # Leverage ratio
        leverage_ratio = ending_balance / ebitda if ebitda > 0 else 0

        # Store Operating Model projection
        projections.append({
            'year': year,
            'calendar_year': most_recent_fy + year,
            'revenue': revenue,
            'ebitda': ebitda,
            'da': da,
            'ebit': ebit,
            'interest': interest_expense,
            'pretax': pretax,
            'taxes': taxes,
            'net_income': net_income,
            'capex': capex,
            'nwc_change': nwc_change,
            'fcf': fcf,
        })

        # Store Debt Schedule
        debt_schedule.append({
            'year': year,
            'beginning_balance': beginning_balance,
            'interest_expense': interest_expense,
            'mandatory_amort': mandatory_amort,
            'cash_available': cash_available,
            'optional_paydown': optional_paydown,
            'ending_balance': ending_balance,
            'leverage_ratio': leverage_ratio,
        })

        # Prepare for next year
        prev_revenue = revenue
        beginning_balance = ending_balance

    return {
        'ticker': summary.get('ticker', 'Unknown'),
        'most_recent_fy': most_recent_fy,
        'entry_revenue': entry_revenue,
        'entry_ebitda': entry_ebitda,
        'new_debt_raised': new_debt_raised,
        'revenue_growth': revenue_growth,
        'growth_is_fallback': growth_is_fallback,
        'ebitda_margin': ebitda_margin,
        'da_pct': da_pct,
        'capex_pct': capex_pct,
        'tax_rate': tax_rate,
        'interest_rate': interest_rate,
        'mandatory_amort_pct': mandatory_amort_pct,
        'projections': projections,
        'debt_schedule': debt_schedule,
    }


def print_full_model(model: dict):
    """Print the full integrated model (Operating Model + Debt Schedule)."""
    if "error" in model:
        print(f"  ERROR: {model['error']}")
        return

    def fmt_currency(val):
        if abs(val) >= 1e9:
            return f"${val/1e9:,.2f}B"
        elif abs(val) >= 1e6:
            return f"${val/1e6:,.2f}M"
        else:
            return f"${val:,.0f}"

    ticker = model['ticker']
    year1 = model['projections'][0]
    final = model['projections'][-1]
    debt1 = model['debt_schedule'][0]
    debt_final = model['debt_schedule'][-1]

    print(f"\n  Full Integrated Model for {ticker}:")
    print(f"  Entry FY: {model['most_recent_fy']}")
    print(f"  New Debt Raised: {fmt_currency(model['new_debt_raised'])}")
    print(f"  Interest Rate: {model['interest_rate']*100:.1f}%")
    print(f"  Mandatory Amort: {model['mandatory_amort_pct']*100:.1f}%/year")
    print()

    # Operating Model Summary
    print(f"  --- OPERATING MODEL ---")
    print(f"  {'Metric':<20}  {'Year 1':>12}  {'Year {}'.format(len(model['projections'])):>12}")
    print("  " + "-" * 48)
    for label, key in [('Revenue', 'revenue'), ('EBITDA', 'ebitda'), ('Interest', 'interest'),
                        ('Net Income', 'net_income'), ('FCF', 'fcf')]:
        print(f"  {label:<20}  {fmt_currency(year1[key]):>12}  {fmt_currency(final[key]):>12}")

    # Debt Schedule Summary
    print()
    print(f"  --- DEBT SCHEDULE ---")
    print(f"  {'Metric':<20}  {'Year 1':>12}  {'Year {}'.format(len(model['debt_schedule'])):>12}")
    print("  " + "-" * 48)
    for label, key in [('Beginning Balance', 'beginning_balance'),
                        ('Interest Expense', 'interest_expense'),
                        ('Mandatory Amort', 'mandatory_amort'),
                        ('Optional Paydown', 'optional_paydown'),
                        ('Ending Balance', 'ending_balance')]:
        print(f"  {label:<20}  {fmt_currency(debt1[key]):>12}  {fmt_currency(debt_final[key]):>12}")

    print()
    print(f"  Leverage Ratio Year 1: {debt1['leverage_ratio']:.1f}x")
    print(f"  Leverage Ratio Year {len(model['debt_schedule'])}: {debt_final['leverage_ratio']:.1f}x")

    # Debt paydown check
    total_paydown = model['new_debt_raised'] - debt_final['ending_balance']
    pct_paid = total_paydown / model['new_debt_raised'] * 100 if model['new_debt_raised'] > 0 else 0
    print(f"  Total Debt Paydown: {fmt_currency(total_paydown)} ({pct_paid:.1f}% of original)")

    # Check for anomalies
    if debt_final['ending_balance'] < 0:
        print("  [WARNING] Ending balance went negative!")
    if debt_final['ending_balance'] > model['new_debt_raised']:
        print("  [WARNING] Ending balance exceeds original debt!")


# =============================================================================
# MOCK DATA FOR TESTING (when SEC credentials not available)
# =============================================================================

def get_mock_data(ticker: str) -> tuple:
    """Generate mock data for testing Excel generation without SEC API access."""
    mock_data = {
        "AAPL": {
            "summary": {
                "ticker": "AAPL",
                "company_name": "Apple Inc.",
                "cik": 320193,
                "sic_code": "3571",
                "sic_description": "Electronic Computers",
                "fiscal_year_end": "09-30",
                "current_price": 195.50,
                "shares_outstanding": 15500000000,
                "fiscal_years": {
                    2024: {
                        "revenue": 383285000000,
                        "operating_income": 114301000000,
                        "da": 11200000000,
                        "ebitda_calculated": 125501000000,
                        "total_debt": 111088000000,
                        "cash": 61555000000,
                        "capex": 9959000000,
                        "missing": []
                    },
                    2023: {
                        "revenue": 394328000000,
                        "operating_income": 114301000000,
                        "da": 11519000000,
                        "ebitda_calculated": 125820000000,
                        "total_debt": 109614000000,
                        "cash": 29965000000,
                        "capex": 10959000000,
                        "missing": []
                    },
                    2022: {
                        "revenue": 365817000000,
                        "operating_income": 119437000000,
                        "da": 11104000000,
                        "ebitda_calculated": 130541000000,
                        "total_debt": 111824000000,
                        "cash": 23646000000,
                        "capex": 10708000000,
                        "missing": []
                    },
                    2021: {
                        "revenue": 274515000000,
                        "operating_income": 108949000000,
                        "da": 11284000000,
                        "ebitda_calculated": 120233000000,
                        "total_debt": 119381000000,
                        "cash": 34940000000,
                        "capex": 11085000000,
                        "missing": []
                    },
                    2020: {
                        "revenue": 274515000000,
                        "operating_income": 66288000000,
                        "da": 11056000000,
                        "ebitda_calculated": 77344000000,
                        "total_debt": 112043000000,
                        "cash": 38016000000,
                        "capex": 7309000000,
                        "missing": []
                    }
                }
            },
            "validation": {
                "status": "pass",
                "missing_hard": [],
                "missing_soft": [],
                "disqualifying_reasons": [],
                "sector_excluded": False,
                "sector_excluded_reason": None,
                "defaults_applied": []
            }
        },
        "FIZZ": {
            "summary": {
                "ticker": "FIZZ",
                "company_name": "National Beverage Corp.",
                "cik": 879993,
                "sic_code": "2086",
                "sic_description": "Bottled & Canned Soft Drinks",
                "fiscal_year_end": "04-30",
                "current_price": 45.20,
                "shares_outstanding": 46200000,
                "fiscal_years": {
                    2024: {
                        "revenue": 1184800000,
                        "operating_income": 174200000,
                        "da": 32500000,
                        "ebitda_calculated": 206700000,
                        "total_debt": None,  # Will be defaulted
                        "cash": 145000000,
                        "capex": 28000000,
                        "missing": ["total_debt"]
                    },
                    2023: {
                        "revenue": 1171500000,
                        "operating_income": 201500000,
                        "da": 30200000,
                        "ebitda_calculated": 231700000,
                        "total_debt": None,
                        "cash": 167000000,
                        "capex": 24000000,
                        "missing": ["total_debt"]
                    },
                    2022: {
                        "revenue": 1148600000,
                        "operating_income": 183400000,
                        "da": 28900000,
                        "ebitda_calculated": 212300000,
                        "total_debt": None,
                        "cash": 189000000,
                        "capex": 31000000,
                        "missing": ["total_debt"]
                    },
                    2021: {
                        "revenue": 1075000000,
                        "operating_income": 205000000,
                        "da": 26800000,
                        "ebitda_calculated": 231800000,
                        "total_debt": None,
                        "cash": 203000000,
                        "capex": 21000000,
                        "missing": ["total_debt"]
                    },
                    2020: {
                        "revenue": 1014000000,
                        "operating_income": 181000000,
                        "da": 25100000,
                        "ebitda_calculated": 206100000,
                        "total_debt": None,
                        "cash": 196000000,
                        "capex": 19000000,
                        "missing": ["total_debt"]
                    }
                }
            },
            "validation": {
                "status": "degraded",
                "missing_hard": [],
                "missing_soft": ["Total debt data is missing for FY2024. This could indicate a genuinely debt-free company."],
                "disqualifying_reasons": [],
                "sector_excluded": False,
                "sector_excluded_reason": None,
                "defaults_applied": ["Total debt will default to $0 — verify this is accurate before proceeding."]
            }
        },
        "CCL": {
            "summary": {
                "ticker": "CCL",
                "company_name": "Carnival Corporation",
                "cik": 815097,
                "sic_code": "4481",
                "sic_description": "Deep Sea Foreign Transportation of Passengers",
                "fiscal_year_end": "11-30",
                "current_price": 18.75,
                "shares_outstanding": 1270000000,
                "fiscal_years": {
                    2024: {
                        "revenue": 24530000000,
                        "operating_income": 3265000000,
                        "da": 2800000000,
                        "ebitda_calculated": 6065000000,
                        "total_debt": 31500000000,  # Heavy debt
                        "cash": 1850000000,
                        "capex": 4200000000,
                        "missing": []
                    },
                    2023: {
                        "revenue": 21630000000,
                        "operating_income": 1958000000,
                        "da": 2600000000,
                        "ebitda_calculated": 4558000000,
                        "total_debt": 33200000000,
                        "cash": 2700000000,
                        "capex": 3500000000,
                        "missing": []
                    },
                    2022: {
                        "revenue": 12169000000,
                        "operating_income": -3865000000,
                        "da": 2500000000,
                        "ebitda_calculated": -1365000000,
                        "total_debt": 35400000000,
                        "cash": 7320000000,
                        "capex": 2100000000,
                        "missing": []
                    },
                    2021: {
                        "revenue": 1908000000,
                        "operating_income": -7810000000,
                        "da": 2380000000,
                        "ebitda_calculated": -5430000000,
                        "total_debt": 33100000000,
                        "cash": 9530000000,
                        "capex": 1400000000,
                        "missing": []
                    },
                    2020: {
                        "revenue": 5595000000,
                        "operating_income": -8878000000,
                        "da": 2340000000,
                        "ebitda_calculated": -6538000000,
                        "total_debt": 26650000000,
                        "cash": 9500000000,
                        "capex": 2600000000,
                        "missing": []
                    }
                }
            },
            "validation": {
                "status": "pass",
                "missing_hard": [],
                "missing_soft": [],
                "disqualifying_reasons": [],
                "sector_excluded": False,
                "sector_excluded_reason": None,
                "defaults_applied": []
            }
        }
    }

    if ticker in mock_data:
        return mock_data[ticker]["summary"], mock_data[ticker]["validation"]
    return None, None


# =============================================================================
# TEST HARNESS
# =============================================================================

if __name__ == "__main__":
    import sys
    sys.path.insert(0, str(Path(__file__).parent))

    # Check for --mock flag to use mock data
    use_mock = "--mock" in sys.argv

    # Allow command-line override for SEC email
    # Usage: python excel_generator.py [sec_email]
    # Or: python excel_generator.py --mock (use mock data)
    for arg in sys.argv[1:]:
        if arg != "--mock" and "@" in arg:
            os.environ["SEC_CONTACT_EMAIL"] = arg

    print("="*70)
    print("AIO LBO EXCEL GENERATOR TEST")
    print("="*70)

    # Test tickers
    test_tickers = ["AAPL", "FIZZ", "CCL"]

    print(f"\nTest tickers:")
    print("  AAPL - Large-cap, clean data")
    print("  FIZZ - Small-cap, debt likely defaulted to $0")
    print("  CCL  - Debt-heavy, should have real debt figures")

    if use_mock:
        print("\n[MOCK MODE] Using mock data for testing")
    else:
        # Import SEC modules only if not using mock
        try:
            from sec_edgar_test import fetch_ticker_data, get_sec_user_agent, get_twelve_data_api_key
            from validator import validate

            sec_user_agent = get_sec_user_agent()
            try:
                twelve_data_key = get_twelve_data_api_key()
            except ValueError:
                twelve_data_key = ""
                print("Note: No Twelve Data key - current price will be missing")
        except ValueError as e:
            print(f"\nWARNING: {e}")
            print("Falling back to mock data mode.\n")
            use_mock = True

    output_dir = Path(__file__).parent / "output"
    output_dir.mkdir(exist_ok=True)

    for ticker in test_tickers:
        print(f"\n{'='*70}")
        print(f"PROCESSING: {ticker}")
        print("="*70)

        if use_mock:
            # Use mock data
            print(f"\n[1/3] Loading mock data...")
            summary, validation = get_mock_data(ticker)
            if not summary:
                print(f"  ERROR: No mock data for {ticker}")
                continue
            print(f"  Loaded mock data for {summary['company_name']}")
        else:
            # Fetch real data
            print(f"\n[1/3] Fetching data from SEC EDGAR...")
            summary = fetch_ticker_data(
                ticker,
                sec_user_agent,
                twelve_data_key,
                verbose=False,
                skip_price=True
            )

            if not summary:
                print(f"  ERROR: Failed to fetch data for {ticker}")
                continue

            # Validate
            print(f"[2/3] Validating...")
            validation = validate(summary)

        print(f"  Status: {validation['status'].upper()}")
        if validation['defaults_applied']:
            print(f"  Defaults applied:")
            for d in validation['defaults_applied']:
                print(f"    - {d}")

        # Generate workbook
        print(f"[{'2' if use_mock else '3'}/3] Generating Excel workbook...")
        validated_data = {
            'summary': summary,
            'validation': validation
        }

        filepath = generate_workbook(validated_data, str(output_dir))

        # Verify formulas
        print(f"\n--- Verifying formulas ---")
        verification = verify_workbook_formulas(filepath)

        # Print key values (formula strings since we can't evaluate)
        print(f"\nKey formula strings from Sources & Uses:")
        for cell, formula in verification['sources_uses_formulas'].items():
            print(f"  {cell}: {formula}")

        # Print Operating Model formulas (Year 1 and Year 5 only) - focus on Interest Expense
        print(f"\nKey formula strings from Operating Model (Interest Expense row):")
        op_formulas = verification.get('operating_model_formulas', {})
        for cell, formula in sorted(op_formulas.items()):
            # Only show columns C (Year 1) and G (Year 5)
            col = cell[0]
            if col in ['C', 'G'] and 'Debt Schedule' in formula:
                print(f"  {cell}: {formula}")

        # Print Debt Schedule formulas (Year 1 and Year 5 only)
        print(f"\nKey formula strings from Debt Schedule (Year 1 & Year 5):")
        debt_formulas = verification.get('debt_schedule_formulas', {})
        for cell, formula in sorted(debt_formulas.items()):
            col = cell[0]
            if col in ['C', 'G']:
                print(f"  {cell}: {formula}")

        # Print Returns tab formulas
        print(f"\nKey formula strings from Returns tab:")
        returns_formulas = verification.get('returns_formulas', {})
        for cell, formula in sorted(returns_formulas.items()):
            print(f"  {cell}: {formula}")

        # Print named ranges being used
        print(f"\nNamed ranges used:")
        key_names = ['EntryRevenue', 'EntryEBITDA', 'EntryMultiple', 'LeverageMultiple',
                     'TransactionFeePct', 'TotalUses', 'NewDebtRaised',
                     'SponsorEquity', 'TotalSources', 'BalanceCheck', 'ExitEBITDA',
                     'RevenueGrowthRate', 'EBITDAMargin', 'DAPct', 'CapExPct', 'NWCPct', 'TaxRate',
                     'InterestRate', 'AmortizationPct', 'ExitDebtBalance']
        for name in key_names:
            if name in verification['named_ranges']:
                print(f"  {name}: {verification['named_ranges'][name]}")

        # Compute and print expected values
        print(f"\n--- Sources & Uses Computed Values ---")
        computed = evaluate_model_values(summary, validation)
        print_computed_values(computed)

        # Compute and print Full Integrated Model (Operating Model + Debt Schedule)
        print(f"\n--- Full Integrated Model (Operating Model + Debt Schedule) ---")
        full_model = evaluate_full_model(summary, validation, exit_year=5)
        print_full_model(full_model)

    print(f"\n{'='*70}")
    print("TEST COMPLETE")
    print("="*70)
    print(f"\nGenerated files are in: {output_dir}")
    print("\nNOTE: openpyxl writes formulas as text. To see computed values,")
    print("open the files in Excel or a compatible spreadsheet application.")
    print("The Balance Check cell should show $0 if formulas are correct.")
