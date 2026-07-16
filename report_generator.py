"""
AIO LBO Report Generator

Reads a generated .xlsx file, extracts computed values with explicit data provenance,
computes a deterministic feasibility score, and calls an LLM to produce a narrative report.
"""

import json
import os
import shutil
import subprocess
import sys
import tempfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Callable, Dict, List, Literal, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def sanitize_for_console(text: str) -> str:
    """
    Replace Unicode characters that can't be displayed on Windows console (cp1252).
    The LLM may use mathematical symbols like ≤ which fail on Windows.
    """
    replacements = {
        '\u2264': '<=',  # ≤
        '\u2265': '>=',  # ≥
        '\u2260': '!=',  # ≠
        '\u00d7': 'x',   # ×
        '\u2013': '-',   # en dash
        '\u2014': '--',  # em dash
        '\u2018': "'",   # left single quote
        '\u2019': "'",   # right single quote
        '\u201c': '"',   # left double quote
        '\u201d': '"',   # right double quote
        '\u2022': '*',   # bullet
        '\u2026': '...', # ellipsis
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text


# =============================================================================
# PART 1: LIBREOFFICE RECALCULATION
# =============================================================================

def find_libreoffice_executable() -> Optional[str]:
    """
    Find the LibreOffice executable on the system.

    Checks:
    1. 'soffice' on PATH
    2. 'libreoffice' on PATH
    3. Common Windows installation paths
    4. Common macOS paths

    Returns the executable path or None if not found.
    """
    # Check PATH first
    for cmd in ['soffice', 'libreoffice']:
        path = shutil.which(cmd)
        if path:
            return path

    # Common Windows installation paths
    windows_paths = [
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        r"C:\Program Files\LibreOffice 7\program\soffice.exe",
        r"C:\Program Files\LibreOffice 24\program\soffice.exe",
    ]

    # Common macOS paths
    mac_paths = [
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
    ]

    # Common Linux paths (beyond PATH)
    linux_paths = [
        "/usr/bin/soffice",
        "/usr/bin/libreoffice",
        "/usr/lib/libreoffice/program/soffice",
    ]

    all_paths = windows_paths + mac_paths + linux_paths

    for path in all_paths:
        if os.path.isfile(path):
            return path

    return None


def recalculate_workbook_excel_com(filepath: str) -> str:
    """
    Recalculate a workbook using Excel COM on Windows.

    This is a fallback when LibreOffice is not available.
    Only works on Windows with Excel installed.

    Returns path to recalculated temporary file.
    """
    if sys.platform != 'win32':
        raise RuntimeError("Excel COM automation only available on Windows")

    try:
        import win32com.client
    except ImportError:
        raise RuntimeError(
            "pywin32 not installed. Run: pip install pywin32\n"
            "Or install LibreOffice for cross-platform support."
        )

    # Create temp directory
    temp_dir = tempfile.mkdtemp(prefix="aiolbo_recalc_")
    input_name = Path(filepath).stem
    output_path = os.path.join(temp_dir, f"{input_name}.xlsx")

    excel = None
    wb = None

    try:
        # Start Excel
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        # Open workbook
        abs_path = os.path.abspath(filepath)
        wb = excel.Workbooks.Open(abs_path)

        # Force recalculation
        wb.Application.CalculateFull()

        # Save to temp location
        abs_output = os.path.abspath(output_path)
        wb.SaveAs(abs_output, FileFormat=51)  # 51 = xlsx

        return output_path

    except Exception as e:
        shutil.rmtree(temp_dir, ignore_errors=True)
        raise RuntimeError(f"Excel COM recalculation failed: {e}")

    finally:
        if wb:
            wb.Close(SaveChanges=False)
        if excel:
            excel.Quit()


def recalculate_workbook(filepath: str, allow_excel_fallback: bool = True) -> str:
    """
    Recalculate a workbook using LibreOffice headless mode.

    This is critical: openpyxl cannot evaluate Excel formulas, so reading
    the original file would give None for all formula cells. We must run
    the file through a real spreadsheet engine first.

    Args:
        filepath: Path to the .xlsx file to recalculate
        allow_excel_fallback: If True, try Excel COM on Windows when LibreOffice unavailable

    Returns:
        Path to the recalculated temporary file

    Raises:
        RuntimeError: If no spreadsheet engine is available or conversion fails
    """
    lo_path = find_libreoffice_executable()

    if not lo_path:
        # Try Excel COM fallback on Windows
        if allow_excel_fallback and sys.platform == 'win32':
            print("  LibreOffice not found, trying Excel COM...")
            return recalculate_workbook_excel_com(filepath)

        raise RuntimeError(
            "No spreadsheet engine found for recalculation.\n\n"
            "Options:\n"
            "1. Install LibreOffice and ensure 'soffice' is on your PATH\n"
            "   Download from: https://www.libreoffice.org/download/download/\n"
            "2. On Windows with Excel installed: pip install pywin32\n\n"
            "This is required because openpyxl cannot evaluate Excel formulas - we must "
            "run the workbook through a real spreadsheet engine to get computed values."
        )

    # Create a temp directory for the output
    temp_dir = tempfile.mkdtemp(prefix="aiolbo_recalc_")

    try:
        # Run LibreOffice headless conversion
        # --convert-to xlsx forces recalculation during the save
        result = subprocess.run(
            [
                lo_path,
                "--headless",
                "--convert-to", "xlsx",
                "--outdir", temp_dir,
                filepath
            ],
            capture_output=True,
            text=True,
            timeout=60  # 60 second timeout
        )

        if result.returncode != 0:
            raise RuntimeError(
                f"LibreOffice conversion failed (exit code {result.returncode}):\n"
                f"stdout: {result.stdout}\n"
                f"stderr: {result.stderr}"
            )

        # Find the output file
        input_name = Path(filepath).stem
        output_path = os.path.join(temp_dir, f"{input_name}.xlsx")

        if not os.path.exists(output_path):
            # LibreOffice might have changed the name slightly
            xlsx_files = list(Path(temp_dir).glob("*.xlsx"))
            if xlsx_files:
                output_path = str(xlsx_files[0])
            else:
                raise RuntimeError(
                    f"LibreOffice conversion produced no output file.\n"
                    f"temp_dir contents: {os.listdir(temp_dir)}\n"
                    f"stdout: {result.stdout}\n"
                    f"stderr: {result.stderr}"
                )

        return output_path

    except subprocess.TimeoutExpired:
        shutil.rmtree(temp_dir, ignore_errors=True)
        raise RuntimeError("LibreOffice conversion timed out after 60 seconds")
    except Exception as e:
        shutil.rmtree(temp_dir, ignore_errors=True)
        raise


# =============================================================================
# PART 1: DATA EXTRACTION WITH PROVENANCE TAGGING
# =============================================================================

SourceType = Literal["fetched", "defaulted", "substituted", "user_assumption", "calculated"]


@dataclass
class TaggedValue:
    """A value with explicit provenance tracking."""
    value: Any
    source: SourceType
    note: Optional[str] = None

    def to_dict(self) -> dict:
        result = {"value": self.value, "source": self.source}
        if self.note:
            result["note"] = self.note
        return result


@dataclass
class ExtractedData:
    """All extracted data from a workbook with provenance tags."""

    # Company info
    company_name: TaggedValue = None
    ticker: TaggedValue = None
    sector: TaggedValue = None

    # Entry deal structure (from Sources & Uses)
    purchase_ev: TaggedValue = None
    entry_multiple: TaggedValue = None
    total_uses: TaggedValue = None
    new_debt_raised: TaggedValue = None
    sponsor_equity: TaggedValue = None

    # Operating trajectory (from Operating Model)
    revenue_year1: TaggedValue = None
    revenue_exit: TaggedValue = None
    ebitda_year1: TaggedValue = None
    ebitda_exit: TaggedValue = None

    # Full projection data for scoring (lists of values by year)
    fcf_by_year: List[float] = field(default_factory=list)
    interest_by_year: List[float] = field(default_factory=list)
    mandatory_amort_by_year: List[float] = field(default_factory=list)

    # Debt trajectory (from Debt Schedule)
    debt_beginning_year1: TaggedValue = None
    debt_ending_year1: TaggedValue = None
    debt_ending_exit: TaggedValue = None
    leverage_ratio_year1: TaggedValue = None
    leverage_ratio_exit: TaggedValue = None

    # Exit (from Returns)
    exit_ebitda: TaggedValue = None
    exit_ev: TaggedValue = None
    exit_debt: TaggedValue = None
    exit_equity: TaggedValue = None

    # Headline returns
    irr: TaggedValue = None
    moic: TaggedValue = None

    # Sensitivity range
    moic_min: TaggedValue = None
    moic_max: TaggedValue = None
    moic_base: TaggedValue = None
    irr_min: TaggedValue = None
    irr_max: TaggedValue = None
    irr_base: TaggedValue = None

    # The 14 Assumptions inputs
    revenue_growth_rate: TaggedValue = None
    ebitda_margin: TaggedValue = None
    offer_premium: TaggedValue = None
    leverage_multiple: TaggedValue = None
    interest_rate: TaggedValue = None
    tax_rate: TaggedValue = None
    capex_pct: TaggedValue = None
    da_pct: TaggedValue = None
    nwc_pct: TaggedValue = None
    exit_year: TaggedValue = None
    exit_multiple: TaggedValue = None
    transaction_fee_pct: TaggedValue = None
    amortization_pct: TaggedValue = None

    # Historical data from Assumptions
    entry_revenue: TaggedValue = None
    entry_ebitda: TaggedValue = None
    total_debt: TaggedValue = None
    cash: TaggedValue = None
    current_price: TaggedValue = None
    shares_outstanding: TaggedValue = None

    def to_dict(self) -> dict:
        """Convert to a dictionary for JSON serialization."""
        result = {}
        for key, value in self.__dict__.items():
            if isinstance(value, TaggedValue):
                result[key] = value.to_dict()
            elif isinstance(value, list):
                result[key] = value
            elif value is not None:
                result[key] = value
        return result

    def get_defaulted_fields(self) -> List[Tuple[str, str]]:
        """Return list of (field_name, note) for all defaulted/substituted fields."""
        results = []
        for key, value in self.__dict__.items():
            if isinstance(value, TaggedValue):
                if value.source in ("defaulted", "substituted") and value.note:
                    results.append((key, value.note))
        return results


def get_cell_provenance(ws, row: int, col: int, assumptions_ws) -> Tuple[SourceType, Optional[str]]:
    """
    Determine the provenance of a cell based on its styling and position.

    Uses the color convention:
    - Blue font = hardcoded input (fetched or user_assumption)
    - Black font = formula (calculated)
    - Orange fill = substituted (fallback calculation)
    - Yellow fill = defaulted (soft requirement missing)

    Returns (source_type, note_if_any)
    """
    cell = ws.cell(row=row, column=col)

    # Check fill color first (overrides font color)
    fill = cell.fill
    if fill and fill.start_color and fill.start_color.rgb:
        rgb = str(fill.start_color.rgb)
        # Yellow = defaulted (FFFF99)
        if rgb in ('FFFF99', '00FFFF99'):
            comment = cell.comment.text if cell.comment else None
            return ("defaulted", comment)
        # Orange = substituted/implausible (FFB366)
        if rgb in ('FFB366', '00FFB366'):
            comment = cell.comment.text if cell.comment else None
            return ("substituted", comment)

    # Check font color
    font = cell.font
    if font and font.color and font.color.rgb:
        rgb = str(font.color.rgb)
        # Blue font = hardcoded input
        if rgb in ('0000CC', '000000CC'):
            # It's an input - but is it fetched data or user assumption?
            # User assumptions are the 14 deal inputs (rows 28-41 in Assumptions)
            # Fetched data is company data (earlier rows)
            return ("user_assumption", None)
        # Black font = formula
        if rgb in ('000000', '00000000'):
            return ("calculated", None)

    # Default to calculated for formula cells
    if cell.value is not None and isinstance(cell.value, str) and str(cell.value).startswith('='):
        return ("calculated", None)

    return ("fetched", None)


def extract_data_from_workbook(filepath: str) -> ExtractedData:
    """
    Extract all relevant data from a recalculated workbook with provenance tagging.

    Args:
        filepath: Path to the recalculated .xlsx file

    Returns:
        ExtractedData object with all values tagged
    """
    # Load in data_only mode to get computed values
    wb = load_workbook(filepath, data_only=True)

    # Also load without data_only to check formulas/styling
    wb_formulas = load_workbook(filepath, data_only=False)

    data = ExtractedData()

    # ==========================================================================
    # ASSUMPTIONS TAB
    # ==========================================================================
    if 'Assumptions' in wb.sheetnames:
        ws = wb['Assumptions']
        ws_fmt = wb_formulas['Assumptions']

        # Company info (rows 3-5 area, column C)
        # Row 3: Company Name
        data.company_name = TaggedValue(
            value=ws.cell(row=3, column=3).value,
            source="fetched"
        )

        # Row 4: Ticker
        data.ticker = TaggedValue(
            value=ws.cell(row=4, column=3).value,
            source="fetched"
        )

        # Row 6: SIC Code / Sector (row 5 is CIK, not sector)
        # Format in Excel: "3571 - Electronic Computers"
        data.sector = TaggedValue(
            value=ws.cell(row=6, column=3).value,
            source="fetched"
        )

        # Historical financials (around rows 15-22)
        # Row 15: Revenue (most recent FY in column C)
        data.entry_revenue = TaggedValue(
            value=ws.cell(row=15, column=3).value,
            source="fetched"
        )

        # Row 18: EBITDA
        data.entry_ebitda = TaggedValue(
            value=ws.cell(row=18, column=3).value,
            source="fetched"
        )

        # Row 21: Total Debt
        debt_cell = ws.cell(row=21, column=3)
        debt_fmt_cell = ws_fmt.cell(row=21, column=3)
        debt_source, debt_note = "fetched", None

        # Check for yellow/defaulted fill (handle both with and without alpha prefix)
        if debt_fmt_cell.fill and debt_fmt_cell.fill.start_color:
            rgb = str(debt_fmt_cell.fill.start_color.rgb)
            # Yellow fill = FFFF99 (may have 00 alpha prefix)
            if 'FFFF99' in rgb.upper():
                debt_source = "defaulted"
                debt_note = debt_fmt_cell.comment.text if debt_fmt_cell.comment else "Total debt not reported, defaulted to $0"

        data.total_debt = TaggedValue(
            value=debt_cell.value or 0,
            source=debt_source,
            note=debt_note
        )

        # Row 22: Cash
        data.cash = TaggedValue(
            value=ws.cell(row=22, column=3).value,
            source="fetched"
        )

        # Row 24: Current Price
        data.current_price = TaggedValue(
            value=ws.cell(row=24, column=3).value,
            source="fetched"
        )

        # Row 25: Shares Outstanding
        data.shares_outstanding = TaggedValue(
            value=ws.cell(row=25, column=3).value,
            source="fetched"
        )

        # Deal assumptions (rows 28-41)
        # Row 28: Revenue Growth Rate
        growth_cell = ws.cell(row=28, column=3)
        growth_fmt_cell = ws_fmt.cell(row=28, column=3)
        growth_source, growth_note = "user_assumption", None

        # Check for orange fill (substituted - fallback growth rate)
        if growth_fmt_cell.fill and growth_fmt_cell.fill.start_color:
            rgb = str(growth_fmt_cell.fill.start_color.rgb)
            # Orange fill = FFB366 (may have 00 alpha prefix)
            if 'FFB366' in rgb.upper():
                growth_source = "substituted"
                growth_note = growth_fmt_cell.comment.text if growth_fmt_cell.comment else "Growth rate fell outside sanity band, using 3% fallback"

        data.revenue_growth_rate = TaggedValue(
            value=growth_cell.value,
            source=growth_source,
            note=growth_note
        )

        # Row 29: EBITDA Margin
        data.ebitda_margin = TaggedValue(
            value=ws.cell(row=29, column=3).value,
            source="user_assumption"
        )

        # Row 30: Entry Multiple
        data.entry_multiple = TaggedValue(
            value=ws.cell(row=30, column=3).value,
            source="user_assumption"
        )

        # Row 31: Offer Premium
        data.offer_premium = TaggedValue(
            value=ws.cell(row=31, column=3).value,
            source="user_assumption"
        )

        # Row 32: Leverage Multiple
        data.leverage_multiple = TaggedValue(
            value=ws.cell(row=32, column=3).value,
            source="user_assumption"
        )

        # Row 33: Interest Rate
        data.interest_rate = TaggedValue(
            value=ws.cell(row=33, column=3).value,
            source="user_assumption"
        )

        # Row 34: Tax Rate
        data.tax_rate = TaggedValue(
            value=ws.cell(row=34, column=3).value,
            source="user_assumption"
        )

        # Row 35: CapEx %
        data.capex_pct = TaggedValue(
            value=ws.cell(row=35, column=3).value,
            source="user_assumption"
        )

        # Row 36: D&A %
        data.da_pct = TaggedValue(
            value=ws.cell(row=36, column=3).value,
            source="user_assumption"
        )

        # Row 37: NWC %
        data.nwc_pct = TaggedValue(
            value=ws.cell(row=37, column=3).value,
            source="user_assumption"
        )

        # Row 38: Exit Year
        exit_year_val = ws.cell(row=38, column=3).value
        data.exit_year = TaggedValue(
            value=int(exit_year_val) if exit_year_val else 5,
            source="user_assumption"
        )

        # Row 39: Exit Multiple
        data.exit_multiple = TaggedValue(
            value=ws.cell(row=39, column=3).value,
            source="user_assumption"
        )

        # Row 40: Transaction Fee %
        data.transaction_fee_pct = TaggedValue(
            value=ws.cell(row=40, column=3).value,
            source="user_assumption"
        )

        # Row 41: Amortization %
        data.amortization_pct = TaggedValue(
            value=ws.cell(row=41, column=3).value,
            source="user_assumption"
        )

    # ==========================================================================
    # SOURCES & USES TAB
    # ==========================================================================
    if 'Sources & Uses' in wb.sheetnames:
        ws = wb['Sources & Uses']

        # Row 3: Purchase EV
        data.purchase_ev = TaggedValue(
            value=ws.cell(row=3, column=3).value,
            source="calculated"
        )

        # Row 6: Total Uses
        data.total_uses = TaggedValue(
            value=ws.cell(row=6, column=3).value,
            source="calculated"
        )

        # Row 11: New Debt Raised
        data.new_debt_raised = TaggedValue(
            value=ws.cell(row=11, column=3).value,
            source="calculated"
        )

        # Row 12: Sponsor Equity
        data.sponsor_equity = TaggedValue(
            value=ws.cell(row=12, column=3).value,
            source="calculated"
        )

    # ==========================================================================
    # OPERATING MODEL TAB
    # ==========================================================================
    exit_year = data.exit_year.value if data.exit_year else 5

    if 'Operating Model' in wb.sheetnames:
        ws = wb['Operating Model']

        # Operating Model structure:
        # Row 5: "Income Statement" section header
        # Row 6: Revenue
        # Row 7: EBITDA
        # Row 20: Free Cash Flow for Debt Paydown
        year1_col = 3  # Column C = Year 1
        exit_col = 2 + exit_year  # Column for exit year (e.g., G for Year 5)

        # Row 6: Revenue
        data.revenue_year1 = TaggedValue(
            value=ws.cell(row=6, column=year1_col).value,
            source="calculated"
        )
        data.revenue_exit = TaggedValue(
            value=ws.cell(row=6, column=exit_col).value,
            source="calculated"
        )

        # Row 7: EBITDA
        data.ebitda_year1 = TaggedValue(
            value=ws.cell(row=7, column=year1_col).value,
            source="calculated"
        )
        data.ebitda_exit = TaggedValue(
            value=ws.cell(row=7, column=exit_col).value,
            source="calculated"
        )

        # Extract FCF for all years (Row 20) for scoring
        for year in range(1, exit_year + 1):
            col = 2 + year
            fcf = ws.cell(row=20, column=col).value
            data.fcf_by_year.append(fcf if fcf is not None else 0)

    # ==========================================================================
    # DEBT SCHEDULE TAB
    # ==========================================================================
    if 'Debt Schedule' in wb.sheetnames:
        ws = wb['Debt Schedule']

        year1_col = 3  # Column C
        exit_col = 2 + exit_year

        # Row 6: Beginning Balance
        data.debt_beginning_year1 = TaggedValue(
            value=ws.cell(row=6, column=year1_col).value,
            source="calculated"
        )

        # Row 7: Interest Expense - extract for all years for scoring
        for year in range(1, exit_year + 1):
            col = 2 + year
            interest = ws.cell(row=7, column=col).value
            data.interest_by_year.append(interest if interest is not None else 0)

        # Row 8: Mandatory Amortization - extract for all years for scoring
        for year in range(1, exit_year + 1):
            col = 2 + year
            amort = ws.cell(row=8, column=col).value
            data.mandatory_amort_by_year.append(amort if amort is not None else 0)

        # Row 11: Ending Balance
        data.debt_ending_year1 = TaggedValue(
            value=ws.cell(row=11, column=year1_col).value,
            source="calculated"
        )
        data.debt_ending_exit = TaggedValue(
            value=ws.cell(row=11, column=exit_col).value,
            source="calculated"
        )

        # Row 14: Leverage Ratio
        data.leverage_ratio_year1 = TaggedValue(
            value=ws.cell(row=14, column=year1_col).value,
            source="calculated"
        )
        data.leverage_ratio_exit = TaggedValue(
            value=ws.cell(row=14, column=exit_col).value,
            source="calculated"
        )

    # ==========================================================================
    # RETURNS TAB
    # ==========================================================================
    if 'Returns' in wb.sheetnames:
        ws = wb['Returns']

        # Row 7: Exit EBITDA
        data.exit_ebitda = TaggedValue(
            value=ws.cell(row=7, column=3).value,
            source="calculated"
        )

        # Row 9: Exit EV
        data.exit_ev = TaggedValue(
            value=ws.cell(row=9, column=3).value,
            source="calculated"
        )

        # Row 10: Exit Debt
        data.exit_debt = TaggedValue(
            value=ws.cell(row=10, column=3).value,
            source="calculated"
        )

        # Row 11: Exit Equity
        data.exit_equity = TaggedValue(
            value=ws.cell(row=11, column=3).value,
            source="calculated"
        )

        # Row 19: IRR (XIRR result)
        data.irr = TaggedValue(
            value=ws.cell(row=19, column=3).value,
            source="calculated"
        )

        # Row 22: MOIC
        data.moic = TaggedValue(
            value=ws.cell(row=22, column=3).value,
            source="calculated"
        )

    # ==========================================================================
    # SENSITIVITY TAB
    # ==========================================================================
    if 'Sensitivity' in wb.sheetnames:
        ws = wb['Sensitivity']

        # The grid is 5x5, starting at row 6 (headers) and row 7-11 (data)
        # MOIC grid: columns C-G (3-7), rows 6-10
        # IRR grid: columns J-N (10-14), rows 6-10

        # Extract all MOIC values from the grid
        moic_values = []
        for row in range(6, 11):  # Rows 6-10 (data rows after header)
            for col in range(3, 8):  # Columns C-G
                val = ws.cell(row=row, column=col).value
                if val is not None and isinstance(val, (int, float)):
                    moic_values.append(val)

        # Extract all IRR values from the grid
        irr_values = []
        for row in range(6, 11):
            for col in range(10, 15):  # Columns J-N
                val = ws.cell(row=row, column=col).value
                if val is not None and isinstance(val, (int, float)):
                    irr_values.append(val)

        # Base case is center cell (row 8, col E for MOIC, col L for IRR)
        moic_base_val = ws.cell(row=8, column=5).value  # E8
        irr_base_val = ws.cell(row=8, column=12).value  # L8

        if moic_values:
            data.moic_min = TaggedValue(value=min(moic_values), source="calculated")
            data.moic_max = TaggedValue(value=max(moic_values), source="calculated")
        data.moic_base = TaggedValue(value=moic_base_val, source="calculated")

        if irr_values:
            data.irr_min = TaggedValue(value=min(irr_values), source="calculated")
            data.irr_max = TaggedValue(value=max(irr_values), source="calculated")
        data.irr_base = TaggedValue(value=irr_base_val, source="calculated")

    wb.close()
    wb_formulas.close()

    return data


# =============================================================================
# PART 2: DETERMINISTIC FEASIBILITY SCORING
# =============================================================================

@dataclass
class ScoreBreakdown:
    """Breakdown of the feasibility score components."""
    irr_score: float  # 0-30 points
    moic_score: float  # 0-20 points
    debt_service_score: float  # 0-25 points
    leverage_reduction_score: float  # 0-15 points
    data_quality_score: float  # 0-10 points
    total_score: float  # 0-100 points

    # Additional context for the LLM
    irr_value: float
    moic_value: float
    min_coverage_ratio: float
    min_coverage_year: int
    leverage_reduction_pct: float
    defaulted_field_count: int

    def to_dict(self) -> dict:
        return {
            "components": {
                "irr": {"score": self.irr_score, "max": 30, "value": self.irr_value},
                "moic": {"score": self.moic_score, "max": 20, "value": self.moic_value},
                "debt_service": {
                    "score": self.debt_service_score,
                    "max": 25,
                    "min_coverage_ratio": self.min_coverage_ratio,
                    "worst_year": self.min_coverage_year
                },
                "leverage_reduction": {
                    "score": self.leverage_reduction_score,
                    "max": 15,
                    "reduction_pct": self.leverage_reduction_pct
                },
                "data_quality": {
                    "score": self.data_quality_score,
                    "max": 10,
                    "defaulted_fields": self.defaulted_field_count
                }
            },
            "total": self.total_score
        }


def compute_feasibility_score(data: ExtractedData) -> ScoreBreakdown:
    """
    Compute the deterministic feasibility score (0-100) from extracted data.

    Components:
    1. IRR vs hurdle rate (30 points max)
    2. MOIC (20 points max)
    3. Debt service coverage (25 points max)
    4. Leverage reduction (15 points max)
    5. Data quality (10 points max)
    """

    # ==========================================================================
    # 1. IRR Score (30 points max)
    # ==========================================================================
    irr_value = data.irr.value if data.irr and data.irr.value else 0

    if irr_value <= 0:
        irr_score = 0
    elif irr_value < 0.10:  # 0% to 10%
        # Scale from 0 to 5 points
        irr_score = (irr_value / 0.10) * 5
    elif irr_value < 0.25:  # 10% to 25%
        # Scale from 5 to 30 points (linear from 10% to 25%)
        irr_score = 5 + ((irr_value - 0.10) / 0.15) * 25
    else:  # 25%+
        irr_score = 30

    # ==========================================================================
    # 2. MOIC Score (20 points max)
    # ==========================================================================
    moic_value = data.moic.value if data.moic and data.moic.value else 0

    if moic_value < 1.0:
        moic_score = 0
    elif moic_value < 3.0:
        # Scale linearly from 0 at 1.0x to 20 at 3.0x
        moic_score = ((moic_value - 1.0) / 2.0) * 20
    else:
        moic_score = 20

    # ==========================================================================
    # 3. Debt Service Coverage Score (25 points max)
    # ==========================================================================
    # Coverage = FCF / Mandatory Amortization ONLY
    #
    # NOTE: FCF for Debt Paydown (from Operating Model) already has Interest
    # Expense subtracted out - it flows through Net Income which is post-interest.
    # So we only divide by Mandatory Amortization (the required principal payment)
    # to test whether operating cash flow covers the REQUIRED debt service.
    # Including interest in the denominator would double-count it.
    coverage_ratios = []

    for i in range(len(data.fcf_by_year)):
        fcf = data.fcf_by_year[i] if i < len(data.fcf_by_year) else 0
        amort = data.mandatory_amort_by_year[i] if i < len(data.mandatory_amort_by_year) else 0

        # Only divide by mandatory amortization (interest already in FCF numerator)
        if amort > 0:
            coverage = fcf / amort
        else:
            coverage = float('inf')  # No mandatory amortization = infinite coverage

        coverage_ratios.append((i + 1, coverage))  # (year, coverage)

    if coverage_ratios:
        min_coverage_year, min_coverage = min(coverage_ratios, key=lambda x: x[1])
    else:
        min_coverage_year, min_coverage = 1, 0

    if min_coverage < 1.0:
        debt_service_score = 0
    elif min_coverage < 2.5:
        # Scale linearly from 0 at 1.0x to 25 at 2.5x
        debt_service_score = ((min_coverage - 1.0) / 1.5) * 25
    else:
        debt_service_score = 25

    # ==========================================================================
    # 4. Leverage Reduction Score (15 points max)
    # ==========================================================================
    entry_leverage = data.leverage_ratio_year1.value if data.leverage_ratio_year1 and data.leverage_ratio_year1.value else 0
    exit_leverage = data.leverage_ratio_exit.value if data.leverage_ratio_exit and data.leverage_ratio_exit.value else 0

    if entry_leverage > 0:
        leverage_reduction_pct = (entry_leverage - exit_leverage) / entry_leverage
    else:
        leverage_reduction_pct = 0

    if leverage_reduction_pct <= 0:
        leverage_reduction_score = 0
    elif leverage_reduction_pct < 0.50:
        # Scale linearly from 0 at 0% to 15 at 50%
        leverage_reduction_score = (leverage_reduction_pct / 0.50) * 15
    else:
        leverage_reduction_score = 15

    # ==========================================================================
    # 5. Data Quality Score (10 points max)
    # ==========================================================================
    defaulted_fields = data.get_defaulted_fields()
    defaulted_count = len(defaulted_fields)

    data_quality_score = max(0, 10 - (defaulted_count * 2))

    # ==========================================================================
    # Total Score
    # ==========================================================================
    total_score = irr_score + moic_score + debt_service_score + leverage_reduction_score + data_quality_score

    return ScoreBreakdown(
        irr_score=round(irr_score, 1),
        moic_score=round(moic_score, 1),
        debt_service_score=round(debt_service_score, 1),
        leverage_reduction_score=round(leverage_reduction_score, 1),
        data_quality_score=round(data_quality_score, 1),
        total_score=round(total_score, 1),
        irr_value=irr_value,
        moic_value=moic_value,
        min_coverage_ratio=min_coverage if min_coverage != float('inf') else 999,
        min_coverage_year=min_coverage_year,
        leverage_reduction_pct=leverage_reduction_pct,
        defaulted_field_count=defaulted_count
    )


# =============================================================================
# PART 3: LLM PROVIDER ABSTRACTION
# =============================================================================

class LLMProviderError(Exception):
    """Error from an LLM provider that does NOT leak the API key."""
    pass


def _call_anthropic(prompt: str, api_key: str) -> str:
    """
    Call Anthropic Claude API.

    Returns the response text.
    Raises LLMProviderError on failure (without leaking the key).
    """
    try:
        import anthropic
    except ImportError:
        raise LLMProviderError(
            "anthropic package not installed. Run: pip install anthropic"
        )

    try:
        client = anthropic.Anthropic(api_key=api_key)

        response = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=2000,
            messages=[
                {"role": "user", "content": prompt}
            ]
        )

        return response.content[0].text

    except anthropic.AuthenticationError:
        raise LLMProviderError("Anthropic API authentication failed. Check your API key.")
    except anthropic.RateLimitError:
        raise LLMProviderError("Anthropic API rate limit exceeded. Try again later.")
    except anthropic.APIStatusError as e:
        raise LLMProviderError(f"Anthropic API error: {e.message}")
    except Exception as e:
        # Generic catch - don't include the full exception which might contain the key
        raise LLMProviderError(f"Anthropic API call failed: {type(e).__name__}")


def _call_openai(prompt: str, api_key: str) -> str:
    """
    Call OpenAI GPT API.

    Returns the response text.
    Raises LLMProviderError on failure (without leaking the key).
    """
    try:
        import openai
    except ImportError:
        raise LLMProviderError(
            "openai package not installed. Run: pip install openai"
        )

    try:
        client = openai.OpenAI(api_key=api_key)

        response = client.chat.completions.create(
            model="gpt-4o",
            max_tokens=2000,
            messages=[
                {"role": "user", "content": prompt}
            ]
        )

        return response.choices[0].message.content

    except openai.AuthenticationError:
        raise LLMProviderError("OpenAI API authentication failed. Check your API key.")
    except openai.RateLimitError:
        raise LLMProviderError("OpenAI API rate limit exceeded. Try again later.")
    except openai.APIStatusError as e:
        raise LLMProviderError(f"OpenAI API error: {e.message}")
    except Exception as e:
        # Generic catch - don't include the full exception which might contain the key
        raise LLMProviderError(f"OpenAI API call failed: {type(e).__name__}")


def _call_gemini(prompt: str, api_key: str) -> str:
    """
    Call Google Gemini API using the new google-genai package.

    Returns the response text.
    Raises LLMProviderError on failure (without leaking the key).
    """
    try:
        from google import genai
        from google.genai import types
    except ImportError:
        raise LLMProviderError(
            "google-genai package not installed. Run: pip install google-genai"
        )

    try:
        client = genai.Client(api_key=api_key)

        response = client.models.generate_content(
            model="gemini-3.5-flash",
            contents=prompt,
            config=types.GenerateContentConfig(
                max_output_tokens=4000,
            )
        )

        return response.text

    except Exception as e:
        error_type = type(e).__name__
        error_msg = str(e).lower()

        # Check for common error patterns without leaking the key
        if 'api_key' in error_msg or 'authentication' in error_msg or 'invalid' in error_msg:
            raise LLMProviderError("Gemini API authentication failed. Check your API key.")
        if 'quota' in error_msg or 'rate' in error_msg or 'limit' in error_msg:
            raise LLMProviderError("Gemini API rate limit or quota exceeded. Try again later.")
        if 'blocked' in error_msg or 'safety' in error_msg:
            raise LLMProviderError("Gemini API blocked the request due to safety filters.")
        if '404' in error_msg or 'not found' in error_msg:
            raise LLMProviderError("Gemini model not found. The model name may have changed.")

        # Generic error - don't leak details that might contain the key
        raise LLMProviderError(f"Gemini API call failed: {error_type}")


# Provider registry - easy to add more providers
LLM_PROVIDERS: Dict[str, Callable[[str, str], str]] = {
    "anthropic": _call_anthropic,
    "openai": _call_openai,
    "gemini": _call_gemini,
}


def generate_narrative(
    structured_data: ExtractedData,
    score_breakdown: ScoreBreakdown,
    provider: str,
    api_key: str,
    prompt_override: str = None
) -> str:
    """
    Generate a narrative report from extracted data and score.

    Args:
        structured_data: Extracted data with provenance tags
        score_breakdown: Computed feasibility score breakdown
        provider: LLM provider name ("anthropic", "openai", or "gemini")
        api_key: API key for the provider (NOT stored, logged, or written to disk)
        prompt_override: Optional custom prompt (used by comparison tool)

    Returns:
        Generated narrative report text

    Raises:
        LLMProviderError: On API failure (without leaking the key)
        ValueError: If provider is not supported
    """
    if provider not in LLM_PROVIDERS:
        raise ValueError(f"Unsupported LLM provider: {provider}. Supported: {list(LLM_PROVIDERS.keys())}")

    # Use custom prompt if provided, otherwise build the standard report prompt
    if prompt_override:
        prompt = prompt_override
    else:
        prompt = build_report_prompt(structured_data, score_breakdown)

    # Call the provider
    provider_fn = LLM_PROVIDERS[provider]
    return provider_fn(prompt, api_key)


# =============================================================================
# PART 4: THE PROMPT
# =============================================================================

def format_currency(value: float) -> str:
    """Format a number as currency (billions/millions)."""
    if value is None:
        return "N/A"
    if abs(value) >= 1e9:
        return f"${value/1e9:,.2f}B"
    elif abs(value) >= 1e6:
        return f"${value/1e6:,.2f}M"
    else:
        return f"${value:,.0f}"


def format_pct(value: float) -> str:
    """Format a number as percentage."""
    if value is None:
        return "N/A"
    return f"{value*100:.1f}%"


def format_multiple(value: float) -> str:
    """Format a number as a multiple."""
    if value is None:
        return "N/A"
    return f"{value:.2f}x"


def build_report_prompt(data: ExtractedData, score: ScoreBreakdown) -> str:
    """
    Build the prompt for the LLM narrative generation.
    """

    # Format the data for the prompt
    defaulted_fields = data.get_defaulted_fields()

    prompt = f"""You are writing a concise LBO analysis report. Use ONLY the data provided below. Do not invent, estimate, or speculate about anything not explicitly stated.

=== COMPANY ===
Company: {data.company_name.value if data.company_name else 'Unknown'}
Ticker: {data.ticker.value if data.ticker else 'Unknown'}
Sector: {data.sector.value if data.sector else 'Unknown'}

=== ENTRY DEAL STRUCTURE ===
Purchase Enterprise Value: {format_currency(data.purchase_ev.value if data.purchase_ev else 0)}
Entry EV/EBITDA Multiple: {format_multiple(data.entry_multiple.value if data.entry_multiple else 0)}
Total Uses (Purchase + Fees): {format_currency(data.total_uses.value if data.total_uses else 0)}
New Debt Raised: {format_currency(data.new_debt_raised.value if data.new_debt_raised else 0)}
Sponsor Equity Required: {format_currency(data.sponsor_equity.value if data.sponsor_equity else 0)}
Leverage Multiple: {format_multiple(data.leverage_multiple.value if data.leverage_multiple else 0)}

=== OPERATING TRAJECTORY ===
Exit Year: {data.exit_year.value if data.exit_year else 5}
Revenue Year 1: {format_currency(data.revenue_year1.value if data.revenue_year1 else 0)}
Revenue Exit Year: {format_currency(data.revenue_exit.value if data.revenue_exit else 0)}
EBITDA Year 1: {format_currency(data.ebitda_year1.value if data.ebitda_year1 else 0)}
EBITDA Exit Year: {format_currency(data.ebitda_exit.value if data.ebitda_exit else 0)}
Revenue Growth Rate: {format_pct(data.revenue_growth_rate.value if data.revenue_growth_rate else 0)}
EBITDA Margin: {format_pct(data.ebitda_margin.value if data.ebitda_margin else 0)}

=== DEBT TRAJECTORY ===
Beginning Debt (Year 1): {format_currency(data.debt_beginning_year1.value if data.debt_beginning_year1 else 0)}
Ending Debt (Year 1): {format_currency(data.debt_ending_year1.value if data.debt_ending_year1 else 0)}
Ending Debt (Exit Year): {format_currency(data.debt_ending_exit.value if data.debt_ending_exit else 0)}
Leverage Ratio Year 1: {format_multiple(data.leverage_ratio_year1.value if data.leverage_ratio_year1 else 0)}
Leverage Ratio Exit Year: {format_multiple(data.leverage_ratio_exit.value if data.leverage_ratio_exit else 0)}
Interest Rate: {format_pct(data.interest_rate.value if data.interest_rate else 0)}

=== EXIT SCENARIO ===
Exit EBITDA: {format_currency(data.exit_ebitda.value if data.exit_ebitda else 0)}
Exit EV/EBITDA Multiple: {format_multiple(data.exit_multiple.value if data.exit_multiple else 0)}
Exit Enterprise Value: {format_currency(data.exit_ev.value if data.exit_ev else 0)}
Exit Debt Balance: {format_currency(data.exit_debt.value if data.exit_debt else 0)}
Exit Equity Value: {format_currency(data.exit_equity.value if data.exit_equity else 0)}

=== HEADLINE RETURNS ===
IRR: {format_pct(data.irr.value if data.irr else 0)}
MOIC: {format_multiple(data.moic.value if data.moic else 0)}

=== SENSITIVITY RANGE (Entry Multiple x Exit Multiple) ===
MOIC Range: {format_multiple(data.moic_min.value if data.moic_min else 0)} to {format_multiple(data.moic_max.value if data.moic_max else 0)} (base case: {format_multiple(data.moic_base.value if data.moic_base else 0)})
IRR Range: {format_pct(data.irr_min.value if data.irr_min else 0)} to {format_pct(data.irr_max.value if data.irr_max else 0)} (base case: {format_pct(data.irr_base.value if data.irr_base else 0)})

=== FEASIBILITY SCORE: {score.total_score}/100 ===

EXACT SCORING THRESHOLDS (you MUST reference these exactly, do NOT substitute your own "industry benchmarks"):
1. IRR vs. Hurdle ({score.irr_score}/30 points): IRR of {format_pct(score.irr_value)}
   - Scoring: 0 points at IRR ≤0%, scales to 5 points at 10% IRR, then scales to full 30 points at 25%+ IRR
2. MOIC ({score.moic_score}/20 points): MOIC of {format_multiple(score.moic_value)}
   - Scoring: 0 points at MOIC ≤1.0x, scales linearly to full 20 points at 3.0x+ MOIC
3. Debt Service Coverage ({score.debt_service_score}/25 points): FCF covers mandatory amortization by {score.min_coverage_ratio:.2f}x minimum in Year {score.min_coverage_year}
   - Scoring: 0 points at coverage ≤1.0x, scales linearly to full 25 points at 2.5x+ coverage
4. Leverage Reduction ({score.leverage_reduction_score}/15 points): {score.leverage_reduction_pct*100:.1f}% reduction over hold period
   - Scoring: 0 points at 0% reduction, scales linearly to full 15 points at 50%+ reduction
5. Data Quality ({score.data_quality_score}/10 points): {score.defaulted_field_count} field(s) with defaulted/substituted values
   - Scoring: 10 points minus 2 points per defaulted/substituted field (minimum 0)

"""

    # Add data quality disclosures
    if defaulted_fields:
        prompt += "=== DATA QUALITY DISCLOSURES (MUST be mentioned in report) ===\n"
        for field_name, note in defaulted_fields:
            prompt += f"- {field_name}: {note}\n"
        prompt += "\n"

    prompt += """=== YOUR TASK ===

Write a concise LBO analysis report with the following sections:

1. **Deal Summary** (one paragraph): What company is being acquired, at what valuation, and how the purchase is financed (debt vs. equity split).

2. **Sources & Uses** (brief, factual): State the total transaction cost, how much is debt, how much is equity.

3. **Operating and Debt Trajectory**: Describe how revenue, EBITDA, and leverage evolve over the hold period. Note debt paydown progress.

4. **Exit Scenario and Returns**: State the exit valuation, equity proceeds, IRR, and MOIC.

5. **Sensitivity Context**: One or two sentences summarizing how returns vary across the entry/exit multiple range tested. Do NOT reproduce the full grid.

6. **Feasibility Assessment** (CRITICAL SCOPE):
   - Reference the computed feasibility score (""" + f"{score.total_score}" + """/100)
   - Explain what's driving the score using the component breakdown (e.g., "the score is held down primarily by weak debt service coverage" or "strong IRR contributes significantly to the score")
   - ONLY assess whether the deal's CAPITAL STRUCTURE is sound: debt service coverage, leverage trajectory, IRR/MOIC vs. typical PE benchmarks, data completeness
   - Do NOT speculate about deal execution probability, regulatory approval, shareholder approval, financing availability, or negotiation outcomes. The model has no information about any of that and you must not invent commentary on it.

7. **Verdict**: A short, plain statement summarizing whether this capital structure appears workable based on the numbers.

IMPORTANT INSTRUCTIONS:
- Stay factual and grounded in the numbers provided. Do not invent, estimate, or speculate about anything not in the data.
- Any field listed under "DATA QUALITY DISCLOSURES" above MUST be explicitly and plainly disclosed in your report (e.g., "Total debt was not reported and was assumed to be $0").
- Keep it concise: a few short paragraphs total, not an exhaustive essay.
- No investment advice framing. Do not say "you should do this deal" or "this is a good investment." Describe what the numbers show.
- Format currency values consistently (use $XXB for billions, $XXM for millions).
- CRITICAL: When explaining score components in the Feasibility Assessment section, you MUST use ONLY the exact scoring thresholds provided above (e.g., "1.0x to 2.5x" for debt service, "1.0x to 3.0x" for MOIC, "10% to 25%" for IRR, "0% to 50%" for leverage reduction). Do NOT substitute your own knowledge of "typical PE benchmarks" or "industry standards" - describe this tool's actual scoring logic accurately.
"""

    return prompt


# =============================================================================
# MAIN REPORT GENERATION FUNCTION
# =============================================================================

def generate_report(
    xlsx_path: str,
    provider: str,
    api_key: str,
    cleanup_temp: bool = True
) -> Tuple[str, ExtractedData, ScoreBreakdown]:
    """
    Generate a complete LBO analysis report from a workbook.

    Args:
        xlsx_path: Path to the generated .xlsx file
        provider: LLM provider ("anthropic" or "openai")
        api_key: API key for the provider
        cleanup_temp: Whether to clean up temporary files (default True)

    Returns:
        Tuple of (report_text, extracted_data, score_breakdown)
    """
    # Step 1: Recalculate the workbook
    print(f"Recalculating workbook with LibreOffice...")
    recalc_path = recalculate_workbook(xlsx_path)
    temp_dir = os.path.dirname(recalc_path)

    try:
        # Step 2: Extract data with provenance
        print(f"Extracting data from recalculated workbook...")
        data = extract_data_from_workbook(recalc_path)

        # Step 3: Compute feasibility score
        print(f"Computing feasibility score...")
        score = compute_feasibility_score(data)

        # Step 4: Generate narrative
        print(f"Generating narrative with {provider}...")
        report = generate_narrative(data, score, provider, api_key)

        return report, data, score

    finally:
        # Clean up temp directory
        if cleanup_temp and temp_dir:
            shutil.rmtree(temp_dir, ignore_errors=True)


# =============================================================================
# CLI / TEST HARNESS
# =============================================================================

def print_score_breakdown(ticker: str, score: ScoreBreakdown):
    """Print a detailed score breakdown for verification."""
    print(f"\n{'='*60}")
    print(f"FEASIBILITY SCORE BREAKDOWN: {ticker}")
    print(f"{'='*60}")
    print(f"1. IRR Score:              {score.irr_score:5.1f} / 30  (IRR = {score.irr_value*100:.1f}%)")
    print(f"2. MOIC Score:             {score.moic_score:5.1f} / 20  (MOIC = {score.moic_value:.2f}x)")
    print(f"3. Debt Service Score:     {score.debt_service_score:5.1f} / 25  (FCF/MandatoryAmort = {score.min_coverage_ratio:.2f}x min in Year {score.min_coverage_year})")
    print(f"4. Leverage Reduction:     {score.leverage_reduction_score:5.1f} / 15  (Reduction = {score.leverage_reduction_pct*100:.1f}%)")
    print(f"5. Data Quality Score:     {score.data_quality_score:5.1f} / 10  ({score.defaulted_field_count} defaulted fields)")
    print(f"{'-'*60}")
    print(f"   TOTAL SCORE:            {score.total_score:5.1f} / 100")
    print(f"{'='*60}")


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Generate LBO analysis report from workbook")
    parser.add_argument("xlsx_path", nargs="?", help="Path to .xlsx file")
    parser.add_argument("--provider", choices=["anthropic", "openai", "gemini"], default="anthropic",
                        help="LLM provider (default: anthropic)")
    parser.add_argument("--api-key", help="API key (or set ANTHROPIC_API_KEY / OPENAI_API_KEY / GEMINI_API_KEY env var)")
    parser.add_argument("--test", action="store_true", help="Run test suite on AAPL, CCL, FIZZ")
    parser.add_argument("--extract-only", action="store_true",
                        help="Only extract data and compute score (no LLM call)")
    parser.add_argument("--show-prompt", action="store_true",
                        help="Show the LLM prompt that would be sent (no API call)")

    args = parser.parse_args()

    # Get API key from argument or environment
    api_key = args.api_key
    if not api_key:
        if args.provider == "anthropic":
            api_key = os.environ.get("ANTHROPIC_API_KEY")
        elif args.provider == "openai":
            api_key = os.environ.get("OPENAI_API_KEY")
        elif args.provider == "gemini":
            api_key = os.environ.get("GEMINI_API_KEY")

    if args.test:
        # Test mode - process all three test tickers
        output_dir = Path(__file__).parent / "output"

        # Find the most recent files for each ticker
        test_tickers = ["AAPL", "CCL", "FIZZ"]

        for ticker in test_tickers:
            print(f"\n{'#'*70}")
            print(f"# PROCESSING: {ticker}")
            print(f"{'#'*70}")

            # Find most recent file for this ticker
            files = sorted(output_dir.glob(f"LBO_{ticker}_*.xlsx"), reverse=True)
            if not files:
                print(f"  ERROR: No file found for {ticker}")
                continue

            xlsx_path = str(files[0])
            print(f"Using: {xlsx_path}")

            try:
                # Recalculate and extract
                recalc_path = recalculate_workbook(xlsx_path)
                temp_dir = os.path.dirname(recalc_path)

                try:
                    data = extract_data_from_workbook(recalc_path)
                    score = compute_feasibility_score(data)

                    # Print score breakdown
                    print_score_breakdown(ticker, score)

                    # Print defaulted fields
                    defaulted = data.get_defaulted_fields()
                    if defaulted:
                        print(f"\nDefaulted/Substituted Fields:")
                        for field_name, note in defaulted:
                            print(f"  - {field_name}: {note}")

                    # Show prompt if requested
                    if args.show_prompt:
                        prompt = build_report_prompt(data, score)
                        print(f"\n{'='*60}")
                        print(f"LLM PROMPT FOR: {ticker}")
                        print(f"{'='*60}")
                        print(prompt)
                    # Generate report if API key provided
                    elif api_key and not args.extract_only:
                        print(f"\nGenerating narrative report...")
                        report = generate_narrative(data, score, args.provider, api_key)
                        print(f"\n{'='*60}")
                        print(f"GENERATED REPORT: {ticker}")
                        print(f"{'='*60}")
                        print(sanitize_for_console(report))
                    elif args.extract_only:
                        print(f"\n(Skipping LLM call - extract-only mode)")
                    else:
                        print(f"\n(Skipping LLM call - no API key provided)")
                        print(f"Set {args.provider.upper()}_API_KEY or use --api-key")

                finally:
                    shutil.rmtree(temp_dir, ignore_errors=True)

            except Exception as e:
                print(f"  ERROR: {e}")
                import traceback
                traceback.print_exc()

    elif args.xlsx_path:
        # Single file mode
        if not api_key and not args.extract_only and not args.show_prompt:
            print(f"ERROR: API key required. Set {args.provider.upper()}_API_KEY or use --api-key")
            sys.exit(1)

        try:
            if args.extract_only or args.show_prompt:
                recalc_path = recalculate_workbook(args.xlsx_path)
                temp_dir = os.path.dirname(recalc_path)
                try:
                    data = extract_data_from_workbook(recalc_path)
                    score = compute_feasibility_score(data)
                    print_score_breakdown("", score)

                    if args.show_prompt:
                        prompt = build_report_prompt(data, score)
                        print("\n" + "="*60)
                        print("LLM PROMPT")
                        print("="*60)
                        print(prompt)
                    else:
                        print("\nExtracted Data (JSON):")
                        print(json.dumps(data.to_dict(), indent=2, default=str))
                finally:
                    shutil.rmtree(temp_dir, ignore_errors=True)
            else:
                report, data, score = generate_report(
                    args.xlsx_path, args.provider, api_key
                )
                print_score_breakdown("", score)
                print("\n" + "="*60)
                print("GENERATED REPORT")
                print("="*60)
                print(sanitize_for_console(report))

        except Exception as e:
            print(f"ERROR: {e}")
            sys.exit(1)

    else:
        parser.print_help()
