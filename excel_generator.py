"""
AIO LBO Excel Generator

Generates Excel workbooks for LBO models from validated SEC EDGAR data.
Currently implements: Assumptions tab, Sources & Uses tab.
"""

import os
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.comments import Comment


# =============================================================================
# STYLE CONSTANTS - IB/PE Modeling Convention
# =============================================================================

# Blue font: hardcoded inputs (API-fetched data + user assumptions)
BLUE_FONT = Font(name='Calibri', size=11, color='0000CC', bold=False)

# Black font: formulas (calculated from other cells)
BLACK_FONT = Font(name='Calibri', size=11, color='000000', bold=False)

# Bold fonts for headers
BLUE_FONT_BOLD = Font(name='Calibri', size=11, color='0000CC', bold=True)
BLACK_FONT_BOLD = Font(name='Calibri', size=11, color='000000', bold=True)

# Section header styling
SECTION_HEADER_FONT = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
SECTION_HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

# Sub-section header styling
SUBSECTION_FONT = Font(name='Calibri', size=11, bold=True, color='000000')
SUBSECTION_FILL = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')

# Defaulted value flag (soft requirement defaulted by validator)
DEFAULTED_FILL = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')

# Error fill (for balance check)
ERROR_FILL = PatternFill(start_color='FF6666', end_color='FF6666', fill_type='solid')

# Implausible calculation flag (e.g., COVID-distorted growth rate)
# Orange fill - distinct from yellow "defaulted" and matches Interest Expense placeholder
IMPLAUSIBLE_CALC_FILL = PatternFill(start_color='FFB366', end_color='FFB366', fill_type='solid')

# Borders
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Alignments
RIGHT_ALIGN = Alignment(horizontal='right')
LEFT_ALIGN = Alignment(horizontal='left')
CENTER_ALIGN = Alignment(horizontal='center')

# Number formats
CURRENCY_FORMAT = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'
CURRENCY_MILLIONS_FORMAT = '_($* #,##0.0,,"M"_);_($* (#,##0.0,,"M");_($* "-"??_);_(@_)'
PERCENT_FORMAT = '0.0%'
MULTIPLE_FORMAT = '0.0"x"'
NUMBER_FORMAT = '#,##0'
SHARES_FORMAT = '#,##0.0,,"M"'  # Shares in millions


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def safe_get(data: dict, *keys, default=None):
    """Safely navigate nested dict structure."""
    current = data
    for key in keys:
        if isinstance(current, dict) and key in current:
            current = current[key]
        else:
            return default
    return current if current is not None else default


def get_most_recent_fy(summary: dict) -> Optional[int]:
    """Get the most recent fiscal year from summary."""
    fiscal_years = summary.get('fiscal_years', {})
    if not fiscal_years:
        return None
    return max(fiscal_years.keys())


def get_fy_data(summary: dict, fy: int) -> dict:
    """Get data for a specific fiscal year."""
    return summary.get('fiscal_years', {}).get(fy, {})


def calculate_revenue_growth_rate(summary: dict) -> tuple[float, bool, str]:
    """
    Calculate revenue growth rate from 5-year history using MEDIAN of YoY changes.

    Uses median instead of mean to reduce sensitivity to extreme outlier years
    (e.g., COVID-era revenue collapses/recoveries).

    Applies a sanity band: if calculated rate is outside -15% to +25%,
    falls back to a conservative 3% default.

    Returns:
        (growth_rate, is_fallback, comment)
        - growth_rate: the rate to use (either calculated median or 3% fallback)
        - is_fallback: True if the sanity band was triggered and 3% fallback was used
        - comment: description of how the value was derived
    """
    import statistics

    fiscal_years = summary.get('fiscal_years', {})
    if len(fiscal_years) < 2:
        return 0.03, True, "Insufficient historical data — defaulted to 3%"

    # Sort years in ascending order
    sorted_years = sorted(fiscal_years.keys())
    growth_rates = []

    for i in range(1, len(sorted_years)):
        prev_year = sorted_years[i - 1]
        curr_year = sorted_years[i]
        prev_rev = fiscal_years.get(prev_year, {}).get('revenue')
        curr_rev = fiscal_years.get(curr_year, {}).get('revenue')

        if prev_rev and curr_rev and prev_rev > 0:
            growth = (curr_rev - prev_rev) / prev_rev
            growth_rates.append(growth)

    if not growth_rates:
        return 0.03, True, "No valid revenue history — defaulted to 3%"

    # Use MEDIAN instead of mean to reduce outlier sensitivity
    median_growth = statistics.median(growth_rates)

    # Sanity band: -15% to +25%
    GROWTH_MIN = -0.15
    GROWTH_MAX = 0.25
    FALLBACK_RATE = 0.03

    if median_growth < GROWTH_MIN or median_growth > GROWTH_MAX:
        return (
            FALLBACK_RATE,
            True,
            f"Historical growth rate ({median_growth*100:.1f}%) was implausible "
            f"(likely COVID-era distortion) — defaulted to 3%, please review and adjust manually."
        )

    return (
        median_growth,
        False,
        f"Default: median of historical YoY growth rates ({median_growth*100:.1f}%)"
    )


def is_defaulted_field(summary: dict, validation_result: dict, field_name: str) -> bool:
    """Check if a field was defaulted by the validator."""
    defaults = validation_result.get('defaults_applied', [])

    # Check if field is mentioned in defaults
    for default_msg in defaults:
        if field_name.lower() in default_msg.lower():
            return True

    # Also check if the field is None in the most recent FY
    most_recent_fy = get_most_recent_fy(summary)
    if most_recent_fy:
        fy_data = get_fy_data(summary, most_recent_fy)
        if fy_data.get(field_name) is None:
            return True

    return False


def create_defined_name(wb: Workbook, name: str, sheet_name: str, cell: str):
    """Create a workbook-level defined name for a cell."""
    from openpyxl.workbook.defined_name import DefinedName

    # Clean the name to be Excel-compatible
    clean_name = name.replace(' ', '_').replace('%', 'Pct').replace('/', '_')
    clean_name = ''.join(c for c in clean_name if c.isalnum() or c == '_')

    # Ensure name doesn't start with a number
    if clean_name[0].isdigit():
        clean_name = '_' + clean_name

    # Build the cell reference - parse cell like "C15" to "$C$15"
    col_letter = ''.join(c for c in cell if c.isalpha())
    row_num = ''.join(c for c in cell if c.isdigit())
    cell_ref = f"${col_letter}${row_num}"

    # Create and add the defined name
    defn = DefinedName(clean_name, attr_text=f"'{sheet_name}'!{cell_ref}")
    wb.defined_names[clean_name] = defn

    return clean_name


# =============================================================================
# ASSUMPTIONS TAB BUILDER
# =============================================================================

def build_assumptions_tab(wb: Workbook, summary: dict, validation_result: dict) -> dict:
    """
    Build the Assumptions tab.

    Returns a dict mapping friendly names to Excel cell references for use by other tabs.
    """
    ws = wb.create_sheet("Assumptions", 0)

    # Column widths
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18

    # Track named ranges
    named_ranges = {}
    current_row = 1

    # ==========================================================================
    # SECTION A: COMPANY DATA
    # ==========================================================================

    # Section header
    ws.cell(row=current_row, column=2, value="SECTION A: COMPANY DATA (Fetched from SEC EDGAR)")
    ws.cell(row=current_row, column=2).font = SECTION_HEADER_FONT
    ws.cell(row=current_row, column=2).fill = SECTION_HEADER_FILL
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=7)
    current_row += 2

    # Company identification
    ws.cell(row=current_row, column=2, value="Company Name").font = BLUE_FONT
    ws.cell(row=current_row, column=3, value=summary.get('company_name', 'Unknown')).font = BLUE_FONT
    current_row += 1

    ws.cell(row=current_row, column=2, value="Ticker").font = BLUE_FONT
    ws.cell(row=current_row, column=3, value=summary.get('ticker', '')).font = BLUE_FONT
    current_row += 1

    ws.cell(row=current_row, column=2, value="CIK").font = BLUE_FONT
    ws.cell(row=current_row, column=3, value=summary.get('cik', '')).font = BLUE_FONT
    current_row += 1

    ws.cell(row=current_row, column=2, value="SIC Code / Sector").font = BLUE_FONT
    sic_code = summary.get('sic_code', '')
    sic_desc = summary.get('sic_description', '')
    ws.cell(row=current_row, column=3, value=f"{sic_code} - {sic_desc}").font = BLUE_FONT
    current_row += 2

    # Current market data
    ws.cell(row=current_row, column=2, value="Market Data").font = SUBSECTION_FONT
    ws.cell(row=current_row, column=2).fill = SUBSECTION_FILL
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
    current_row += 1

    # Current share price
    ws.cell(row=current_row, column=2, value="Current Share Price").font = BLUE_FONT
    price_cell = ws.cell(row=current_row, column=3)
    current_price = summary.get('current_price')
    if current_price is not None:
        price_cell.value = current_price
        price_cell.number_format = '$#,##0.00'
    else:
        price_cell.value = 0
        price_cell.comment = Comment("Price not available - update manually", "System")
    price_cell.font = BLUE_FONT
    price_ref = f"C{current_row}"
    named_ranges['CurrentPrice'] = create_defined_name(wb, 'CurrentPrice', 'Assumptions', price_ref)
    current_row += 1

    # Shares outstanding
    ws.cell(row=current_row, column=2, value="Shares Outstanding").font = BLUE_FONT
    shares_cell = ws.cell(row=current_row, column=3)
    shares = summary.get('shares_outstanding')
    if shares is not None:
        shares_cell.value = shares
        shares_cell.number_format = '#,##0'
    else:
        shares_cell.value = 0
        shares_cell.comment = Comment("Shares not available - update manually", "System")
    shares_cell.font = BLUE_FONT
    shares_ref = f"C{current_row}"
    named_ranges['SharesOutstanding'] = create_defined_name(wb, 'SharesOutstanding', 'Assumptions', shares_ref)
    current_row += 1

    # Market cap (FORMULA - black font)
    ws.cell(row=current_row, column=2, value="Market Cap").font = BLACK_FONT
    mktcap_cell = ws.cell(row=current_row, column=3)
    mktcap_cell.value = f"={price_ref}*{shares_ref}"
    mktcap_cell.font = BLACK_FONT
    mktcap_cell.number_format = CURRENCY_FORMAT
    named_ranges['MarketCap'] = create_defined_name(wb, 'MarketCap', 'Assumptions', f"C{current_row}")
    current_row += 2

    # ==========================================================================
    # 5-YEAR HISTORICAL FINANCIALS
    # ==========================================================================

    ws.cell(row=current_row, column=2, value="Historical Financials (5-Year)").font = SUBSECTION_FONT
    ws.cell(row=current_row, column=2).fill = SUBSECTION_FILL
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=7)
    current_row += 1

    # Get fiscal years sorted (most recent first)
    fiscal_years = summary.get('fiscal_years', {})
    sorted_fys = sorted(fiscal_years.keys(), reverse=True)[:5]

    # Header row with fiscal years
    ws.cell(row=current_row, column=2, value="Metric").font = BLUE_FONT_BOLD
    for i, fy in enumerate(sorted_fys):
        col = 3 + i
        ws.cell(row=current_row, column=col, value=f"FY{fy}").font = BLUE_FONT_BOLD
        ws.cell(row=current_row, column=col).alignment = CENTER_ALIGN
    current_row += 1

    # Revenue row
    revenue_row = current_row
    ws.cell(row=current_row, column=2, value="Revenue").font = BLUE_FONT
    for i, fy in enumerate(sorted_fys):
        col = 3 + i
        cell = ws.cell(row=current_row, column=col)
        fy_data = fiscal_years.get(fy, {})
        revenue = fy_data.get('revenue')
        if revenue is not None:
            cell.value = revenue
        else:
            cell.value = 0
            cell.fill = DEFAULTED_FILL
            cell.comment = Comment("Defaulted — not reported", "System")
        cell.font = BLUE_FONT
        cell.number_format = CURRENCY_FORMAT
    current_row += 1

    # Operating Income row
    ws.cell(row=current_row, column=2, value="Operating Income").font = BLUE_FONT
    for i, fy in enumerate(sorted_fys):
        col = 3 + i
        cell = ws.cell(row=current_row, column=col)
        fy_data = fiscal_years.get(fy, {})
        op_income = fy_data.get('operating_income')
        if op_income is not None:
            cell.value = op_income
        else:
            cell.value = 0
            cell.fill = DEFAULTED_FILL
            cell.comment = Comment("Defaulted — not reported", "System")
        cell.font = BLUE_FONT
        cell.number_format = CURRENCY_FORMAT
    op_income_row = current_row
    current_row += 1

    # D&A row
    ws.cell(row=current_row, column=2, value="D&A").font = BLUE_FONT
    for i, fy in enumerate(sorted_fys):
        col = 3 + i
        cell = ws.cell(row=current_row, column=col)
        fy_data = fiscal_years.get(fy, {})
        da = fy_data.get('da')
        if da is not None:
            cell.value = da
        else:
            cell.value = 0
            cell.fill = DEFAULTED_FILL
            cell.comment = Comment("Defaulted — not reported", "System")
        cell.font = BLUE_FONT
        cell.number_format = CURRENCY_FORMAT
    da_row = current_row
    current_row += 1

    # EBITDA row (FORMULA - black font)
    ebitda_row = current_row
    ws.cell(row=current_row, column=2, value="EBITDA (Calculated)").font = BLACK_FONT
    for i, fy in enumerate(sorted_fys):
        col = 3 + i
        col_letter = get_column_letter(col)
        cell = ws.cell(row=current_row, column=col)
        cell.value = f"={col_letter}{op_income_row}+{col_letter}{da_row}"
        cell.font = BLACK_FONT
        cell.number_format = CURRENCY_FORMAT
    current_row += 2

    # Balance sheet items - most recent only
    ws.cell(row=current_row, column=2, value="Most Recent Balance Sheet Items").font = SUBSECTION_FONT
    ws.cell(row=current_row, column=2).fill = SUBSECTION_FILL
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
    current_row += 1

    most_recent_fy = sorted_fys[0] if sorted_fys else None
    most_recent_data = fiscal_years.get(most_recent_fy, {}) if most_recent_fy else {}

    # Total Debt
    ws.cell(row=current_row, column=2, value="Total Debt").font = BLUE_FONT
    debt_cell = ws.cell(row=current_row, column=3)
    total_debt = most_recent_data.get('total_debt')
    debt_defaulted = is_defaulted_field(summary, validation_result, 'total_debt')
    if total_debt is not None:
        debt_cell.value = total_debt
    else:
        debt_cell.value = 0
        debt_defaulted = True
    if debt_defaulted:
        debt_cell.fill = DEFAULTED_FILL
        debt_cell.comment = Comment("Defaulted — not reported (may be debt-free)", "System")
    debt_cell.font = BLUE_FONT
    debt_cell.number_format = CURRENCY_FORMAT
    named_ranges['TotalDebt'] = create_defined_name(wb, 'TotalDebt', 'Assumptions', f"C{current_row}")
    current_row += 1

    # Cash
    ws.cell(row=current_row, column=2, value="Cash & Equivalents").font = BLUE_FONT
    cash_cell = ws.cell(row=current_row, column=3)
    cash = most_recent_data.get('cash')
    cash_defaulted = is_defaulted_field(summary, validation_result, 'cash')
    if cash is not None:
        cash_cell.value = cash
    else:
        cash_cell.value = 0
        cash_defaulted = True
    if cash_defaulted:
        cash_cell.fill = DEFAULTED_FILL
        cash_cell.comment = Comment("Defaulted — not reported", "System")
    cash_cell.font = BLUE_FONT
    cash_cell.number_format = CURRENCY_FORMAT
    named_ranges['Cash'] = create_defined_name(wb, 'Cash', 'Assumptions', f"C{current_row}")
    current_row += 1

    # CapEx
    ws.cell(row=current_row, column=2, value="Capital Expenditures").font = BLUE_FONT
    capex_cell = ws.cell(row=current_row, column=3)
    capex = most_recent_data.get('capex')
    capex_defaulted = is_defaulted_field(summary, validation_result, 'capex')
    if capex is not None:
        capex_cell.value = capex
    else:
        capex_cell.value = 0
        capex_defaulted = True
    if capex_defaulted:
        capex_cell.fill = DEFAULTED_FILL
        capex_cell.comment = Comment("Defaulted — not reported", "System")
    capex_cell.font = BLUE_FONT
    capex_cell.number_format = CURRENCY_FORMAT
    named_ranges['CapEx'] = create_defined_name(wb, 'CapEx', 'Assumptions', f"C{current_row}")
    capex_row = current_row
    current_row += 3

    # ==========================================================================
    # SECTION B: DEAL ASSUMPTIONS
    # ==========================================================================

    ws.cell(row=current_row, column=2, value="SECTION B: DEAL ASSUMPTIONS (User Inputs)")
    ws.cell(row=current_row, column=2).font = SECTION_HEADER_FONT
    ws.cell(row=current_row, column=2).fill = SECTION_HEADER_FILL
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)
    current_row += 2

    # Calculate default values
    growth_rate, growth_is_fallback, growth_comment = calculate_revenue_growth_rate(summary)

    # Most recent revenue and EBITDA for ratio calculations
    most_recent_revenue = most_recent_data.get('revenue', 0) or 1  # Avoid div by zero
    most_recent_ebitda = most_recent_data.get('ebitda_calculated')
    if most_recent_ebitda is None:
        op_inc = most_recent_data.get('operating_income', 0) or 0
        da = most_recent_data.get('da', 0) or 0
        most_recent_ebitda = op_inc + da

    most_recent_capex = most_recent_data.get('capex', 0) or 0
    most_recent_da = most_recent_data.get('da', 0) or 0

    ebitda_margin = most_recent_ebitda / most_recent_revenue if most_recent_revenue > 0 else 0.15
    capex_pct = most_recent_capex / most_recent_revenue if most_recent_revenue > 0 else 0.02
    da_pct = most_recent_da / most_recent_revenue if most_recent_revenue > 0 else 0.03

    # If capex was defaulted, use 2%
    if capex_defaulted:
        capex_pct = 0.02

    # --- Deal Assumptions Inputs ---

    # Revenue Growth Rate
    ws.cell(row=current_row, column=2, value="Revenue Growth Rate").font = BLUE_FONT
    growth_cell = ws.cell(row=current_row, column=3)
    growth_cell.value = growth_rate
    growth_cell.font = BLUE_FONT
    growth_cell.number_format = PERCENT_FORMAT
    growth_cell.comment = Comment(growth_comment, "System")
    if growth_is_fallback:
        # Apply orange fill for implausible/fallback calculation
        growth_cell.fill = IMPLAUSIBLE_CALC_FILL
    named_ranges['RevenueGrowthRate'] = create_defined_name(wb, 'RevenueGrowthRate', 'Assumptions', f"C{current_row}")
    current_row += 1

    # EBITDA Margin
    ws.cell(row=current_row, column=2, value="EBITDA Margin %").font = BLUE_FONT
    margin_cell = ws.cell(row=current_row, column=3)
    margin_cell.value = ebitda_margin
    margin_cell.font = BLUE_FONT
    margin_cell.number_format = PERCENT_FORMAT
    margin_cell.comment = Comment(f"Default: most recent FY EBITDA / Revenue", "System")
    named_ranges['EBITDAMargin'] = create_defined_name(wb, 'EBITDAMargin', 'Assumptions', f"C{current_row}")
    current_row += 1

    # Entry EV/EBITDA Multiple
    ws.cell(row=current_row, column=2, value="Entry EV/EBITDA Multiple").font = BLUE_FONT
    entry_mult_cell = ws.cell(row=current_row, column=3)
    entry_mult_cell.value = 8.0
    entry_mult_cell.font = BLUE_FONT
    entry_mult_cell.number_format = MULTIPLE_FORMAT
    entry_mult_cell.comment = Comment("Default: 8.0x placeholder — adjust based on sector comps", "System")
    entry_mult_ref = f"C{current_row}"
    named_ranges['EntryMultiple'] = create_defined_name(wb, 'EntryMultiple', 'Assumptions', entry_mult_ref)
    current_row += 1

    # Offer Premium
    ws.cell(row=current_row, column=2, value="Offer Premium %").font = BLUE_FONT
    premium_cell = ws.cell(row=current_row, column=3)
    premium_cell.value = 0.25
    premium_cell.font = BLUE_FONT
    premium_cell.number_format = PERCENT_FORMAT
    named_ranges['OfferPremium'] = create_defined_name(wb, 'OfferPremium', 'Assumptions', f"C{current_row}")
    current_row += 1

    # Leverage Multiple
    ws.cell(row=current_row, column=2, value="Leverage Multiple").font = BLUE_FONT
    leverage_cell = ws.cell(row=current_row, column=3)
    leverage_cell.value = 5.5
    leverage_cell.font = BLUE_FONT
    leverage_cell.number_format = MULTIPLE_FORMAT
    named_ranges['LeverageMultiple'] = create_defined_name(wb, 'LeverageMultiple', 'Assumptions', f"C{current_row}")
    current_row += 1

    # Interest Rate
    ws.cell(row=current_row, column=2, value="Interest Rate").font = BLUE_FONT
    interest_cell = ws.cell(row=current_row, column=3)
    interest_cell.value = 0.08
    interest_cell.font = BLUE_FONT
    interest_cell.number_format = PERCENT_FORMAT
    named_ranges['InterestRate'] = create_defined_name(wb, 'InterestRate', 'Assumptions', f"C{current_row}")
    current_row += 1

    # Tax Rate
    ws.cell(row=current_row, column=2, value="Tax Rate").font = BLUE_FONT
    tax_cell = ws.cell(row=current_row, column=3)
    tax_cell.value = 0.25
    tax_cell.font = BLUE_FONT
    tax_cell.number_format = PERCENT_FORMAT
    named_ranges['TaxRate'] = create_defined_name(wb, 'TaxRate', 'Assumptions', f"C{current_row}")
    current_row += 1

    # CapEx %
    ws.cell(row=current_row, column=2, value="CapEx % of Revenue").font = BLUE_FONT
    capex_pct_cell = ws.cell(row=current_row, column=3)
    capex_pct_cell.value = capex_pct
    capex_pct_cell.font = BLUE_FONT
    capex_pct_cell.number_format = PERCENT_FORMAT
    if capex_defaulted:
        capex_pct_cell.comment = Comment("Default: 2% (CapEx not reported)", "System")
    else:
        capex_pct_cell.comment = Comment("Default: most recent FY CapEx / Revenue", "System")
    named_ranges['CapExPct'] = create_defined_name(wb, 'CapExPct', 'Assumptions', f"C{current_row}")
    current_row += 1

    # D&A %
    ws.cell(row=current_row, column=2, value="D&A % of Revenue").font = BLUE_FONT
    da_pct_cell = ws.cell(row=current_row, column=3)
    da_pct_cell.value = da_pct
    da_pct_cell.font = BLUE_FONT
    da_pct_cell.number_format = PERCENT_FORMAT
    da_pct_cell.comment = Comment("Default: most recent FY D&A / Revenue", "System")
    named_ranges['DAPct'] = create_defined_name(wb, 'DAPct', 'Assumptions', f"C{current_row}")
    current_row += 1

    # Change in NWC %
    ws.cell(row=current_row, column=2, value="Change in NWC % of Revenue").font = BLUE_FONT
    nwc_cell = ws.cell(row=current_row, column=3)
    nwc_cell.value = 0.0
    nwc_cell.font = BLUE_FONT
    nwc_cell.number_format = PERCENT_FORMAT
    named_ranges['NWCPct'] = create_defined_name(wb, 'NWCPct', 'Assumptions', f"C{current_row}")
    current_row += 1

    # Exit Year
    ws.cell(row=current_row, column=2, value="Exit Year").font = BLUE_FONT
    exit_year_cell = ws.cell(row=current_row, column=3)
    exit_year_cell.value = 5
    exit_year_cell.font = BLUE_FONT
    exit_year_cell.number_format = '0'
    named_ranges['ExitYear'] = create_defined_name(wb, 'ExitYear', 'Assumptions', f"C{current_row}")
    current_row += 1

    # Exit EV/EBITDA Multiple (FORMULA by default - references entry multiple)
    ws.cell(row=current_row, column=2, value="Exit EV/EBITDA Multiple").font = BLACK_FONT
    exit_mult_cell = ws.cell(row=current_row, column=3)
    exit_mult_cell.value = f"={entry_mult_ref}"  # Formula referencing entry multiple
    exit_mult_cell.font = BLACK_FONT
    exit_mult_cell.number_format = MULTIPLE_FORMAT
    exit_mult_cell.comment = Comment(
        "Default: equals Entry Multiple (no expansion/contraction). "
        "Override with hardcoded value to model multiple expansion/contraction.",
        "System"
    )
    named_ranges['ExitMultiple'] = create_defined_name(wb, 'ExitMultiple', 'Assumptions', f"C{current_row}")
    current_row += 1

    # Transaction Fee %
    ws.cell(row=current_row, column=2, value="Transaction Fee %").font = BLUE_FONT
    txn_fee_cell = ws.cell(row=current_row, column=3)
    txn_fee_cell.value = 0.02
    txn_fee_cell.font = BLUE_FONT
    txn_fee_cell.number_format = PERCENT_FORMAT
    named_ranges['TransactionFeePct'] = create_defined_name(wb, 'TransactionFeePct', 'Assumptions', f"C{current_row}")
    current_row += 1

    # Mandatory Amortization %
    ws.cell(row=current_row, column=2, value="Mandatory Amortization %").font = BLUE_FONT
    amort_cell = ws.cell(row=current_row, column=3)
    amort_cell.value = 0.01
    amort_cell.font = BLUE_FONT
    amort_cell.number_format = PERCENT_FORMAT
    named_ranges['AmortizationPct'] = create_defined_name(wb, 'AmortizationPct', 'Assumptions', f"C{current_row}")
    current_row += 2

    # Store Entry Revenue and EBITDA references (most recent FY values for projections)
    # Column C is most recent fiscal year
    if sorted_fys:
        most_recent_col = get_column_letter(3)  # Column C is most recent
        entry_revenue_ref = f"{most_recent_col}{revenue_row}"
        entry_ebitda_ref = f"{most_recent_col}{ebitda_row}"
        named_ranges['EntryRevenue'] = create_defined_name(wb, 'EntryRevenue', 'Assumptions', entry_revenue_ref)
        named_ranges['EntryEBITDA'] = create_defined_name(wb, 'EntryEBITDA', 'Assumptions', entry_ebitda_ref)

    return named_ranges


# =============================================================================
# SOURCES & USES TAB BUILDER
# =============================================================================

def build_sources_uses_tab(wb: Workbook, named_ranges: dict):
    """
    Build the Sources & Uses tab.

    All values are formulas (black font) referencing the Assumptions tab.
    """
    ws = wb.create_sheet("Sources & Uses", 1)

    # Column widths
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15

    current_row = 1

    # ==========================================================================
    # USES OF FUNDS
    # ==========================================================================

    ws.cell(row=current_row, column=2, value="USES OF FUNDS")
    ws.cell(row=current_row, column=2).font = SECTION_HEADER_FONT
    ws.cell(row=current_row, column=2).fill = SECTION_HEADER_FILL
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
    current_row += 2

    # Purchase Enterprise Value = Entry EBITDA × Entry Multiple
    ws.cell(row=current_row, column=2, value="Purchase Enterprise Value").font = BLACK_FONT
    pev_cell = ws.cell(row=current_row, column=3)
    pev_cell.value = "=EntryEBITDA*EntryMultiple"
    pev_cell.font = BLACK_FONT
    pev_cell.number_format = CURRENCY_FORMAT
    pev_row = current_row
    current_row += 1

    # Transaction Fees = Purchase EV × Fee %
    ws.cell(row=current_row, column=2, value="Transaction Fees").font = BLACK_FONT
    fees_cell = ws.cell(row=current_row, column=3)
    fees_cell.value = f"=C{pev_row}*TransactionFeePct"
    fees_cell.font = BLACK_FONT
    fees_cell.number_format = CURRENCY_FORMAT
    fees_row = current_row
    current_row += 1

    # Total Uses = sum
    current_row += 1
    ws.cell(row=current_row, column=2, value="Total Uses").font = BLACK_FONT_BOLD
    total_uses_cell = ws.cell(row=current_row, column=3)
    total_uses_cell.value = f"=C{pev_row}+C{fees_row}"
    total_uses_cell.font = BLACK_FONT_BOLD
    total_uses_cell.number_format = CURRENCY_FORMAT
    total_uses_cell.border = Border(top=Side(style='thin'), bottom=Side(style='double'))
    total_uses_row = current_row

    # Create named range for Total Uses
    from openpyxl.workbook.defined_name import DefinedName
    defn = DefinedName('TotalUses', attr_text=f"'Sources & Uses'!C{total_uses_row}")
    wb.defined_names['TotalUses'] = defn

    current_row += 3

    # ==========================================================================
    # SOURCES OF FUNDS
    # ==========================================================================

    ws.cell(row=current_row, column=2, value="SOURCES OF FUNDS")
    ws.cell(row=current_row, column=2).font = SECTION_HEADER_FONT
    ws.cell(row=current_row, column=2).fill = SECTION_HEADER_FILL
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
    current_row += 2

    # New Debt Raised = Entry EBITDA × Leverage Multiple
    ws.cell(row=current_row, column=2, value="New Debt Raised").font = BLACK_FONT
    debt_cell = ws.cell(row=current_row, column=3)
    debt_cell.value = "=EntryEBITDA*LeverageMultiple"
    debt_cell.font = BLACK_FONT
    debt_cell.number_format = CURRENCY_FORMAT
    debt_row = current_row

    # Create named range for NewDebtRaised
    defn = DefinedName('NewDebtRaised', attr_text=f"'Sources & Uses'!C{debt_row}")
    wb.defined_names['NewDebtRaised'] = defn

    current_row += 1

    # Sponsor Equity Check = Total Uses - New Debt Raised
    ws.cell(row=current_row, column=2, value="Sponsor Equity Check").font = BLACK_FONT
    equity_cell = ws.cell(row=current_row, column=3)
    equity_cell.value = f"=C{total_uses_row}-C{debt_row}"
    equity_cell.font = BLACK_FONT
    equity_cell.number_format = CURRENCY_FORMAT
    equity_row = current_row

    # Create named range for SponsorEquity
    defn = DefinedName('SponsorEquity', attr_text=f"'Sources & Uses'!C{equity_row}")
    wb.defined_names['SponsorEquity'] = defn

    current_row += 1

    # Total Sources = New Debt + Sponsor Equity
    current_row += 1
    ws.cell(row=current_row, column=2, value="Total Sources").font = BLACK_FONT_BOLD
    total_sources_cell = ws.cell(row=current_row, column=3)
    total_sources_cell.value = f"=C{debt_row}+C{equity_row}"
    total_sources_cell.font = BLACK_FONT_BOLD
    total_sources_cell.number_format = CURRENCY_FORMAT
    total_sources_cell.border = Border(top=Side(style='thin'), bottom=Side(style='double'))
    total_sources_row = current_row

    # Create named range for TotalSources
    defn = DefinedName('TotalSources', attr_text=f"'Sources & Uses'!C{total_sources_row}")
    wb.defined_names['TotalSources'] = defn

    current_row += 3

    # ==========================================================================
    # BALANCE CHECK
    # ==========================================================================

    ws.cell(row=current_row, column=2, value="Balance Check").font = SUBSECTION_FONT
    ws.cell(row=current_row, column=2).fill = SUBSECTION_FILL
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
    current_row += 1

    ws.cell(row=current_row, column=2, value="Sources - Uses (should = 0)").font = BLACK_FONT
    balance_cell = ws.cell(row=current_row, column=3)
    balance_cell.value = f"=C{total_sources_row}-C{total_uses_row}"
    balance_cell.font = BLACK_FONT
    balance_cell.number_format = CURRENCY_FORMAT
    balance_row = current_row

    # Add conditional formatting: red fill if not zero
    ws.conditional_formatting.add(
        f"C{balance_row}",
        FormulaRule(
            formula=[f"C{balance_row}<>0"],
            fill=ERROR_FILL
        )
    )

    # Create named range for BalanceCheck
    defn = DefinedName('BalanceCheck', attr_text=f"'Sources & Uses'!C{balance_row}")
    wb.defined_names['BalanceCheck'] = defn

    return {
        'pev_row': pev_row,
        'total_uses_row': total_uses_row,
        'debt_row': debt_row,
        'equity_row': equity_row,
        'total_sources_row': total_sources_row,
        'balance_row': balance_row
    }


# =============================================================================
# OPERATING MODEL TAB BUILDER
# =============================================================================

# Placeholder styling for Interest Expense (will be wired to Debt Schedule later)
PLACEHOLDER_FILL = PatternFill(start_color='FFB366', end_color='FFB366', fill_type='solid')


def build_operating_model_tab(wb: Workbook, summary: dict, named_ranges: dict) -> dict:
    """
    Build the Operating Model tab.

    Projects Revenue through Free Cash Flow for each year from Year 1 to Exit Year.
    All values are formulas (black font) referencing the Assumptions tab.

    Returns a dict with row references for use by future tabs (Debt Schedule, Returns).
    """
    ws = wb.create_sheet("Operating Model", 2)

    # Get exit year from assumptions (default 5)
    exit_year = 5  # Default - will be read from named range in formulas

    # Get most recent fiscal year for calendar year labels
    fiscal_years = summary.get('fiscal_years', {})
    if fiscal_years:
        most_recent_fy = max(fiscal_years.keys())
    else:
        most_recent_fy = 2024  # Fallback

    # Column widths
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 35

    # Year columns start at C
    for i in range(exit_year):
        col_letter = get_column_letter(3 + i)
        ws.column_dimensions[col_letter].width = 16

    current_row = 1

    # ==========================================================================
    # HEADER
    # ==========================================================================

    ws.cell(row=current_row, column=2, value="OPERATING MODEL (Projected)")
    ws.cell(row=current_row, column=2).font = SECTION_HEADER_FONT
    ws.cell(row=current_row, column=2).fill = SECTION_HEADER_FILL
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=2 + exit_year)
    current_row += 2

    # Year headers
    ws.cell(row=current_row, column=2, value="").font = BLACK_FONT_BOLD
    for year in range(1, exit_year + 1):
        col = 2 + year  # Column C = Year 1, D = Year 2, etc.
        calendar_year = most_recent_fy + year
        cell = ws.cell(row=current_row, column=col)
        cell.value = f"Year {year}\n(FY{calendar_year})"
        cell.font = BLACK_FONT_BOLD
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    header_row = current_row
    current_row += 2

    # ==========================================================================
    # INCOME STATEMENT SECTION
    # ==========================================================================

    ws.cell(row=current_row, column=2, value="Income Statement").font = SUBSECTION_FONT
    ws.cell(row=current_row, column=2).fill = SUBSECTION_FILL
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=2 + exit_year)
    current_row += 1

    # --- Revenue ---
    revenue_row = current_row
    ws.cell(row=current_row, column=2, value="Revenue").font = BLACK_FONT
    for year in range(1, exit_year + 1):
        col = 2 + year
        col_letter = get_column_letter(col)
        cell = ws.cell(row=current_row, column=col)

        if year == 1:
            # Year 1: Most recent FY Revenue × (1 + Growth Rate)
            # Need to find where revenue is in Assumptions - it's in column C of the historical section
            # We need a named range for Entry Revenue - let's reference the historical data
            # Actually, we should create a named range for this in build_assumptions_tab
            # For now, reference the cell directly using the most recent revenue
            # The Entry EBITDA is at a known location, but Entry Revenue isn't named
            # Let's use a formula that finds it: we'll create EntryRevenue named range
            cell.value = "=EntryRevenue*(1+RevenueGrowthRate)"
        else:
            # Year 2+: Prior year Revenue × (1 + Growth Rate)
            prev_col = get_column_letter(col - 1)
            cell.value = f"={prev_col}{revenue_row}*(1+RevenueGrowthRate)"

        cell.font = BLACK_FONT
        cell.number_format = CURRENCY_FORMAT
    current_row += 1

    # --- EBITDA ---
    ebitda_row = current_row
    ws.cell(row=current_row, column=2, value="EBITDA").font = BLACK_FONT
    for year in range(1, exit_year + 1):
        col = 2 + year
        col_letter = get_column_letter(col)
        cell = ws.cell(row=current_row, column=col)
        cell.value = f"={col_letter}{revenue_row}*EBITDAMargin"
        cell.font = BLACK_FONT
        cell.number_format = CURRENCY_FORMAT
    current_row += 1

    # --- D&A ---
    da_row = current_row
    ws.cell(row=current_row, column=2, value="Less: D&A").font = BLACK_FONT
    for year in range(1, exit_year + 1):
        col = 2 + year
        col_letter = get_column_letter(col)
        cell = ws.cell(row=current_row, column=col)
        cell.value = f"={col_letter}{revenue_row}*DAPct"
        cell.font = BLACK_FONT
        cell.number_format = CURRENCY_FORMAT
    current_row += 1

    # --- EBIT ---
    ebit_row = current_row
    ws.cell(row=current_row, column=2, value="EBIT").font = BLACK_FONT_BOLD
    for year in range(1, exit_year + 1):
        col = 2 + year
        col_letter = get_column_letter(col)
        cell = ws.cell(row=current_row, column=col)
        cell.value = f"={col_letter}{ebitda_row}-{col_letter}{da_row}"
        cell.font = BLACK_FONT_BOLD
        cell.number_format = CURRENCY_FORMAT
        cell.border = Border(top=Side(style='thin'))
    current_row += 1

    # --- Interest Expense (PLACEHOLDER) ---
    interest_row = current_row
    ws.cell(row=current_row, column=2, value="Less: Interest Expense").font = BLACK_FONT
    for year in range(1, exit_year + 1):
        col = 2 + year
        cell = ws.cell(row=current_row, column=col)
        cell.value = 0  # Hardcoded placeholder
        cell.font = BLACK_FONT
        cell.number_format = CURRENCY_FORMAT
        cell.fill = PLACEHOLDER_FILL
        cell.comment = Comment(
            "PLACEHOLDER - will be replaced with a reference to the Debt Schedule tab once built.",
            "System"
        )
    current_row += 1

    # --- Pre-Tax Income ---
    pretax_row = current_row
    ws.cell(row=current_row, column=2, value="Pre-Tax Income").font = BLACK_FONT
    for year in range(1, exit_year + 1):
        col = 2 + year
        col_letter = get_column_letter(col)
        cell = ws.cell(row=current_row, column=col)
        cell.value = f"={col_letter}{ebit_row}-{col_letter}{interest_row}"
        cell.font = BLACK_FONT
        cell.number_format = CURRENCY_FORMAT
    current_row += 1

    # --- Taxes (floored at 0) ---
    taxes_row = current_row
    ws.cell(row=current_row, column=2, value="Less: Taxes").font = BLACK_FONT
    for year in range(1, exit_year + 1):
        col = 2 + year
        col_letter = get_column_letter(col)
        cell = ws.cell(row=current_row, column=col)
        # MAX(0, Pre-Tax Income × TaxRate) - no negative taxes
        cell.value = f"=MAX(0,{col_letter}{pretax_row}*TaxRate)"
        cell.font = BLACK_FONT
        cell.number_format = CURRENCY_FORMAT
    current_row += 1

    # --- Net Income ---
    net_income_row = current_row
    ws.cell(row=current_row, column=2, value="Net Income").font = BLACK_FONT_BOLD
    for year in range(1, exit_year + 1):
        col = 2 + year
        col_letter = get_column_letter(col)
        cell = ws.cell(row=current_row, column=col)
        cell.value = f"={col_letter}{pretax_row}-{col_letter}{taxes_row}"
        cell.font = BLACK_FONT_BOLD
        cell.number_format = CURRENCY_FORMAT
        cell.border = Border(top=Side(style='thin'), bottom=Side(style='double'))
    current_row += 2

    # ==========================================================================
    # FREE CASH FLOW SECTION
    # ==========================================================================

    ws.cell(row=current_row, column=2, value="Free Cash Flow Build").font = SUBSECTION_FONT
    ws.cell(row=current_row, column=2).fill = SUBSECTION_FILL
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=2 + exit_year)
    current_row += 1

    # --- Net Income (reference) ---
    fcf_net_income_row = current_row
    ws.cell(row=current_row, column=2, value="Net Income").font = BLACK_FONT
    for year in range(1, exit_year + 1):
        col = 2 + year
        col_letter = get_column_letter(col)
        cell = ws.cell(row=current_row, column=col)
        cell.value = f"={col_letter}{net_income_row}"
        cell.font = BLACK_FONT
        cell.number_format = CURRENCY_FORMAT
    current_row += 1

    # --- Plus: D&A (add back) ---
    fcf_da_row = current_row
    ws.cell(row=current_row, column=2, value="Plus: D&A (non-cash add-back)").font = BLACK_FONT
    for year in range(1, exit_year + 1):
        col = 2 + year
        col_letter = get_column_letter(col)
        cell = ws.cell(row=current_row, column=col)
        cell.value = f"={col_letter}{da_row}"
        cell.font = BLACK_FONT
        cell.number_format = CURRENCY_FORMAT
    current_row += 1

    # --- Less: CapEx ---
    capex_row = current_row
    ws.cell(row=current_row, column=2, value="Less: CapEx").font = BLACK_FONT
    for year in range(1, exit_year + 1):
        col = 2 + year
        col_letter = get_column_letter(col)
        cell = ws.cell(row=current_row, column=col)
        cell.value = f"={col_letter}{revenue_row}*CapExPct"
        cell.font = BLACK_FONT
        cell.number_format = CURRENCY_FORMAT
    current_row += 1

    # --- Less: Change in NWC ---
    nwc_row = current_row
    ws.cell(row=current_row, column=2, value="Less: Change in NWC").font = BLACK_FONT
    for year in range(1, exit_year + 1):
        col = 2 + year
        col_letter = get_column_letter(col)
        cell = ws.cell(row=current_row, column=col)
        # NWC change = Revenue × NWCPct (consumes cash when positive)
        cell.value = f"={col_letter}{revenue_row}*NWCPct"
        cell.font = BLACK_FONT
        cell.number_format = CURRENCY_FORMAT
    current_row += 1

    # --- Free Cash Flow ---
    fcf_row = current_row
    ws.cell(row=current_row, column=2, value="Free Cash Flow for Debt Paydown").font = BLACK_FONT_BOLD
    for year in range(1, exit_year + 1):
        col = 2 + year
        col_letter = get_column_letter(col)
        cell = ws.cell(row=current_row, column=col)
        # FCF = Net Income + D&A - CapEx - Change in NWC
        cell.value = f"={col_letter}{fcf_net_income_row}+{col_letter}{fcf_da_row}-{col_letter}{capex_row}-{col_letter}{nwc_row}"
        cell.font = BLACK_FONT_BOLD
        cell.number_format = CURRENCY_FORMAT
        cell.border = Border(top=Side(style='thin'), bottom=Side(style='double'))
    current_row += 2

    # ==========================================================================
    # KEY METRICS (for quick reference)
    # ==========================================================================

    ws.cell(row=current_row, column=2, value="Key Metrics").font = SUBSECTION_FONT
    ws.cell(row=current_row, column=2).fill = SUBSECTION_FILL
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=2 + exit_year)
    current_row += 1

    # --- EBITDA Margin % ---
    ws.cell(row=current_row, column=2, value="EBITDA Margin %").font = BLACK_FONT
    for year in range(1, exit_year + 1):
        col = 2 + year
        col_letter = get_column_letter(col)
        cell = ws.cell(row=current_row, column=col)
        cell.value = f"={col_letter}{ebitda_row}/{col_letter}{revenue_row}"
        cell.font = BLACK_FONT
        cell.number_format = PERCENT_FORMAT
    current_row += 1

    # --- FCF Conversion % (FCF / EBITDA) ---
    ws.cell(row=current_row, column=2, value="FCF Conversion % (FCF/EBITDA)").font = BLACK_FONT
    for year in range(1, exit_year + 1):
        col = 2 + year
        col_letter = get_column_letter(col)
        cell = ws.cell(row=current_row, column=col)
        cell.value = f"=IF({col_letter}{ebitda_row}=0,0,{col_letter}{fcf_row}/{col_letter}{ebitda_row})"
        cell.font = BLACK_FONT
        cell.number_format = PERCENT_FORMAT

    # Create named ranges for key rows that will be used by Debt Schedule and Returns tabs
    from openpyxl.workbook.defined_name import DefinedName

    # Create named ranges for the Operating Model outputs
    # These use the exit year column (last year) for exit calculations
    exit_col = get_column_letter(2 + exit_year)

    # Exit Year EBITDA (for exit valuation)
    defn = DefinedName('ExitEBITDA', attr_text=f"'Operating Model'!{exit_col}{ebitda_row}")
    wb.defined_names['ExitEBITDA'] = defn

    # Create row references for each year's FCF (for debt paydown)
    # Store these as a dict to return
    row_refs = {
        'revenue_row': revenue_row,
        'ebitda_row': ebitda_row,
        'da_row': da_row,
        'ebit_row': ebit_row,
        'interest_row': interest_row,
        'pretax_row': pretax_row,
        'taxes_row': taxes_row,
        'net_income_row': net_income_row,
        'capex_row': capex_row,
        'nwc_row': nwc_row,
        'fcf_row': fcf_row,
        'exit_year': exit_year,
        'year_start_col': 3,  # Column C
    }

    return row_refs


# =============================================================================
# MAIN GENERATOR FUNCTION
# =============================================================================

def generate_workbook(validated_summary: dict, output_dir: str = None) -> str:
    """
    Generate an LBO model Excel workbook from a validated summary.

    Args:
        validated_summary: Dict containing both the summary data and validation result.
                          Expected structure:
                          {
                              'summary': {...},  # From fetch_ticker_data()
                              'validation': {...}  # From validate()
                          }
                          OR just the summary dict directly (validation defaults to empty)
        output_dir: Directory to save the file (defaults to current directory)

    Returns:
        Path to the generated .xlsx file
    """
    # Handle both formats: wrapped or direct summary
    if 'summary' in validated_summary and 'validation' in validated_summary:
        summary = validated_summary['summary']
        validation = validated_summary['validation']
    else:
        summary = validated_summary
        validation = {}

    # Create workbook
    wb = Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # Build Assumptions tab
    named_ranges = build_assumptions_tab(wb, summary, validation)

    # Build Sources & Uses tab
    su_refs = build_sources_uses_tab(wb, named_ranges)

    # Build Operating Model tab
    op_model_refs = build_operating_model_tab(wb, summary, named_ranges)

    # Generate filename
    ticker = summary.get('ticker', 'UNKNOWN')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"LBO_{ticker}_{timestamp}.xlsx"

    # Determine output path
    if output_dir:
        output_path = Path(output_dir) / filename
    else:
        output_path = Path.cwd() / filename

    # Save workbook
    wb.save(str(output_path))

    print(f"Generated workbook: {output_path}")

    return str(output_path)


# =============================================================================
# FORMULA VERIFICATION (since openpyxl doesn't evaluate formulas)
# =============================================================================

def verify_workbook_formulas(filepath: str) -> dict:
    """
    Load a generated workbook and print the formula strings for verification.

    Since openpyxl writes formulas as text and can't evaluate them,
    this function reads back the formulas so you can spot-check the logic.

    Returns a dict with key formula strings.
    """
    wb = load_workbook(filepath)

    result = {
        'filepath': filepath,
        'sheets': list(wb.sheetnames),
        'assumptions_formulas': {},
        'sources_uses_formulas': {},
        'operating_model_formulas': {},
        'named_ranges': {}
    }

    # Get named ranges
    for name in wb.defined_names:
        defn = wb.defined_names[name]
        result['named_ranges'][name] = str(defn.attr_text)

    # Check Assumptions tab
    if 'Assumptions' in wb.sheetnames:
        ws = wb['Assumptions']
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    cell_ref = f"{get_column_letter(cell.column)}{cell.row}"
                    result['assumptions_formulas'][cell_ref] = cell.value

    # Check Sources & Uses tab
    if 'Sources & Uses' in wb.sheetnames:
        ws = wb['Sources & Uses']
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    cell_ref = f"{get_column_letter(cell.column)}{cell.row}"
                    result['sources_uses_formulas'][cell_ref] = cell.value

    # Check Operating Model tab
    if 'Operating Model' in wb.sheetnames:
        ws = wb['Operating Model']
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    cell_ref = f"{get_column_letter(cell.column)}{cell.row}"
                    result['operating_model_formulas'][cell_ref] = cell.value

    return result


def print_verification_report(verification: dict):
    """Print a human-readable verification report."""
    print(f"\n{'='*70}")
    print(f"FORMULA VERIFICATION REPORT")
    print(f"{'='*70}")
    print(f"File: {verification['filepath']}")
    print(f"Sheets: {', '.join(verification['sheets'])}")

    print(f"\n--- Named Ranges ({len(verification['named_ranges'])}) ---")
    for name, ref in sorted(verification['named_ranges'].items()):
        print(f"  {name}: {ref}")

    print(f"\n--- Assumptions Tab Formulas ---")
    for cell, formula in verification['assumptions_formulas'].items():
        print(f"  {cell}: {formula}")

    print(f"\n--- Sources & Uses Tab Formulas ---")
    for cell, formula in verification['sources_uses_formulas'].items():
        print(f"  {cell}: {formula}")

    # Key formulas to verify
    print(f"\n--- KEY FORMULAS TO VERIFY ---")
    su_formulas = verification['sources_uses_formulas']

    print("\nExpected logic:")
    print("  Purchase EV = EntryEBITDA × EntryMultiple")
    print("  Transaction Fees = Purchase EV × TransactionFeePct")
    print("  Total Uses = Purchase EV + Fees")
    print("  New Debt = EntryEBITDA × LeverageMultiple")
    print("  Sponsor Equity = Total Uses - New Debt")
    print("  Total Sources = New Debt + Sponsor Equity")
    print("  Balance Check = Total Sources - Total Uses (should = 0)")


def evaluate_model_values(summary: dict, validation: dict) -> dict:
    """
    Evaluate the key model values manually to verify formula logic.

    Since openpyxl can't evaluate Excel formulas, this function computes
    the expected values using the same logic as the formulas. This allows
    verification without opening Excel.

    Returns a dict with computed values.
    """
    fiscal_years = summary.get('fiscal_years', {})
    if not fiscal_years:
        return {"error": "No fiscal year data"}

    most_recent_fy = max(fiscal_years.keys())
    fy_data = fiscal_years[most_recent_fy]

    # Get Entry EBITDA (most recent FY)
    ebitda = fy_data.get('ebitda_calculated')
    if ebitda is None:
        op_inc = fy_data.get('operating_income', 0) or 0
        da = fy_data.get('da', 0) or 0
        ebitda = op_inc + da

    # Deal assumptions (defaults from build_assumptions_tab)
    entry_multiple = 8.0
    leverage_multiple = 5.5
    transaction_fee_pct = 0.02

    # Computed values
    purchase_ev = ebitda * entry_multiple
    transaction_fees = purchase_ev * transaction_fee_pct
    total_uses = purchase_ev + transaction_fees
    new_debt = ebitda * leverage_multiple
    sponsor_equity = total_uses - new_debt
    total_sources = new_debt + sponsor_equity
    balance_check = total_sources - total_uses

    return {
        "ticker": summary.get('ticker', 'Unknown'),
        "most_recent_fy": most_recent_fy,
        "entry_ebitda": ebitda,
        "entry_multiple": entry_multiple,
        "leverage_multiple": leverage_multiple,
        "transaction_fee_pct": transaction_fee_pct,
        "purchase_ev": purchase_ev,
        "transaction_fees": transaction_fees,
        "total_uses": total_uses,
        "new_debt": new_debt,
        "sponsor_equity": sponsor_equity,
        "total_sources": total_sources,
        "balance_check": balance_check,
        "balance_ok": abs(balance_check) < 0.01  # Should be exactly 0
    }


def print_computed_values(values: dict):
    """Print the computed model values in a readable format."""
    if "error" in values:
        print(f"  ERROR: {values['error']}")
        return

    def fmt_currency(val):
        if val >= 1e9:
            return f"${val/1e9:,.2f}B"
        elif val >= 1e6:
            return f"${val/1e6:,.2f}M"
        else:
            return f"${val:,.0f}"

    print(f"\n  Computed Values for {values['ticker']} (FY{values['most_recent_fy']}):")
    print(f"  {'-'*50}")
    print(f"  Entry EBITDA:      {fmt_currency(values['entry_ebitda'])}")
    print(f"  Entry Multiple:    {values['entry_multiple']:.1f}x")
    print(f"  Leverage Multiple: {values['leverage_multiple']:.1f}x")
    print(f"  Transaction Fee %: {values['transaction_fee_pct']*100:.1f}%")
    print(f"  {'-'*50}")
    print(f"  USES:")
    print(f"    Purchase EV:     {fmt_currency(values['purchase_ev'])}")
    print(f"    Transaction Fees:{fmt_currency(values['transaction_fees'])}")
    print(f"    Total Uses:      {fmt_currency(values['total_uses'])}")
    print(f"  {'-'*50}")
    print(f"  SOURCES:")
    print(f"    New Debt:        {fmt_currency(values['new_debt'])}")
    print(f"    Sponsor Equity:  {fmt_currency(values['sponsor_equity'])}")
    print(f"    Total Sources:   {fmt_currency(values['total_sources'])}")
    print(f"  {'-'*50}")
    print(f"  Balance Check:     {fmt_currency(values['balance_check'])} {'[OK]' if values['balance_ok'] else '[ERROR]'}")


def evaluate_operating_model(summary: dict, validation: dict, exit_year: int = 5) -> dict:
    """
    Evaluate the Operating Model projections manually to verify formula logic.

    Returns a dict with projected values for each year.
    """
    fiscal_years = summary.get('fiscal_years', {})
    if not fiscal_years:
        return {"error": "No fiscal year data"}

    most_recent_fy = max(fiscal_years.keys())
    fy_data = fiscal_years[most_recent_fy]

    # Entry values
    entry_revenue = fy_data.get('revenue', 0) or 0
    entry_ebitda = fy_data.get('ebitda_calculated')
    if entry_ebitda is None:
        op_inc = fy_data.get('operating_income', 0) or 0
        da = fy_data.get('da', 0) or 0
        entry_ebitda = op_inc + da

    # Calculate growth rate using the same logic as the Excel generator
    revenue_growth, growth_is_fallback, _ = calculate_revenue_growth_rate(summary)

    # Assumptions (defaults)
    ebitda_margin = entry_ebitda / entry_revenue if entry_revenue > 0 else 0.15
    da_pct = (fy_data.get('da', 0) or 0) / entry_revenue if entry_revenue > 0 else 0.03
    capex_pct = (fy_data.get('capex', 0) or 0) / entry_revenue if entry_revenue > 0 else 0.02
    nwc_pct = 0.0
    tax_rate = 0.25
    interest_expense = 0  # Placeholder

    projections = []
    prev_revenue = entry_revenue

    for year in range(1, exit_year + 1):
        # Revenue
        revenue = prev_revenue * (1 + revenue_growth)

        # EBITDA
        ebitda = revenue * ebitda_margin

        # D&A
        da = revenue * da_pct

        # EBIT
        ebit = ebitda - da

        # Pre-Tax Income
        pretax = ebit - interest_expense

        # Taxes (floored at 0)
        taxes = max(0, pretax * tax_rate)

        # Net Income
        net_income = pretax - taxes

        # FCF components
        capex = revenue * capex_pct
        nwc_change = revenue * nwc_pct

        # FCF
        fcf = net_income + da - capex - nwc_change

        projections.append({
            'year': year,
            'calendar_year': most_recent_fy + year,
            'revenue': revenue,
            'ebitda': ebitda,
            'da': da,
            'ebit': ebit,
            'interest': interest_expense,
            'pretax': pretax,
            'taxes': taxes,
            'net_income': net_income,
            'capex': capex,
            'nwc_change': nwc_change,
            'fcf': fcf,
        })

        prev_revenue = revenue

    return {
        'ticker': summary.get('ticker', 'Unknown'),
        'most_recent_fy': most_recent_fy,
        'entry_revenue': entry_revenue,
        'entry_ebitda': entry_ebitda,
        'revenue_growth': revenue_growth,
        'growth_is_fallback': growth_is_fallback,
        'ebitda_margin': ebitda_margin,
        'da_pct': da_pct,
        'capex_pct': capex_pct,
        'tax_rate': tax_rate,
        'projections': projections,
    }


def print_operating_model(op_model: dict):
    """Print Operating Model projections in a readable format."""
    if "error" in op_model:
        print(f"  ERROR: {op_model['error']}")
        return

    def fmt_currency(val):
        if abs(val) >= 1e9:
            return f"${val/1e9:,.2f}B"
        elif abs(val) >= 1e6:
            return f"${val/1e6:,.2f}M"
        else:
            return f"${val:,.0f}"

    print(f"\n  Operating Model for {op_model['ticker']}:")
    print(f"  Entry FY: {op_model['most_recent_fy']}")
    print(f"  Entry Revenue: {fmt_currency(op_model['entry_revenue'])}")
    growth_flag = " [FALLBACK - implausible calc]" if op_model.get('growth_is_fallback') else ""
    print(f"  Revenue Growth: {op_model['revenue_growth']*100:.1f}%{growth_flag}")
    print(f"  EBITDA Margin: {op_model['ebitda_margin']*100:.1f}%")
    print()

    # Print header
    years = op_model['projections']
    header = f"  {'Metric':<20}"
    for p in [years[0], years[-1]]:  # Year 1 and final year
        header += f"  Year {p['year']:>3}"
    print(header)
    print("  " + "-" * (20 + 12 * 2))

    # Print key rows
    metrics = [
        ('Revenue', 'revenue'),
        ('EBITDA', 'ebitda'),
        ('D&A', 'da'),
        ('EBIT', 'ebit'),
        ('Interest', 'interest'),
        ('Pre-Tax', 'pretax'),
        ('Taxes', 'taxes'),
        ('Net Income', 'net_income'),
        ('CapEx', 'capex'),
        ('NWC Change', 'nwc_change'),
        ('FCF', 'fcf'),
    ]

    for label, key in metrics:
        row = f"  {label:<20}"
        for p in [years[0], years[-1]]:
            row += f"  {fmt_currency(p[key]):>10}"
        print(row)

    # FCF conversion check
    year1 = years[0]
    final = years[-1]
    print()
    print(f"  FCF/EBITDA Year 1: {year1['fcf']/year1['ebitda']*100:.1f}%")
    print(f"  FCF/EBITDA Year {final['year']}: {final['fcf']/final['ebitda']*100:.1f}%")


# =============================================================================
# MOCK DATA FOR TESTING (when SEC credentials not available)
# =============================================================================

def get_mock_data(ticker: str) -> tuple:
    """Generate mock data for testing Excel generation without SEC API access."""
    mock_data = {
        "AAPL": {
            "summary": {
                "ticker": "AAPL",
                "company_name": "Apple Inc.",
                "cik": 320193,
                "sic_code": "3571",
                "sic_description": "Electronic Computers",
                "fiscal_year_end": "09-30",
                "current_price": 195.50,
                "shares_outstanding": 15500000000,
                "fiscal_years": {
                    2024: {
                        "revenue": 383285000000,
                        "operating_income": 114301000000,
                        "da": 11200000000,
                        "ebitda_calculated": 125501000000,
                        "total_debt": 111088000000,
                        "cash": 61555000000,
                        "capex": 9959000000,
                        "missing": []
                    },
                    2023: {
                        "revenue": 394328000000,
                        "operating_income": 114301000000,
                        "da": 11519000000,
                        "ebitda_calculated": 125820000000,
                        "total_debt": 109614000000,
                        "cash": 29965000000,
                        "capex": 10959000000,
                        "missing": []
                    },
                    2022: {
                        "revenue": 365817000000,
                        "operating_income": 119437000000,
                        "da": 11104000000,
                        "ebitda_calculated": 130541000000,
                        "total_debt": 111824000000,
                        "cash": 23646000000,
                        "capex": 10708000000,
                        "missing": []
                    },
                    2021: {
                        "revenue": 274515000000,
                        "operating_income": 108949000000,
                        "da": 11284000000,
                        "ebitda_calculated": 120233000000,
                        "total_debt": 119381000000,
                        "cash": 34940000000,
                        "capex": 11085000000,
                        "missing": []
                    },
                    2020: {
                        "revenue": 274515000000,
                        "operating_income": 66288000000,
                        "da": 11056000000,
                        "ebitda_calculated": 77344000000,
                        "total_debt": 112043000000,
                        "cash": 38016000000,
                        "capex": 7309000000,
                        "missing": []
                    }
                }
            },
            "validation": {
                "status": "pass",
                "missing_hard": [],
                "missing_soft": [],
                "disqualifying_reasons": [],
                "sector_excluded": False,
                "sector_excluded_reason": None,
                "defaults_applied": []
            }
        },
        "FIZZ": {
            "summary": {
                "ticker": "FIZZ",
                "company_name": "National Beverage Corp.",
                "cik": 879993,
                "sic_code": "2086",
                "sic_description": "Bottled & Canned Soft Drinks",
                "fiscal_year_end": "04-30",
                "current_price": 45.20,
                "shares_outstanding": 46200000,
                "fiscal_years": {
                    2024: {
                        "revenue": 1184800000,
                        "operating_income": 174200000,
                        "da": 32500000,
                        "ebitda_calculated": 206700000,
                        "total_debt": None,  # Will be defaulted
                        "cash": 145000000,
                        "capex": 28000000,
                        "missing": ["total_debt"]
                    },
                    2023: {
                        "revenue": 1171500000,
                        "operating_income": 201500000,
                        "da": 30200000,
                        "ebitda_calculated": 231700000,
                        "total_debt": None,
                        "cash": 167000000,
                        "capex": 24000000,
                        "missing": ["total_debt"]
                    },
                    2022: {
                        "revenue": 1148600000,
                        "operating_income": 183400000,
                        "da": 28900000,
                        "ebitda_calculated": 212300000,
                        "total_debt": None,
                        "cash": 189000000,
                        "capex": 31000000,
                        "missing": ["total_debt"]
                    },
                    2021: {
                        "revenue": 1075000000,
                        "operating_income": 205000000,
                        "da": 26800000,
                        "ebitda_calculated": 231800000,
                        "total_debt": None,
                        "cash": 203000000,
                        "capex": 21000000,
                        "missing": ["total_debt"]
                    },
                    2020: {
                        "revenue": 1014000000,
                        "operating_income": 181000000,
                        "da": 25100000,
                        "ebitda_calculated": 206100000,
                        "total_debt": None,
                        "cash": 196000000,
                        "capex": 19000000,
                        "missing": ["total_debt"]
                    }
                }
            },
            "validation": {
                "status": "degraded",
                "missing_hard": [],
                "missing_soft": ["Total debt data is missing for FY2024. This could indicate a genuinely debt-free company."],
                "disqualifying_reasons": [],
                "sector_excluded": False,
                "sector_excluded_reason": None,
                "defaults_applied": ["Total debt will default to $0 — verify this is accurate before proceeding."]
            }
        },
        "CCL": {
            "summary": {
                "ticker": "CCL",
                "company_name": "Carnival Corporation",
                "cik": 815097,
                "sic_code": "4481",
                "sic_description": "Deep Sea Foreign Transportation of Passengers",
                "fiscal_year_end": "11-30",
                "current_price": 18.75,
                "shares_outstanding": 1270000000,
                "fiscal_years": {
                    2024: {
                        "revenue": 24530000000,
                        "operating_income": 3265000000,
                        "da": 2800000000,
                        "ebitda_calculated": 6065000000,
                        "total_debt": 31500000000,  # Heavy debt
                        "cash": 1850000000,
                        "capex": 4200000000,
                        "missing": []
                    },
                    2023: {
                        "revenue": 21630000000,
                        "operating_income": 1958000000,
                        "da": 2600000000,
                        "ebitda_calculated": 4558000000,
                        "total_debt": 33200000000,
                        "cash": 2700000000,
                        "capex": 3500000000,
                        "missing": []
                    },
                    2022: {
                        "revenue": 12169000000,
                        "operating_income": -3865000000,
                        "da": 2500000000,
                        "ebitda_calculated": -1365000000,
                        "total_debt": 35400000000,
                        "cash": 7320000000,
                        "capex": 2100000000,
                        "missing": []
                    },
                    2021: {
                        "revenue": 1908000000,
                        "operating_income": -7810000000,
                        "da": 2380000000,
                        "ebitda_calculated": -5430000000,
                        "total_debt": 33100000000,
                        "cash": 9530000000,
                        "capex": 1400000000,
                        "missing": []
                    },
                    2020: {
                        "revenue": 5595000000,
                        "operating_income": -8878000000,
                        "da": 2340000000,
                        "ebitda_calculated": -6538000000,
                        "total_debt": 26650000000,
                        "cash": 9500000000,
                        "capex": 2600000000,
                        "missing": []
                    }
                }
            },
            "validation": {
                "status": "pass",
                "missing_hard": [],
                "missing_soft": [],
                "disqualifying_reasons": [],
                "sector_excluded": False,
                "sector_excluded_reason": None,
                "defaults_applied": []
            }
        }
    }

    if ticker in mock_data:
        return mock_data[ticker]["summary"], mock_data[ticker]["validation"]
    return None, None


# =============================================================================
# TEST HARNESS
# =============================================================================

if __name__ == "__main__":
    import sys
    sys.path.insert(0, str(Path(__file__).parent))

    # Check for --mock flag to use mock data
    use_mock = "--mock" in sys.argv

    # Allow command-line override for SEC email
    # Usage: python excel_generator.py [sec_email]
    # Or: python excel_generator.py --mock (use mock data)
    for arg in sys.argv[1:]:
        if arg != "--mock" and "@" in arg:
            os.environ["SEC_CONTACT_EMAIL"] = arg

    print("="*70)
    print("AIO LBO EXCEL GENERATOR TEST")
    print("="*70)

    # Test tickers
    test_tickers = ["AAPL", "FIZZ", "CCL"]

    print(f"\nTest tickers:")
    print("  AAPL - Large-cap, clean data")
    print("  FIZZ - Small-cap, debt likely defaulted to $0")
    print("  CCL  - Debt-heavy, should have real debt figures")

    if use_mock:
        print("\n[MOCK MODE] Using mock data for testing")
    else:
        # Import SEC modules only if not using mock
        try:
            from sec_edgar_test import fetch_ticker_data, get_sec_user_agent, get_twelve_data_api_key
            from validator import validate

            sec_user_agent = get_sec_user_agent()
            try:
                twelve_data_key = get_twelve_data_api_key()
            except ValueError:
                twelve_data_key = ""
                print("Note: No Twelve Data key - current price will be missing")
        except ValueError as e:
            print(f"\nWARNING: {e}")
            print("Falling back to mock data mode.\n")
            use_mock = True

    output_dir = Path(__file__).parent / "output"
    output_dir.mkdir(exist_ok=True)

    for ticker in test_tickers:
        print(f"\n{'='*70}")
        print(f"PROCESSING: {ticker}")
        print("="*70)

        if use_mock:
            # Use mock data
            print(f"\n[1/3] Loading mock data...")
            summary, validation = get_mock_data(ticker)
            if not summary:
                print(f"  ERROR: No mock data for {ticker}")
                continue
            print(f"  Loaded mock data for {summary['company_name']}")
        else:
            # Fetch real data
            print(f"\n[1/3] Fetching data from SEC EDGAR...")
            summary = fetch_ticker_data(
                ticker,
                sec_user_agent,
                twelve_data_key,
                verbose=False,
                skip_price=True
            )

            if not summary:
                print(f"  ERROR: Failed to fetch data for {ticker}")
                continue

            # Validate
            print(f"[2/3] Validating...")
            validation = validate(summary)

        print(f"  Status: {validation['status'].upper()}")
        if validation['defaults_applied']:
            print(f"  Defaults applied:")
            for d in validation['defaults_applied']:
                print(f"    - {d}")

        # Generate workbook
        print(f"[{'2' if use_mock else '3'}/3] Generating Excel workbook...")
        validated_data = {
            'summary': summary,
            'validation': validation
        }

        filepath = generate_workbook(validated_data, str(output_dir))

        # Verify formulas
        print(f"\n--- Verifying formulas ---")
        verification = verify_workbook_formulas(filepath)

        # Print key values (formula strings since we can't evaluate)
        print(f"\nKey formula strings from Sources & Uses:")
        for cell, formula in verification['sources_uses_formulas'].items():
            print(f"  {cell}: {formula}")

        # Print Operating Model formulas (Year 1 and Year 5 only)
        print(f"\nKey formula strings from Operating Model (Year 1 & Year 5):")
        op_formulas = verification.get('operating_model_formulas', {})
        for cell, formula in sorted(op_formulas.items()):
            # Only show columns C (Year 1) and G (Year 5)
            col = cell[0]
            if col in ['C', 'G']:
                print(f"  {cell}: {formula}")

        # Print named ranges being used
        print(f"\nNamed ranges used:")
        key_names = ['EntryRevenue', 'EntryEBITDA', 'EntryMultiple', 'LeverageMultiple',
                     'TransactionFeePct', 'TotalUses', 'NewDebtRaised',
                     'SponsorEquity', 'TotalSources', 'BalanceCheck', 'ExitEBITDA',
                     'RevenueGrowthRate', 'EBITDAMargin', 'DAPct', 'CapExPct', 'NWCPct', 'TaxRate']
        for name in key_names:
            if name in verification['named_ranges']:
                print(f"  {name}: {verification['named_ranges'][name]}")

        # Compute and print expected values
        print(f"\n--- Sources & Uses Computed Values ---")
        computed = evaluate_model_values(summary, validation)
        print_computed_values(computed)

        # Compute and print Operating Model projections
        print(f"\n--- Operating Model Computed Values ---")
        op_model = evaluate_operating_model(summary, validation, exit_year=5)
        print_operating_model(op_model)

    print(f"\n{'='*70}")
    print("TEST COMPLETE")
    print("="*70)
    print(f"\nGenerated files are in: {output_dir}")
    print("\nNOTE: openpyxl writes formulas as text. To see computed values,")
    print("open the files in Excel or a compatible spreadsheet application.")
    print("The Balance Check cell should show $0 if formulas are correct.")
